import os
import sys
import json
import sqlite3
import threading
import subprocess

# --- Auto-dependency installer ---
def install_dependencies():
    # Added pymupdf and matplotlib for PDF page rendering and plotting
    required = {"customtkinter", "pandas", "openpyxl", "google-generativeai", "pillow", "pymupdf", "matplotlib"}
    try:
        import pkg_resources
        installed = {pkg.key for pkg in pkg_resources.working_set}
    except Exception:
        try:
            import importlib.metadata
            installed = {dist.metadata['Name'].lower() for dist in importlib.metadata.distributions()}
        except Exception:
            installed = set()
            
    missing = required - installed
    if missing:
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        ans = messagebox.askyesno(
            "Install Dependencies",
            f"The following required libraries are missing: {', '.join(missing)}.\n\nWould you like the app to install them automatically now?"
        )
        if ans:
            progress = tk.Toplevel()
            progress.title("Installing...")
            progress.geometry("350x120")
            progress.resizable(False, False)
            # Center the window
            screen_width = progress.winfo_screenwidth()
            screen_height = progress.winfo_screenheight()
            x = (screen_width / 2) - (350 / 2)
            y = (screen_height / 2) - (120 / 2)
            progress.geometry(f"+{int(x)}+{int(y)}")
            
            lbl = tk.Label(progress, text="Installing dependencies, please wait...\n(This might take a minute)", pady=20, font=("Segoe UI", 10))
            lbl.pack()
            progress.update()
            
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
                messagebox.showinfo("Success", "All dependencies installed successfully! Starting the app...")
                progress.destroy()
                root.destroy()
                # Restart the script
                os.execv(sys.executable, [sys.executable] + sys.argv)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to install dependencies: {e}\nPlease run: pip install customtkinter pandas openpyxl google-generativeai pillow pymupdf matplotlib")
                sys.exit(1)
        else:
            sys.exit(0)

# Check and install dependencies
install_dependencies()

# --- Main App Imports ---
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from PIL import Image
from google import genai
from google.genai import types
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import webbrowser

# --- Configurations ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")
DB_FILE = os.path.join(SCRIPT_DIR, "quotes.db")

# --- Database Initialization ---
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # Table for tracking successfully processed files
    c.execute("""
        CREATE TABLE IF NOT EXISTS processed_files (
            filename TEXT PRIMARY KEY,
            processed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Table for storing all quotation metrics
    c.execute("""
        CREATE TABLE IF NOT EXISTS extracted_quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            supplier TEXT,
            product TEXT,
            spec TEXT,
            color TEXT,
            elastic TEXT,
            price REAL,
            unit TEXT,
            moq TEXT,
            packing TEXT,
            term TEXT,
            lead_time TEXT
        )
    """)
    # Table for supplier contacts
    c.execute("""
        CREATE TABLE IF NOT EXISTS supplier_contacts (
            supplier TEXT PRIMARY KEY,
            contact_info TEXT,
            source_file TEXT,
            extracted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try:
        c.execute("ALTER TABLE supplier_contacts ADD COLUMN source_file TEXT")
    except sqlite3.OperationalError:
        pass
    
    # Self-healing migration for validity_date and sourcing_risk columns in quotes table
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN validity_date TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN sourcing_risk TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN attached_media TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN review_status TEXT DEFAULT 'Needs Review'")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN reviewed_by TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN reviewed_at TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN review_notes TEXT")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN supplier_master_id INTEGER")
    except sqlite3.OperationalError:
        pass
    try:
        c.execute("ALTER TABLE extracted_quotes ADD COLUMN product_master_id INTEGER")
    except sqlite3.OperationalError:
        pass

    # Table for enterprise audit trail of user-visible procurement changes
    c.execute("""
        CREATE TABLE IF NOT EXISTS quote_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            quote_id INTEGER,
            action TEXT,
            previous_status TEXT,
            new_status TEXT,
            note TEXT,
            actor TEXT DEFAULT 'Local User',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        UPDATE extracted_quotes
        SET review_status = 'Needs Review'
        WHERE review_status IS NULL OR review_status = ''
    """)

    # Enterprise master data tables for controlled supplier/product records
    c.execute("""
        CREATE TABLE IF NOT EXISTS supplier_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            legal_name TEXT,
            display_name TEXT UNIQUE,
            country TEXT,
            city TEXT,
            contact_person TEXT,
            email TEXT,
            phone TEXT,
            category TEXT,
            status TEXT DEFAULT 'Active',
            payment_terms TEXT,
            certifications TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS product_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT UNIQUE,
            category TEXT,
            standard_specs TEXT,
            packaging TEXT,
            carton_cbm REAL,
            compliance_requirements TEXT,
            target_price REAL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS master_data_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT,
            entity_id INTEGER,
            action TEXT,
            note TEXT,
            actor TEXT DEFAULT 'Local User',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # RFQ and PO workflow registers for enterprise procurement control
    c.execute("""
        CREATE TABLE IF NOT EXISTS rfq_register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfq_number TEXT UNIQUE,
            product_master_id INTEGER,
            product_name TEXT,
            target_quantity TEXT,
            price_terms TEXT,
            lead_time TEXT,
            payment_terms TEXT,
            selected_suppliers TEXT,
            status TEXT DEFAULT 'Draft',
            pdf_path TEXT,
            specs TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS po_register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT UNIQUE,
            supplier_master_id INTEGER,
            product_master_id INTEGER,
            quote_id INTEGER,
            supplier_name TEXT,
            product_name TEXT,
            quantity INTEGER,
            unit_cost REAL,
            total_value REAL,
            payment_terms TEXT,
            delivery_address TEXT,
            status TEXT DEFAULT 'Issued',
            pdf_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS workflow_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workflow_type TEXT,
            workflow_id INTEGER,
            action TEXT,
            status TEXT,
            note TEXT,
            actor TEXT DEFAULT 'Local User',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
        
    # Table for chatbot history persistence
    c.execute("""
        CREATE TABLE IF NOT EXISTS chat_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender TEXT,
            message TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Table for tracking supplier compliance & factory audits
    c.execute("""
        CREATE TABLE IF NOT EXISTS supplier_compliance (
            supplier TEXT PRIMARY KEY,
            has_ce INTEGER,
            has_fda INTEGER,
            has_iso INTEGER,
            has_bsci INTEGER,
            has_sgs INTEGER,
            audit_score REAL,
            defect_rate REAL
        )
    """)
    # Table for operational incidents & defects log
    c.execute("""
        CREATE TABLE IF NOT EXISTS supplier_incidents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier TEXT,
            incident_type TEXT,
            description TEXT,
            severity TEXT,
            logged_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Table for supplier quote historical price trend tracking
    c.execute("""
        CREATE TABLE IF NOT EXISTS price_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier TEXT,
            product TEXT,
            price REAL,
            log_date TEXT DEFAULT (date('now'))
        )
    """)
    conn.commit()
    
    # Auto-populate contacts from existing quotes
    c.execute("SELECT DISTINCT supplier FROM extracted_quotes WHERE supplier IS NOT NULL AND supplier != 'Unknown'")
    suppliers = [row[0] for row in c.fetchall()]
    for s in suppliers:
        c.execute("SELECT filename FROM extracted_quotes WHERE supplier = ? LIMIT 1", (s,))
        q_row = c.fetchone()
        filename = q_row[0] if q_row else "N/A"
        
        c.execute("SELECT COUNT(*) FROM supplier_contacts WHERE supplier = ?", (s,))
        if c.fetchone()[0] == 0:
            c.execute("INSERT INTO supplier_contacts (supplier, contact_info, source_file) VALUES (?, ?, ?)", 
                      (s, "Contact info placeholder. Click '✏ Edit Info' to fill manually, or process a new quote from this supplier to auto-extract details.", filename))
        else:
            c.execute("UPDATE supplier_contacts SET source_file = ? WHERE supplier = ? AND (source_file IS NULL OR source_file = '')", (filename, s))
    conn.commit()
    conn.close()

class App(ctk.CTk):
    THEME = {
        "app_bg": "#F5F7FA",
        "surface": "#FFFFFF",
        "surface_alt": "#F1F5F9",
        "surface_soft": "#F8FAFC",
        "input_bg": "#FBFCFE",
        "sidebar": "#111827",
        "sidebar_hover": "#1F2937",
        "primary": "#1E4E8C",
        "primary_hover": "#173E70",
        "success": "#15803D",
        "success_hover": "#116A33",
        "warning": "#B45309",
        "warning_hover": "#92400E",
        "danger": "#B91C1C",
        "danger_hover": "#991B1B",
        "text": "#111827",
        "muted": "#6B7280",
        "border": "#D9E2EC",
        "border_strong": "#B8C5D3",
        "table_header": "#E5EAF1",
        "table_selected": "#D7E7F7",
        "shadow": "#E8EDF3",
        "success_soft": "#ECFDF3",
        "success_border": "#BBF7D0",
        "warning_soft": "#FFF7ED",
        "warning_border": "#FED7AA",
        "danger_soft": "#FEF2F2",
        "danger_border": "#FECACA",
        "info_soft": "#EFF6FF",
        "info_border": "#BFDBFE",
    }

    def apply_app_theme(self):
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.configure(fg_color=self.THEME["app_bg"])

    def make_button(self, parent, text, command=None, variant="primary", **kwargs):
        palette = {
            "primary": (self.THEME["primary"], self.THEME["primary_hover"], "white"),
            "secondary": (self.THEME["surface_alt"], "#E2E8F0", self.THEME["text"]),
            "success": (self.THEME["success"], self.THEME["success_hover"], "white"),
            "warning": (self.THEME["warning"], self.THEME["warning_hover"], "white"),
            "danger": (self.THEME["danger"], self.THEME["danger_hover"], "white"),
            "ghost": ("transparent", self.THEME["sidebar_hover"], "#D1D5DB"),
        }
        fg, hover, text_color = palette.get(variant, palette["primary"])
        defaults = {
            "fg_color": fg,
            "hover_color": hover,
            "text_color": text_color,
            "corner_radius": 6,
            "height": 32,
            "font": ctk.CTkFont(size=12, weight="bold"),
        }
        defaults.update(kwargs)
        return ctk.CTkButton(parent, text=text, command=command, **defaults)

    def style_card(self, widget, variant="surface"):
        variants = {
            "surface": (self.THEME["surface"], self.THEME["border"]),
            "soft": (self.THEME["surface_soft"], self.THEME["border"]),
            "info": (self.THEME["info_soft"], self.THEME["info_border"]),
            "success": (self.THEME["success_soft"], self.THEME["success_border"]),
            "warning": (self.THEME["warning_soft"], self.THEME["warning_border"]),
            "danger": (self.THEME["danger_soft"], self.THEME["danger_border"]),
        }
        fg, border = variants.get(variant, variants["surface"])
        try:
            widget.configure(fg_color=fg, border_color=border, border_width=1, corner_radius=8)
        except Exception:
            pass
        return widget

    def style_text_output(self, widget):
        try:
            widget.configure(
                fg_color=self.THEME["input_bg"],
                text_color=self.THEME["text"],
                border_color=self.THEME["border"],
                border_width=1,
                corner_radius=8,
                font=("Segoe UI", 11),
            )
        except Exception:
            pass
        return widget

    def make_empty_state(self, parent, title, detail=""):
        box = ctk.CTkFrame(parent, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        box.pack(fill="x", padx=10, pady=10)
        ctk.CTkLabel(box, text=title, font=ctk.CTkFont(size=13, weight="bold"), text_color=self.THEME["text"]).pack(anchor="w", padx=14, pady=(12, 2))
        if detail:
            ctk.CTkLabel(box, text=detail, font=ctk.CTkFont(size=11), text_color=self.THEME["muted"], wraplength=520, justify="left").pack(anchor="w", padx=14, pady=(0, 12))
        return box

    def make_tabview(self, parent):
        return ctk.CTkTabview(
            parent,
            fg_color=self.THEME["surface"],
            border_color=self.THEME["border"],
            border_width=1,
            corner_radius=8,
            segmented_button_fg_color=self.THEME["surface_alt"],
            segmented_button_selected_color=self.THEME["primary"],
            segmented_button_selected_hover_color=self.THEME["primary_hover"],
            segmented_button_unselected_color=self.THEME["surface_alt"],
            segmented_button_unselected_hover_color="#DDE6F0",
            text_color=self.THEME["text"],
            segmented_button_font=ctk.CTkFont(size=12),
        )

    def style_workspace_table(self, style_name="Treeview"):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            style_name,
            background=self.THEME["surface"],
            foreground=self.THEME["text"],
            rowheight=28,
            fieldbackground=self.THEME["surface"],
            borderwidth=0,
            font=("Segoe UI", 9),
        )
        style.map(
            style_name,
            background=[("selected", self.THEME["table_selected"])],
            foreground=[("selected", self.THEME["text"])],
        )
        style.configure(
            f"{style_name}.Heading",
            background=self.THEME["table_header"],
            foreground=self.THEME["text"],
            borderwidth=0,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
        )

    def _is_inside_sidebar(self, widget):
        parent = widget
        while parent is not None:
            if parent is getattr(self, "sidebar_frame", None):
                return True
            try:
                parent = parent.master
            except Exception:
                return False
        return False

    def apply_legacy_light_polish(self, root=None):
        root = root or self
        dark_surfaces = {"#2b2b2b", "#2c2c2c", "#3c3c3c", "#3a3a3a"}
        default_grey_surfaces = {
            "gray86",
            "grey86",
            "gray92",
            "grey92",
            "gray90",
            "grey90",
            "gray89",
            "grey89",
            "gray88",
            "grey88",
            "gray85",
            "grey85",
            "gray82",
            "grey82",
            "gray80",
            "grey80",
            "#dbdbdb",
            "#d9d9d9",
            "#d3d3d3",
            "#e5e5e5",
            "#ebebeb",
            "#f0f0f0",
            "#f2f2f2",
        }
        danger_surfaces = {"#3c2424", "#4d1e1e"}

        def normalize_color(value):
            if isinstance(value, (list, tuple)) and value:
                return str(value[0]).lower()
            return str(value).lower()

        def safe_configure(widget, **kwargs):
            for key, value in kwargs.items():
                try:
                    widget.configure(**{key: value})
                except Exception:
                    pass

        def visit(widget):
            if self._is_inside_sidebar(widget):
                return

            try:
                fg = normalize_color(widget.cget("fg_color"))
                if fg in dark_surfaces:
                    safe_configure(
                        widget,
                        fg_color=self.THEME["surface"],
                        border_color=self.THEME["border"],
                        border_width=1,
                        corner_radius=8,
                    )
                elif fg in default_grey_surfaces and isinstance(widget, (ctk.CTkFrame, ctk.CTkScrollableFrame)):
                    fill = self.THEME["surface_soft"] if isinstance(widget, ctk.CTkScrollableFrame) else self.THEME["surface"]
                    safe_configure(widget, fg_color=fill, border_color=self.THEME["border"], border_width=1, corner_radius=8)
                elif fg in danger_surfaces:
                    safe_configure(widget, fg_color=self.THEME["danger_soft"], border_color=self.THEME["danger_border"], border_width=1)
                elif fg == "#1f538d" and isinstance(widget, ctk.CTkFrame):
                    safe_configure(widget, fg_color=self.THEME["info_soft"], border_color=self.THEME["info_border"], border_width=1)
            except Exception:
                pass

            try:
                if isinstance(widget, ctk.CTkLabel):
                    color = normalize_color(widget.cget("text_color"))
                    if color in {"grey", "gray", "lightgrey", "lightgray", "#cccccc", "#808080"}:
                        widget.configure(text_color=self.THEME["muted"])
                    elif color == "white":
                        widget.configure(text_color=self.THEME["text"])
            except Exception:
                pass

            try:
                if isinstance(widget, ctk.CTkTextbox):
                    fg = normalize_color(widget.cget("fg_color"))
                    if fg in dark_surfaces or fg in default_grey_surfaces:
                        safe_configure(
                            widget,
                            fg_color=self.THEME["input_bg"],
                            text_color=self.THEME["text"],
                            border_color=self.THEME["border"],
                            border_width=1,
                            corner_radius=8,
                        )
                elif isinstance(widget, (ctk.CTkEntry, ctk.CTkComboBox)):
                    safe_configure(
                        widget,
                        fg_color=self.THEME["input_bg"],
                        border_color=self.THEME["border_strong"],
                        button_color="#AAB4C0",
                        button_hover_color="#94A3B8",
                        text_color=self.THEME["text"],
                    )
                elif isinstance(widget, ctk.CTkSlider):
                    safe_configure(
                        widget,
                        button_color="#3B82C4",
                        button_hover_color=self.THEME["primary"],
                        progress_color="#3B82C4",
                    )
                elif isinstance(widget, ctk.CTkCheckBox):
                    safe_configure(
                        widget,
                        fg_color="#3B82C4",
                        hover_color=self.THEME["primary_hover"],
                        border_color=self.THEME["border_strong"],
                        text_color=self.THEME["text"],
                    )
            except Exception:
                pass

            for child in widget.winfo_children():
                visit(child)

        visit(root)

    def generate_with_fallback(self, content_list, prompt, json_response=True):
        if hasattr(self, 'api_provider') and self.api_provider == "Custom OpenAI/Luna":
            import urllib.request
            import json
            import base64
            try:
                url = f"{self.custom_base_url.rstrip('/')}/chat/completions"
                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {self.api_key}"
                }
                
                content_payload = []
                for item in content_list:
                    if isinstance(item, dict) and "data" in item:
                        mime = item.get("mime_type", "image/jpeg")
                        b64_data = base64.b64encode(item["data"]).decode('utf-8')
                        content_payload.append({
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime};base64,{b64_data}"
                            }
                        })
                    else:
                        content_payload.append({
                            "type": "text",
                            "text": str(item)
                        })
                        
                if prompt:
                    content_payload.append({
                        "type": "text",
                        "text": prompt
                    })
                    
                payload = {
                    "model": self.custom_model,
                    "messages": [{"role": "user", "content": content_payload}]
                }
                if json_response:
                    payload["response_format"] = {"type": "json_object"}
                    
                req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), headers=headers, method="POST")
                with urllib.request.urlopen(req, timeout=30) as response:
                    res_data = json.loads(response.read().decode())
                    return res_data["choices"][0]["message"]["content"]
            except Exception as e:
                raise Exception(f"Custom OpenAI/Luna API failed: {e}")

        models_to_try = [
            'gemini-3.1-flash-lite',
            'gemini-3.5-flash-lite',
            'gemini-3.6-flash',
            'gemini-2.0-flash-lite',
            'gemini-2.0-flash'
        ]
        
        if hasattr(self, 'last_working_model') and self.last_working_model in models_to_try:
            models_to_try.remove(self.last_working_model)
            models_to_try.insert(0, self.last_working_model)
            
        last_error = None
        for model_name in models_to_try:
            for attempt in range(3):
                try:
                    client = genai.Client(api_key=self.api_key)
                    contents = []
                    for item in content_list:
                        if isinstance(item, dict) and "data" in item:
                            part = types.Part.from_bytes(
                                data=item["data"],
                                mime_type=item["mime_type"]
                            )
                            contents.append(part)
                        else:
                            contents.append(item)
                    
                    contents.append(prompt)
                    
                    config_params = {}
                    if json_response:
                        config_params["response_mime_type"] = "application/json"
                    
                    config = types.GenerateContentConfig(**config_params)
                    
                    response = client.models.generate_content(
                        model=model_name,
                        contents=contents,
                        config=config
                    )
                    
                    # Successfully generated content! Save as last working model
                    self.last_working_model = model_name
                    return response.text
                except Exception as e:
                    last_error = e
                    err_msg = str(e).lower()
                    
                    # Check for quota or credit depletion error
                    if "resource_exhausted" in err_msg or "429" in err_msg or "quota" in err_msg or "credits are depleted" in err_msg or "unsupported location" in err_msg or "unsupported region" in err_msg or "unsupported api key" in err_msg:
                        # If the quota limit is 0 (blocked on this model), don't retry, just proceed to next model!
                        if "limit: 0" in err_msg:
                            break
                        
                        import time
                        print(f"Model {model_name} rate limit (429) hit. Sleeping 10s before retry (Attempt {attempt+1}/3)...")
                        time.sleep(10.0)
                        continue
                    else:
                        print(f"Model {model_name} failed: {e}")
                        break
        raise last_error

    def __init__(self):
        super().__init__()

        self.title("ProcureAI Enterprise")
        self.geometry("1440x820")
        self.minsize(1180, 720)
        self.apply_app_theme()

        self.api_key = ""
        self.api_provider = "Google Gemini"
        self.custom_base_url = "https://api.openai.com/v1"
        self.custom_model = "gpt-5.6-luna"
        self.selected_folder = ""
        self.files_list = []
        self.extracted_data = [] # List of dicts
        self.quote_id_counter = 1
        self.is_extracting = False
        self.spinner_idx = 0
        self._is_closing = False
        
        self.exchange_rates = {
            "USD": 1.0,
            "CNY": 7.25,
            "EUR": 0.92
        }
        threading.Thread(target=self.fetch_live_exchange_rates, daemon=True).start()
        self.current_preview_path = ""

        # Chat popup loading variables
        self.chat_is_extracting = False
        self.chat_spinner_idx = 0

        # Initialize database tables
        init_db()        # Configure grid layout: Left Sidebar (0), Middle Workspace (1, weight 1), Right Preview (2)
        self.grid_columnconfigure(0, weight=0, minsize=240)
        self.grid_columnconfigure(1, weight=3)
        self.grid_columnconfigure(2, weight=0, minsize=360)
        self.grid_rowconfigure(0, weight=1)

        # --- LEFT PANEL: Sidebar Navigation Menu ---
        self.sidebar_frame = ctk.CTkFrame(self, width=240, corner_radius=0, fg_color=self.THEME["sidebar"])
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        
        self.logo_lbl = ctk.CTkLabel(self.sidebar_frame, text="ProcureAI", font=ctk.CTkFont(size=22, weight="bold"), text_color="white")
        self.logo_lbl.pack(pady=(24, 2), padx=18, anchor="w")
        self.logo_subtitle = ctk.CTkLabel(self.sidebar_frame, text="Sourcing command center", font=ctk.CTkFont(size=11), text_color="#9CA3AF")
        self.logo_subtitle.pack(pady=(0, 22), padx=18, anchor="w")

        # Container for navigation list
        self.nav_container = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        self.nav_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Navigation buttons mapping
        self.sidebar_buttons = {}
        nav_items = [
            ("📊 Sourcing & Analysis", "Sourcing Analysis"),
            ("🏆 Scorecard & Compliance", "Scorecard Compliance"),
            ("📦 Logistics & Costing", "Logistics Costing"),
            ("📝 RFQs & Outreach", "RFQs Outreach"),
            ("🔍 Customs & AI Search", "Customs AI Search"),
            ("⚙️ Settings & System", "Settings Directory")
        ]

        nav_items = [
            ("Dashboard", "Dashboard"),
            ("Quotes", "Sourcing Analysis"),
            ("Suppliers & Settings", "Settings Directory"),
            ("Costing", "Logistics Costing"),
            ("RFQ & Negotiation", "RFQs Outreach"),
            ("Compliance", "Scorecard Compliance"),
            ("Trade & Customs", "Customs AI Search")
        ]

        for label, page_key in nav_items:
            btn = self.make_button(
                self.nav_container, 
                text=label, 
                anchor="w", 
                variant="ghost",
                command=lambda pk=page_key: self.show_page(pk),
                height=38
            )
            btn.pack(fill="x", pady=3, padx=0)
            self.sidebar_buttons[page_key] = btn

        # --- MIDDLE PANEL: Active Page Container ---
        self.sidebar_visible = True
        self.sourcing_files_visible = True
        self.document_preview_visible = True

        self.right_frame = ctk.CTkFrame(self, corner_radius=0, fg_color=self.THEME["app_bg"])
        self.right_frame.grid(row=0, column=1, sticky="nsew", padx=14, pady=14)
        self.right_frame.grid_columnconfigure(0, weight=1)
        self.right_frame.grid_rowconfigure(0, weight=0) # Control header row
        self.right_frame.grid_rowconfigure(1, weight=1) # Page content row

        # Control header bar
        self.top_control_bar = ctk.CTkFrame(self.right_frame, height=42, fg_color="transparent")
        self.top_control_bar.grid(row=0, column=0, sticky="ew", padx=2, pady=(0, 8))

        self.btn_toggle_nav = ctk.CTkButton(self.top_control_bar, text="☰ Hide Navigation", fg_color="#1f538d", hover_color="#153e6b", font=ctk.CTkFont(size=11, weight="bold"), width=130, height=28, command=self.toggle_navigation_sidebar)
        self.btn_toggle_nav.pack(side="left", padx=5)
        self.btn_toggle_nav.configure(text="Hide Navigation", fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], height=30)

        self.btn_toggle_files = ctk.CTkButton(self.top_control_bar, text="📁 Hide Source Files", fg_color="#1f538d", hover_color="#153e6b", font=ctk.CTkFont(size=11, weight="bold"), width=140, height=28, command=self.toggle_sourcing_files)
        self.btn_toggle_files.configure(text="Hide Source Files", fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], height=30)
        # Will be packed dynamically inside show_page()

        self.btn_toggle_preview = ctk.CTkButton(self.top_control_bar, text="📄 Hide Preview", fg_color="#1f538d", hover_color="#153e6b", font=ctk.CTkFont(size=11, weight="bold"), width=110, height=28, command=self.toggle_document_preview)
        self.btn_toggle_preview.configure(text="Hide Preview", fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], height=30)
        # Will be packed dynamically inside show_page()

        self.pages_container = ctk.CTkFrame(self.right_frame, fg_color="transparent")
        self.pages_container.grid(row=1, column=0, sticky="nsew")
        self.pages_container.grid_columnconfigure(0, weight=1)
        self.pages_container.grid_rowconfigure(0, weight=1)

        self.pages = {}
        for name in ["Dashboard", "Sourcing Analysis", "Scorecard Compliance", "Logistics Costing", "RFQs Outreach", "Customs AI Search", "Settings Directory"]:
            self.pages[name] = ctk.CTkFrame(self.pages_container, fg_color="transparent")
            self.pages[name].grid_columnconfigure(0, weight=1)
            self.pages[name].grid_rowconfigure(0, weight=1)

        self.setup_dashboard_page()

        # ---------------------------------------------
        # 1. Workspace: Sourcing & Analysis
        # ---------------------------------------------
        self.sourcing_tabview = self.make_tabview(self.pages["Sourcing Analysis"])
        self.sourcing_tabview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        tab_comp = self.sourcing_tabview.add("📊 Quotes Comparison")
        self.tab_comp_ref = tab_comp
        self.matrix_tab = self.sourcing_tabview.add("🧮 Sourcing Matrix")
        tab_charts = self.sourcing_tabview.add("📈 Visual Charts")
        tab_insights = self.sourcing_tabview.add("💡 AI Sourcing Insights")
        tab_hedge = self.sourcing_tabview.add("💵 Currency Hedging")
        tab_history = self.sourcing_tabview.add("📈 Price History")

        # Split frame layout inside Quotes Comparison tab
        tab_comp.grid_columnconfigure(0, weight=0, minsize=260)
        tab_comp.grid_columnconfigure(1, weight=1)
        tab_comp.grid_rowconfigure(0, weight=1)

        self.sourcing_files_subframe = ctk.CTkFrame(tab_comp, width=260, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        self.sourcing_files_subframe.grid(row=0, column=0, sticky="nsew", padx=(5, 10), pady=10)
        
        self.sourcing_grid_subframe = ctk.CTkFrame(tab_comp, fg_color="transparent")
        self.sourcing_grid_subframe.grid(row=0, column=1, sticky="nsew", padx=(5, 10), pady=10)
        self.sourcing_grid_subframe.grid_columnconfigure(0, weight=1)
        self.sourcing_grid_subframe.grid_rowconfigure(2, weight=1)

        # Folder selection / Direct Path Box
        self.folder_frame = ctk.CTkFrame(self.sourcing_files_subframe, fg_color="transparent")
        self.folder_frame.pack(pady=15, padx=10, fill="x")

        self.load_btn_frame = ctk.CTkFrame(self.folder_frame, fg_color="transparent")
        self.load_btn_frame.pack(fill="x", pady=(0, 5))

        self.btn_select_folder = ctk.CTkButton(self.load_btn_frame, text="Select Folder", command=self.select_folder, width=115)
        self.btn_select_folder.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], corner_radius=6)
        self.btn_select_folder.pack(side="left", fill="x", expand=True, padx=(0, 2))

        self.btn_select_files = ctk.CTkButton(self.load_btn_frame, text="+ Add Files", fg_color="#1f538d", hover_color="#153e6b", command=self.select_files, width=115)
        self.btn_select_files.configure(fg_color=self.THEME["primary"], hover_color=self.THEME["primary_hover"], text_color="white", corner_radius=6)
        self.btn_select_files.pack(side="right", fill="x", expand=True, padx=(2, 0))

        self.folder_entry = ctk.CTkEntry(self.folder_frame, placeholder_text="Or paste folder path here...", height=34, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border_strong"])
        self.folder_entry.pack(pady=2, fill="x")
        self.folder_entry.bind("<Return>", lambda event: self.on_path_entered())
        self.folder_entry.bind("<FocusOut>", lambda event: self.on_path_entered())

        # Files Queue List Views (Unsynced and Synced separated)
        self.unsynced_lbl = ctk.CTkLabel(self.sourcing_files_subframe, text="⏳ Unsynced Queue:", anchor="w", font=ctk.CTkFont(weight="bold"))
        self.unsynced_lbl.configure(text="Pending Queue", font=ctk.CTkFont(size=13, weight="bold"), text_color=self.THEME["text"])
        self.unsynced_lbl.pack(fill="x", padx=10, pady=(10, 3))

        self.files_box_unsynced = tk.Listbox(self.sourcing_files_subframe, bg="#F8FAFC", fg="#111827", borderwidth=0, highlightthickness=1, highlightbackground="#D9E2EC", selectbackground="#D7E7F7", selectforeground="#111827", font=("Segoe UI", 10), height=8)
        self.files_box_unsynced.pack(fill="both", expand=True, padx=10, pady=(2, 5))

        self.synced_lbl = ctk.CTkLabel(self.sourcing_files_subframe, text="✅ Synced Quotes:", anchor="w", font=ctk.CTkFont(weight="bold"))
        self.synced_lbl.configure(text="Synced Quotes", font=ctk.CTkFont(size=13, weight="bold"), text_color=self.THEME["text"])
        self.synced_lbl.pack(fill="x", padx=10, pady=(10, 3))

        self.files_box_synced = tk.Listbox(self.sourcing_files_subframe, bg="#F8FAFC", fg="#111827", borderwidth=0, highlightthickness=1, highlightbackground="#D9E2EC", selectbackground="#D7E7F7", selectforeground="#111827", font=("Segoe UI", 10), height=8)
        self.files_box_synced.pack(fill="both", expand=True, padx=10, pady=(2, 5))

        # Control Buttons
        self.btn_start = ctk.CTkButton(self.sourcing_files_subframe, text="Start Extraction", state="disabled", command=self.start_extraction_thread)
        self.btn_start.configure(fg_color=self.THEME["success"], hover_color=self.THEME["success_hover"], text_color="white", corner_radius=6)
        self.btn_start.pack(fill="x", padx=10, pady=5)

        self.btn_organize = ctk.CTkButton(self.sourcing_files_subframe, text="📁 Organize Files", fg_color="#6e4513", hover_color="#52320b", command=self.start_file_organizer_thread)
        self.btn_organize.configure(text="Organize Files", fg_color=self.THEME["warning"], hover_color=self.THEME["warning_hover"], text_color="white", corner_radius=6)
        self.btn_organize.pack(fill="x", padx=10, pady=5)

        self.progress_bar = ctk.CTkProgressBar(self.sourcing_files_subframe)
        self.progress_bar.pack(fill="x", padx=10, pady=5)
        self.progress_bar.set(0)

        # ---------------------------------------------
        # 2. Workspace: Scorecard & Compliance
        # ---------------------------------------------
        self.scorecard_tabview = self.make_tabview(self.pages["Scorecard Compliance"])
        self.scorecard_tabview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        tab_scorecard = self.scorecard_tabview.add("🏆 Supplier Scorecard")
        tab_qc = self.scorecard_tabview.add("🏢 Factory Audit & QC")
        tab_incidents = self.scorecard_tabview.add("⚠️ Defect Log")

        # ---------------------------------------------
        # 3. Workspace: Logistics & Costing
        # ---------------------------------------------
        self.logistics_tabview = self.make_tabview(self.pages["Logistics Costing"])
        self.logistics_tabview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        tab_landed = self.logistics_tabview.add("📦 Landed Cost Simulator")
        tab_opt = self.logistics_tabview.add("🎯 Purchase Optimizer")
        tab_packing = self.logistics_tabview.add("🚢 Container Packing")
        tab_timeline = self.logistics_tabview.add("📅 Gantt Timeline")
        tab_prof = self.logistics_tabview.add("💰 Profit Simulator")
        tab_po = self.logistics_tabview.add("📄 PO Generator")

        tab_po_register = self.logistics_tabview.add("PO Register")

        # ---------------------------------------------
        # 4. Workspace: RFQs & Outreach
        # ---------------------------------------------
        self.rfqs_tabview = self.make_tabview(self.pages["RFQs Outreach"])
        self.rfqs_tabview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        tab_rfq = self.rfqs_tabview.add("📝 RFQ Generator")
        tab_neg = self.rfqs_tabview.add("💬 AI Negotiation")

        tab_rfq_register = self.rfqs_tabview.add("RFQ Register")

        # ---------------------------------------------
        # 5. Workspace: Customs & AI Search
        # ---------------------------------------------
        self.search_tabview = self.make_tabview(self.pages["Customs AI Search"])
        self.search_tabview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        tab_search = self.search_tabview.add("🔍 AI Visual Search")
        tab_uae = self.search_tabview.add("🇦🇪 UAE Customs & HS Code")
        tab_barriers = self.search_tabview.add("🌍 Global Trade Barriers")

        # ---------------------------------------------
        # 6. Workspace: Settings & System
        # ---------------------------------------------
        self.settings_tabview = self.make_tabview(self.pages["Settings Directory"])
        self.settings_tabview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        
        tab_settings = self.settings_tabview.add("⚙️ Settings & API")
        tab_dir = self.settings_tabview.add("📇 Supplier Directory")
        tab_master = self.settings_tabview.add("Master Data")
        
        settings_card = ctk.CTkFrame(tab_settings, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        settings_card.pack(pady=20, padx=20, fill="both", expand=True)
        
        ctk.CTkLabel(settings_card, text="⚙️ Sourcing API & Environment Settings", font=ctk.CTkFont(size=18, weight="bold"), anchor="w").pack(pady=(20, 15), padx=25, fill="x")
        
        ctk.CTkLabel(settings_card, text="Configure the AI provider used for quote extraction, recommendations, and compliance helpers.", font=ctk.CTkFont(size=12), anchor="w", text_color=self.THEME["muted"]).pack(pady=(0, 18), padx=25, fill="x")

        # Form Container for consistent padding & alignment
        form_frame = ctk.CTkFrame(settings_card, fg_color="transparent")
        form_frame.pack(padx=25, fill="x")

        # Provider selector
        self.provider_lbl = ctk.CTkLabel(form_frame, text="API Provider:", anchor="w")
        self.provider_lbl.configure(text_color=self.THEME["text"], font=ctk.CTkFont(size=12, weight="bold"))
        self.provider_lbl.pack(fill="x", pady=(5, 2), anchor="w")
        
        self.provider_cb = ctk.CTkComboBox(form_frame, values=["Google Gemini", "Custom OpenAI/Luna"], command=self.on_provider_changed, width=400)
        self.provider_cb.pack(pady=2, anchor="w")
        self.provider_cb.set("Google Gemini")

        # API Key
        self.api_lbl = ctk.CTkLabel(form_frame, text="Gemini API Key:", anchor="w")
        self.api_lbl.configure(text_color=self.THEME["text"], font=ctk.CTkFont(size=12, weight="bold"))
        self.api_lbl.pack(fill="x", pady=(5, 2), anchor="w")
        
        self.api_entry = ctk.CTkEntry(form_frame, placeholder_text="AIzaSy...", show="*", width=400)
        self.api_entry.configure(height=34, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border_strong"])
        self.api_entry.pack(pady=2, anchor="w")
        if self.api_key:
            self.api_entry.insert(0, self.api_key)

        # Custom API inputs
        self.custom_api_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        
        self.base_url_lbl = ctk.CTkLabel(self.custom_api_frame, text="Base URL:", anchor="w")
        self.base_url_lbl.configure(text_color=self.THEME["text"], font=ctk.CTkFont(size=12, weight="bold"))
        self.base_url_lbl.pack(fill="x", pady=(5, 2), anchor="w")
        self.base_url_entry = ctk.CTkEntry(self.custom_api_frame, placeholder_text="https://api.openai.com/v1", width=400)
        self.base_url_entry.configure(height=34, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border_strong"])
        self.base_url_entry.pack(pady=2, anchor="w")
        self.base_url_entry.insert(0, "https://api.openai.com/v1")
        
        self.model_lbl = ctk.CTkLabel(self.custom_api_frame, text="Model Name:", anchor="w")
        self.model_lbl.configure(text_color=self.THEME["text"], font=ctk.CTkFont(size=12, weight="bold"))
        self.model_lbl.pack(fill="x", pady=(5, 2), anchor="w")
        self.model_entry = ctk.CTkEntry(self.custom_api_frame, placeholder_text="gpt-5.6-luna", width=400)
        self.model_entry.configure(height=34, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border_strong"])
        self.model_entry.pack(pady=2, anchor="w")
        self.model_entry.insert(0, "gpt-5.6-luna")

        self.btn_save_api = ctk.CTkButton(form_frame, text="Save & Test Key", command=self.save_and_test_key, width=200)
        self.btn_save_api.configure(fg_color=self.THEME["primary"], hover_color=self.THEME["primary_hover"], corner_radius=6, height=36)
        self.btn_save_api.pack(pady=15, anchor="w")

        # Safety Zone card inside Settings page
        safety_card = ctk.CTkFrame(settings_card, fg_color="#FEF2F2", border_color="#FECACA", border_width=1, corner_radius=8)
        safety_card.pack(pady=20, padx=25, fill="x")
        
        ctk.CTkLabel(safety_card, text="🚨 Danger Zone", font=ctk.CTkFont(weight="bold"), text_color="#ff8888", anchor="w").pack(anchor="w", padx=15, pady=(10, 5), fill="x")
        
        self.btn_clear_all = ctk.CTkButton(safety_card, text="🧹 Clear All Data", fg_color="#a83232", hover_color="#8c2626", command=self.clear_all_data)
        self.btn_clear_all.configure(text="Clear All Data", fg_color=self.THEME["danger"], hover_color=self.THEME["danger_hover"], corner_radius=6, height=34)
        self.btn_clear_all.pack(side="left", padx=15, pady=(5, 15))
        
        ctk.CTkLabel(safety_card, text="Warning: Clicking this button permanently purges all quotation records, scorecard data, compliance audits, and attachments from the local database.", font=ctk.CTkFont(size=11), text_color="grey", anchor="w").pack(side="left", padx=10, pady=(5, 15), fill="x", expand=True)

        # --- Re-grid Sourcing Grid elements inside subframe ---
        self.table_ctrl_frame = ctk.CTkFrame(self.sourcing_grid_subframe, fg_color="transparent")
        self.table_ctrl_frame.grid(row=0, column=0, sticky="ew", padx=12, pady=(14, 8))

        self.table_lbl = ctk.CTkLabel(self.table_ctrl_frame, text="Extracted Quotes Comparison", font=ctk.CTkFont(size=18, weight="bold"))
        self.table_lbl.configure(font=ctk.CTkFont(size=20, weight="bold"), text_color=self.THEME["text"])
        self.table_lbl.pack(side="left")

        self.btn_add_row = ctk.CTkButton(self.table_ctrl_frame, text="+ Add Row", width=90, command=self.add_empty_row)
        self.btn_add_row.configure(fg_color=self.THEME["primary"], hover_color=self.THEME["primary_hover"], text_color="white", corner_radius=6, height=34)
        self.btn_add_row.pack(side="right", padx=5)

        self.btn_delete_row = ctk.CTkButton(self.table_ctrl_frame, text="- Delete Row", width=90, fg_color="#a83232", hover_color="#8c2626", command=self.delete_selected_row)
        self.btn_delete_row.configure(fg_color=self.THEME["danger"], hover_color=self.THEME["danger_hover"], text_color="white", corner_radius=6, height=34)
        self.btn_delete_row.pack(side="right", padx=5)

        self.btn_reject_quote = self.make_button(self.table_ctrl_frame, text="Reject", width=78, command=lambda: self.set_selected_quote_status("Rejected"), variant="danger")
        self.btn_reject_quote.pack(side="right", padx=5)

        self.btn_review_quote = self.make_button(self.table_ctrl_frame, text="Review", width=78, command=lambda: self.set_selected_quote_status("Needs Review"), variant="warning")
        self.btn_review_quote.pack(side="right", padx=5)

        self.btn_approve_quote = self.make_button(self.table_ctrl_frame, text="Approve", width=82, command=lambda: self.set_selected_quote_status("Approved"), variant="success")
        self.btn_approve_quote.pack(side="right", padx=5)

        self.btn_edit_row = ctk.CTkButton(self.table_ctrl_frame, text="✏ Edit Row", width=90, command=self.edit_selected_row)
        self.btn_edit_row.configure(text="Edit Row", fg_color="#3B82C4", hover_color="#2C6DA6", text_color="white", corner_radius=6, height=34)
        self.btn_edit_row.pack(side="right", padx=5)

        self.btn_attach_media = ctk.CTkButton(self.table_ctrl_frame, text="📎 Attach Media", width=100, command=self.attach_media_to_selected)
        self.btn_attach_media.configure(text="Attach Media", fg_color="#3B82C4", hover_color="#2C6DA6", text_color="white", corner_radius=6, height=34)
        self.btn_attach_media.pack(side="right", padx=5)

        self.btn_paste_chat = ctk.CTkButton(self.table_ctrl_frame, text="📋 Paste Chat", width=90, fg_color="#1f538d", command=self.open_paste_chat_window)
        self.btn_paste_chat.configure(text="Paste Chat", fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], corner_radius=6, height=34)
        self.btn_paste_chat.pack(side="right", padx=5)

        # Row 1: Search box
        self.search_frame = ctk.CTkFrame(self.sourcing_grid_subframe, fg_color="transparent")
        self.search_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        
        self.search_entry = ctk.CTkEntry(self.search_frame, placeholder_text="🔍 Type to filter by supplier, product, color, specs, etc...", height=28)
        self.search_entry.configure(
            placeholder_text="Search supplier, product, color, specs, terms...",
            height=36,
            fg_color=self.THEME["surface_soft"],
            border_color=self.THEME["border_strong"],
            text_color=self.THEME["text"]
        )
        self.search_entry.pack(fill="x", expand=True, padx=5, pady=5)
        self.search_entry.bind("<KeyRelease>", lambda event: self.filter_table())

        # Row 2: Treeview styled table
        self.setup_table(self.sourcing_grid_subframe)

        # Row 4: AI Procurement Chatbot Panel
        self.chat_panel = ctk.CTkFrame(self.sourcing_grid_subframe, height=180, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        self.chat_panel.grid(row=4, column=0, sticky="ew", padx=12, pady=(8, 5))
        self.chat_panel.grid_propagate(False)
        self.chat_panel.grid_columnconfigure(0, weight=1)
        self.chat_panel.grid_rowconfigure(1, weight=1)

        self.chat_title = ctk.CTkLabel(self.chat_panel, text="💬 AI Procurement Assistant", font=ctk.CTkFont(size=13, weight="bold"))
        self.chat_title.configure(text="AI Procurement Assistant", font=ctk.CTkFont(size=14, weight="bold"), text_color=self.THEME["text"])
        self.chat_title.grid(row=0, column=0, columnspan=2, padx=12, pady=(8, 3), sticky="w")

        self.chat_log = ctk.CTkTextbox(self.chat_panel, wrap="word", font=("Segoe UI", 9))
        self.chat_log.configure(fg_color=self.THEME["surface_soft"], text_color=self.THEME["text"], border_width=0)
        self.chat_log.grid(row=1, column=0, columnspan=2, padx=12, pady=3, sticky="nsew")

        self.chat_input_frame = ctk.CTkFrame(self.chat_panel, fg_color="transparent")
        self.chat_input_frame.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="ew")
        self.chat_input_frame.grid_columnconfigure(0, weight=1)

        self.chat_entry = ctk.CTkEntry(self.chat_input_frame, placeholder_text="Ask AI assistant...", height=26)
        self.chat_entry.configure(height=32, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border_strong"])
        self.chat_entry.grid(row=0, column=0, sticky="ew", padx=(0, 5))
        self.chat_entry.bind("<Return>", lambda event: self.send_chat_message())

        self.btn_chat_send = ctk.CTkButton(self.chat_input_frame, text="Send", width=60, height=26, command=self.send_chat_message)
        self.btn_chat_send.configure(width=70, height=32, fg_color=self.THEME["primary"], hover_color=self.THEME["primary_hover"], corner_radius=6)
        self.btn_chat_send.grid(row=0, column=1, sticky="e")

        # Row 5: Export Buttons & Currency dropdown
        self.export_frame = ctk.CTkFrame(self.sourcing_grid_subframe, fg_color="transparent")
        self.export_frame.grid(row=5, column=0, sticky="ew", padx=12, pady=10)

        self.currency_lbl = ctk.CTkLabel(self.export_frame, text="Currency:")
        self.currency_lbl.pack(side="left", padx=(10, 5))
        
        self.currency_cb = ctk.CTkComboBox(self.export_frame, values=["USD ($)", "CNY (¥)", "EUR (€)"], command=self.change_currency, width=120)
        self.currency_cb.pack(side="left", padx=5)

        self.currency_status_lbl = ctk.CTkLabel(self.export_frame, text="Fetching live rates...", text_color="grey", font=ctk.CTkFont(size=11))
        self.currency_status_lbl.pack(side="left", padx=10)

        self.btn_export_excel = ctk.CTkButton(self.export_frame, text="Export to Excel", fg_color="#1f7d44", hover_color="#15592e", command=self.export_to_excel)
        self.btn_export_excel.configure(fg_color=self.THEME["success"], hover_color=self.THEME["success_hover"], corner_radius=6, height=34)
        self.btn_export_excel.pack(side="right", padx=5)

        self.btn_export_csv = ctk.CTkButton(self.export_frame, text="Export to CSV", command=self.export_to_csv)
        self.btn_export_csv.configure(fg_color="#3B82C4", hover_color="#2C6DA6", corner_radius=6, height=34)
        self.btn_export_csv.pack(side="right", padx=5)



        # --- PAGE 2: Supplier Directory ---
        tab_dir = self.settings_tabview.tab("📇 Supplier Directory")
        tab_dir.grid_columnconfigure(0, weight=1)
        tab_dir.grid_rowconfigure(1, weight=1)

        self.dir_header_frame = ctk.CTkFrame(tab_dir, fg_color="transparent")
        self.dir_header_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=10)

        self.dir_title = ctk.CTkLabel(self.dir_header_frame, text="Supplier Contact Directory", font=ctk.CTkFont(size=18, weight="bold"))
        self.dir_title.pack(side="left")

        self.btn_sync_dir = ctk.CTkButton(self.dir_header_frame, text="🔄 Sync from Quotes", width=140, fg_color="#1f538d", command=self.sync_suppliers_from_quotes)
        self.btn_sync_dir.pack(side="right", padx=5)

        self.btn_extract_contacts = ctk.CTkButton(self.dir_header_frame, text="🔍 Auto-Extract Contacts", width=160, fg_color="#1f7d44", hover_color="#15592e", command=self.start_contact_extraction_thread)
        self.btn_extract_contacts.pack(side="right", padx=5)

        self.directory_scroll_frame = ctk.CTkScrollableFrame(tab_dir, fg_color=self.THEME["surface_soft"])
        self.directory_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=15, pady=10)

        self.setup_master_data_tab()

        # --- PAGE 3: Visual Price Comparison Charts ---
        tab_charts = self.sourcing_tabview.tab("📈 Visual Charts")
        tab_charts.grid_columnconfigure(0, weight=1)
        tab_charts.grid_rowconfigure(1, weight=1)

        self.chart_ctrl_frame = ctk.CTkFrame(tab_charts, fg_color="transparent")
        self.chart_ctrl_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

        self.chart_title_lbl = ctk.CTkLabel(self.chart_ctrl_frame, text="Price Comparison Dashboard", font=ctk.CTkFont(size=18, weight="bold"))
        self.chart_title_lbl.pack(side="left")

        self.chart_category_cb = ctk.CTkComboBox(self.chart_ctrl_frame, values=["All", "Cap", "Apron", "Sleeve", "Shoe", "Mask"], command=lambda choice: self.draw_chart(), width=150)
        self.chart_category_cb.pack(side="right", padx=10)

        self.chart_lbl = ctk.CTkLabel(self.chart_ctrl_frame, text="Filter Category:")
        self.chart_lbl.pack(side="right", padx=5)

        self.chart_display_frame = ctk.CTkFrame(tab_charts, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        self.chart_display_frame.grid(row=1, column=0, sticky="nsew", padx=15, pady=10)

        # Setup Sourcing Insights tab
        self.setup_sourcing_insights_tab()

        # Setup Scorecard tab
        self.setup_scorecard_tab()

        # Setup Timeline tab
        self.setup_timeline_tab()

        # Setup Landed Cost Simulator tab
        self.setup_landed_cost_tab()

        # Setup Purchase Optimizer tab
        self.setup_purchase_optimizer_tab()

        # Setup RFQ Generator tab
        self.setup_rfq_generator_tab()
        self.setup_rfq_register_tab()

        # Setup Profit Simulator tab
        self.setup_profit_simulator_tab()

        # Setup Factory Audit & QC tab
        self.setup_factory_qc_tab()

        # Setup Sourcing Matrix tab
        self.setup_sourcing_matrix_tab()

        # Setup AI Visual Search tab
        self.setup_visual_search_tab()

        # Setup UAE Customs & HS Code tab
        self.setup_uae_customs_tab()

        # Setup Container Packing tab
        self.setup_container_packing_tab()

        # Setup Currency Hedging tab
        self.setup_currency_hedging_tab()

        # Setup Defect Log tab
        self.setup_defect_log_tab()

        # Setup PO Generator tab
        self.setup_po_generator_tab()
        self.setup_po_register_tab()

        # Setup Price History & Trend Tracker tab
        self.setup_price_history_tab()

        # Setup AI Negotiation tab
        self.setup_ai_negotiation_tab()

        # Setup Global Trade Barriers tab
        self.setup_global_barriers_tab()

        # --- RIGHT PANEL 2: Document Preview Sidebar ---
        self.preview_frame = ctk.CTkFrame(self, width=360, corner_radius=0, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1)
        self.preview_frame.grid(row=0, column=2, sticky="nsew", padx=(0, 14), pady=14)
        self.preview_frame.grid_columnconfigure(0, weight=1)
        self.preview_frame.grid_rowconfigure(3, weight=1)

        self.preview_title = ctk.CTkLabel(self.preview_frame, text="Document Preview", font=ctk.CTkFont(size=18, weight="bold"), text_color=self.THEME["text"])
        self.preview_title.grid(row=0, column=0, pady=15, padx=10, sticky="ew")

        self.preview_filename_lbl = ctk.CTkLabel(self.preview_frame, text="Select a row to preview", wraplength=330, text_color=self.THEME["muted"])
        self.preview_filename_lbl.grid(row=1, column=0, pady=5, padx=10, sticky="ew")

        # Horizontal Media Gallery Bar Frame
        self.preview_gallery_bar = ctk.CTkFrame(self.preview_frame, fg_color="transparent")
        self.preview_gallery_bar.grid(row=2, column=0, pady=5, padx=15, sticky="ew")

        # Container for preview media
        self.preview_display_frame = ctk.CTkFrame(self.preview_frame, fg_color=self.THEME["surface_alt"], corner_radius=8)
        self.preview_display_frame.grid(row=3, column=0, sticky="nsew", padx=15, pady=10)
        self.preview_display_frame.grid_columnconfigure(0, weight=1)
        self.preview_display_frame.grid_rowconfigure(0, weight=1)

        self.preview_image_lbl = ctk.CTkLabel(self.preview_display_frame, text="No document selected", text_color=self.THEME["muted"], font=ctk.CTkFont(size=14, weight="bold"))
        self.preview_image_lbl.pack(pady=120, padx=10)

        # Scrolling text box for text file rendering
        self.preview_text_box = ctk.CTkTextbox(self.preview_display_frame, wrap="word")

        # Metrics Overlay Frame inside preview_frame
        self.preview_metrics_frame = ctk.CTkFrame(self.preview_frame, fg_color=self.THEME["surface_alt"], corner_radius=8)
        self.preview_metrics_frame.grid(row=4, column=0, sticky="ew", padx=15, pady=5)
        self.update_preview_metrics_overlay(None)

        # System open button
        self.btn_open_external = ctk.CTkButton(self.preview_frame, text="Open File Externally", state="disabled", command=self.open_file_externally, fg_color="#3B82C4", hover_color="#2C6DA6", corner_radius=6, height=34)
        self.btn_open_external.grid(row=5, column=0, pady=10, padx=15, sticky="ew")

        # Load configurations & history
        self.load_config()
        self.load_chat_history_from_db()
        self.load_all_quotes_from_db()
        self.apply_legacy_light_polish()
        self.show_page("Dashboard")
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_dashboard_page(self):
        self.style_workspace_table("Treeview")
        page = self.pages["Dashboard"]
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(2, weight=1)

        header = ctk.CTkFrame(page, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=6, pady=(2, 12))
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Procurement Overview",
            font=ctk.CTkFont(size=26, weight="bold"),
            text_color=self.THEME["text"],
        ).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(
            header,
            text="Live sourcing status, supplier risk, and next actions from your quote database.",
            font=ctk.CTkFont(size=12),
            text_color=self.THEME["muted"],
        ).grid(row=1, column=0, sticky="w", pady=(3, 0))

        self.dashboard_cards_frame = ctk.CTkFrame(page, fg_color="transparent")
        self.dashboard_cards_frame.grid(row=1, column=0, sticky="ew", padx=0, pady=(0, 10))
        for idx in range(8):
            self.dashboard_cards_frame.grid_columnconfigure(idx, weight=1, uniform="dash_cards")

        self.dashboard_cards = {}
        card_specs = [
            ("review", "Needs Review", "0"),
            ("approved", "Approved", "0"),
            ("rfq_open", "Open RFQs", "0"),
            ("po_open", "Open POs", "0"),
            ("po_value", "Open PO Value", "$0"),
            ("quotes", "Quote Rows", "0"),
            ("suppliers", "Suppliers", "0"),
            ("expired", "Expired Quotes", "0"),
        ]
        for idx, (key, title, value) in enumerate(card_specs):
            card = ctk.CTkFrame(
                self.dashboard_cards_frame,
                fg_color=self.THEME["surface"],
                border_color=self.THEME["border"],
                border_width=1,
                corner_radius=8,
            )
            card.grid(row=0, column=idx, sticky="nsew", padx=6)
            ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=11, weight="bold"), text_color=self.THEME["muted"]).pack(anchor="w", padx=14, pady=(12, 2))
            value_lbl = ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=20, weight="bold"), text_color=self.THEME["text"])
            value_lbl.pack(anchor="w", padx=14, pady=(0, 12))
            self.dashboard_cards[key] = value_lbl

        body = ctk.CTkFrame(page, fg_color="transparent")
        body.grid(row=2, column=0, sticky="nsew", padx=0, pady=0)
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)

        left = ctk.CTkFrame(body, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        left.grid(row=0, column=0, sticky="nsew", padx=(6, 8), pady=6)
        left.grid_columnconfigure(0, weight=1)
        left.grid_rowconfigure(2, weight=1)

        left_header = ctk.CTkFrame(left, fg_color="transparent")
        left_header.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 8))
        left_header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(left_header, text="Quote Approval Queue", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w")
        self.make_button(left_header, "Open Quotes", command=self.open_dashboard_quote_in_quotes, variant="secondary", width=96).grid(row=0, column=1, padx=4)
        self.make_button(left_header, "Approve", command=lambda: self.update_dashboard_selected_quote("Approved"), variant="success", width=82).grid(row=0, column=2, padx=4)
        self.make_button(left_header, "Reject", command=lambda: self.update_dashboard_selected_quote("Rejected"), variant="danger", width=74).grid(row=0, column=3, padx=4)
        self.make_button(left_header, "Send to RFQ", command=self.send_dashboard_quote_to_rfq, variant="primary", width=92).grid(row=0, column=4, padx=4)

        ctk.CTkLabel(
            left,
            text="Review quote rows before they can feed downstream purchase orders.",
            font=ctk.CTkFont(size=11),
            text_color=self.THEME["muted"],
        ).grid(row=1, column=0, sticky="w", padx=16, pady=(0, 6))

        cols = ("id", "supplier", "product", "price", "issue")
        self.dashboard_approval_tree = ttk.Treeview(left, columns=cols, show="headings", style="Treeview")
        for col, label, width in [
            ("id", "ID", 55),
            ("supplier", "Supplier", 210),
            ("product", "Product", 150),
            ("price", "Unit Price", 90),
            ("issue", "Review Warning", 230),
        ]:
            self.dashboard_approval_tree.heading(col, text=label)
            self.dashboard_approval_tree.column(col, width=width, anchor="w")
        self.dashboard_approval_tree.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 16))
        self.dashboard_approval_tree.tag_configure("needs_review", background="#FEF3C7", foreground="#92400E")
        self.dashboard_approval_tree.bind("<Double-1>", lambda event: self.open_dashboard_quote_in_quotes())

        right = ctk.CTkFrame(body, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 6), pady=6)
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=2)
        right.grid_rowconfigure(3, weight=1)
        ctk.CTkLabel(right, text="Workflow Audit Timeline", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w", padx=16, pady=(14, 8))
        self.dashboard_timeline_box = ctk.CTkTextbox(right, wrap="word", fg_color=self.THEME["surface_alt"], text_color=self.THEME["text"], border_width=0)
        self.dashboard_timeline_box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 12))

        ctk.CTkLabel(right, text="Recommended Next Actions", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=2, column=0, sticky="w", padx=16, pady=(4, 8))
        self.dashboard_actions_box = ctk.CTkTextbox(right, wrap="word", fg_color=self.THEME["surface_alt"], text_color=self.THEME["text"], border_width=0)
        self.dashboard_actions_box.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 16))

    def update_dashboard_page(self):
        if not hasattr(self, "dashboard_cards"):
            return

        data = self.extracted_data or []
        workflow = self.get_workflow_dashboard_snapshot()
        suppliers = {self.clean_supplier_name(r.get("supplier")) for r in data if r.get("supplier") and r.get("supplier") != "Unknown"}
        decision_data = self.get_decision_quotes(data) if data else []
        numeric_prices = []
        for r in decision_data:
            try:
                price = float(r.get("price"))
                if price > 0:
                    numeric_prices.append((price, r))
            except Exception:
                pass

        import datetime
        today = datetime.date.today()
        expired_count = 0
        active_count = 0
        for r in data:
            try:
                val = datetime.datetime.strptime(str(r.get("validity_date")), "%Y-%m-%d").date()
                if val < today:
                    expired_count += 1
                else:
                    active_count += 1
            except Exception:
                pass

        high_risk = [r for r in data if "high" in str(r.get("sourcing_risk") or "").lower()]
        needs_review = [r for r in data if (r.get("review_status") or "Needs Review") == "Needs Review"]
        approved = [r for r in data if r.get("review_status") == "Approved"]

        self.dashboard_cards["quotes"].configure(text=str(len(data)))
        self.dashboard_cards["suppliers"].configure(text=str(len(suppliers)))
        self.dashboard_cards["review"].configure(text=str(len(needs_review)), text_color=self.THEME["warning"] if needs_review else self.THEME["text"])
        self.dashboard_cards["approved"].configure(text=str(len(approved)), text_color=self.THEME["success"] if approved else self.THEME["text"])
        self.dashboard_cards["expired"].configure(text=str(expired_count), text_color=self.THEME["danger"] if expired_count else self.THEME["text"])
        self.dashboard_cards["rfq_open"].configure(text=str(workflow["open_rfqs"]), text_color=self.THEME["primary"] if workflow["open_rfqs"] else self.THEME["text"])
        self.dashboard_cards["po_open"].configure(text=str(workflow["open_pos"]), text_color=self.THEME["primary"] if workflow["open_pos"] else self.THEME["text"])
        self.dashboard_cards["po_value"].configure(text=f"${workflow['open_po_value']:,.0f}", text_color=self.THEME["success"] if workflow["open_po_value"] else self.THEME["text"])

        self.dashboard_approval_tree.delete(*self.dashboard_approval_tree.get_children())
        review_rows = sorted(needs_review, key=lambda row: self.quote_review_priority(row), reverse=True)[:18]
        for r in review_rows:
            price = r.get("price")
            try:
                price_txt = f"${float(price):.5f}"
            except Exception:
                price_txt = "N/A"
            self.dashboard_approval_tree.insert(
                "",
                tk.END,
                values=(
                    r.get("id"),
                    self.clean_supplier_name(r.get("supplier")),
                    r.get("product") or "N/A",
                    price_txt,
                    self.quote_warning_summary(r),
                ),
                tags=("needs_review",),
            )

        actions = []
        if not data:
            actions.append("Select a supplier quote folder and run extraction to populate the workspace.")
        if needs_review:
            actions.append(f"Verify and approve {len(needs_review)} quote rows before using them for sourcing decisions.")
        if expired_count:
            actions.append(f"Request refreshed pricing for {expired_count} expired quote rows.")
        if high_risk:
            actions.append(f"Review payment terms and supplier exposure for {len(high_risk)} high-risk quote rows.")
        if active_count and numeric_prices:
            best_supplier = self.clean_supplier_name(min(numeric_prices, key=lambda item: item[0])[1].get("supplier"))
            actions.append(f"Use {best_supplier} as the first benchmark in negotiations based on lowest unit price.")
        if workflow["open_rfqs"]:
            actions.append(f"Follow up on {workflow['open_rfqs']} open RFQ record(s) until they are closed or cancelled.")
        if workflow["open_pos"]:
            actions.append(f"Track {workflow['open_pos']} active PO record(s) through supplier acceptance, shipment, and closure.")
        actions.append("Keep supplier contact records current before broadcasting RFQs.")

        timeline_text = "\n\n".join(workflow["timeline"]) if workflow["timeline"] else "No RFQ, PO, or quote approval events have been logged yet."
        self.dashboard_timeline_box.configure(state="normal")
        self.dashboard_timeline_box.delete("1.0", tk.END)
        self.dashboard_timeline_box.insert("1.0", timeline_text)
        self.dashboard_timeline_box.configure(state="disabled")

        self.dashboard_actions_box.configure(state="normal")
        self.dashboard_actions_box.delete("1.0", tk.END)
        self.dashboard_actions_box.insert("1.0", "\n\n".join(f"- {a}" for a in actions))
        self.dashboard_actions_box.configure(state="disabled")

    def get_workflow_dashboard_snapshot(self):
        snapshot = {"open_rfqs": 0, "open_pos": 0, "open_po_value": 0.0, "timeline": []}
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT COUNT(*) FROM rfq_register WHERE status NOT IN ('Closed', 'Cancelled')")
            snapshot["open_rfqs"] = c.fetchone()[0] or 0
            c.execute("SELECT COUNT(*), COALESCE(SUM(total_value), 0) FROM po_register WHERE status NOT IN ('Closed', 'Cancelled')")
            po_count, po_value = c.fetchone()
            snapshot["open_pos"] = po_count or 0
            snapshot["open_po_value"] = float(po_value or 0)

            events = []
            c.execute("""
                SELECT created_at, workflow_type, workflow_id, action, status, note
                FROM workflow_audit_log
                ORDER BY datetime(created_at) DESC, id DESC
                LIMIT 12
            """)
            for created_at, workflow_type, workflow_id, action, status, note in c.fetchall():
                events.append((created_at, f"{created_at} | {workflow_type}-{workflow_id} | {action} -> {status}\n{note or ''}"))

            c.execute("""
                SELECT q.created_at, 'QUOTE', q.quote_id, q.action, q.new_status, COALESCE(e.supplier, ''), COALESCE(e.product, '')
                FROM quote_audit_log q
                LEFT JOIN extracted_quotes e ON e.id = q.quote_id
                ORDER BY datetime(q.created_at) DESC, q.id DESC
                LIMIT 8
            """)
            for created_at, workflow_type, quote_id, action, status, supplier, product in c.fetchall():
                note = f"{supplier} / {product}".strip(" /")
                events.append((created_at, f"{created_at} | {workflow_type}-{quote_id} | {action} -> {status}\n{note}"))

            conn.close()
            events.sort(key=lambda item: item[0] or "", reverse=True)
            snapshot["timeline"] = [text for _, text in events[:12]]
        except Exception as e:
            snapshot["timeline"] = [f"Unable to load workflow timeline: {e}"]
        return snapshot

    def quote_review_priority(self, row):
        score = 0
        if not row.get("price"):
            score += 4
        if not row.get("moq") or str(row.get("moq")).upper() == "N/A":
            score += 2
        if not row.get("lead_time") or str(row.get("lead_time")).upper() == "N/A":
            score += 2
        if "high" in str(row.get("sourcing_risk") or "").lower():
            score += 3
        try:
            import datetime
            validity = datetime.datetime.strptime(str(row.get("validity_date")), "%Y-%m-%d").date()
            if validity < datetime.date.today():
                score += 4
        except Exception:
            pass
        return score

    def quote_warning_summary(self, row):
        warnings = []
        if not row.get("price"):
            warnings.append("missing price")
        if not row.get("moq") or str(row.get("moq")).upper() == "N/A":
            warnings.append("missing MOQ")
        if not row.get("lead_time") or str(row.get("lead_time")).upper() == "N/A":
            warnings.append("missing lead time")
        if "high" in str(row.get("sourcing_risk") or "").lower():
            warnings.append("high risk")
        try:
            import datetime
            validity = datetime.datetime.strptime(str(row.get("validity_date")), "%Y-%m-%d").date()
            if validity < datetime.date.today():
                warnings.append("expired")
        except Exception:
            pass
        return ", ".join(warnings) if warnings else "ready for approval"

    def get_dashboard_selected_quote_ids(self):
        if not hasattr(self, "dashboard_approval_tree"):
            return []
        ids = []
        for item in self.dashboard_approval_tree.selection():
            vals = self.dashboard_approval_tree.item(item, "values")
            if vals:
                try:
                    ids.append(int(vals[0]))
                except Exception:
                    pass
        return ids

    def update_dashboard_selected_quote(self, status):
        quote_ids = self.get_dashboard_selected_quote_ids()
        if not quote_ids:
            messagebox.showwarning("Select Quote", "Please select a quote in the approval queue first.")
            return
        self.set_quote_status_by_ids(quote_ids, status)

    def open_dashboard_quote_in_quotes(self):
        quote_ids = self.get_dashboard_selected_quote_ids()
        self.show_page("Sourcing Analysis")
        try:
            self.sourcing_tabview.set("📊 Quotes Comparison")
        except Exception:
            pass
        if quote_ids and hasattr(self, "tree"):
            target = str(quote_ids[0])
            for item in self.tree.get_children():
                vals = self.tree.item(item, "values")
                if vals and str(vals[0]) == target:
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)
                    break

    def send_dashboard_quote_to_rfq(self):
        quote_ids = self.get_dashboard_selected_quote_ids()
        if not quote_ids:
            messagebox.showwarning("Select Quote", "Please select a quote in the approval queue first.")
            return
        quote = next((r for r in self.extracted_data if r.get("id") == quote_ids[0]), None)
        if not quote:
            return
        self.show_page("RFQs Outreach")
        try:
            self.rfqs_tabview.set("📝 RFQ Generator")
        except Exception:
            pass
        product = self.clean_product_name(quote.get("product"))
        self.rfq_product_cb.set("Custom")
        self.rfq_name_entry.delete(0, tk.END)
        self.rfq_name_entry.insert(0, product)
        self.rfq_specs_text.delete("1.0", tk.END)
        specs = "\n".join(x for x in [quote.get("spec"), quote.get("color"), quote.get("elastic"), quote.get("packing")] if x and x != "N/A")
        self.rfq_specs_text.insert("1.0", specs or f"High-quality {product} matching approved sourcing requirements.")
        messagebox.showinfo("RFQ Draft Ready", f"Loaded quote {quote_ids[0]} into the RFQ generator.")

    def on_closing(self):
        self.is_extracting = False
        self.chat_is_extracting = False
        self._is_closing = True
        try:
            for after_id in self.tk.call("after", "info"):
                try:
                    self.after_cancel(after_id)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            import matplotlib.pyplot as plt
            plt.close("all")
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass

    def get_validity_display(self, date_str):
        if not date_str or date_str.lower() in ["n/a", "null", "none", ""]:
            return "Unknown"
        import re
        match = re.search(r'\d{4}-\d{2}-\d{2}', date_str)
        if match:
            try:
                import datetime
                q_date = datetime.datetime.strptime(match.group(0), "%Y-%m-%d").date()
                today = datetime.date.today()
                if q_date < today:
                    return f"🔴 Expired ({match.group(0)})"
                else:
                    return f"🟢 Active ({match.group(0)})"
            except Exception:
                pass
        return date_str

    def get_risk_display(self, risk_str, supplier_name=None):
        cleaned_sup = self.clean_supplier_name(supplier_name) if supplier_name else None
        incidents = self.supplier_incidents_map.get(cleaned_sup, []) if hasattr(self, 'supplier_incidents_map') and cleaned_sup else []
        
        max_severity = None
        incident_types = []
        for sev, inc_type in incidents:
            incident_types.append(inc_type)
            if sev == "High":
                max_severity = "High"
            elif sev == "Medium" and max_severity != "High":
                max_severity = "Medium"
            elif sev == "Low" and max_severity not in ["High", "Medium"]:
                max_severity = "Low"
                
        if max_severity == "High":
            return f"🔴 High Risk (Incident: {', '.join(set(incident_types))})"
        elif max_severity == "Medium":
            return f"🟡 Medium Risk (Incident: {', '.join(set(incident_types))})"
        elif max_severity == "Low" and (not risk_str or risk_str.lower() in ["n/a", "null", "none", ""]):
            return "🟢 Low Risk (Logged Incident: Low)"

        if not risk_str or risk_str.lower() in ["n/a", "null", "none", ""]:
            return "🟢 Low Risk"
        r_lower = risk_str.lower()
        if "high" in r_lower or "prepayment" in r_lower or "expired" in r_lower:
            return f"🔴 {risk_str}"
        elif "medium" in r_lower or "warning" in r_lower or "long lead" in r_lower:
            return f"🟡 {risk_str}"
        else:
            return f"🟢 {risk_str}"

    def change_currency(self, choice):
        # Update column header with current currency name
        currency_choice = self.currency_cb.get()
        symbol = "$"
        if "CNY" in currency_choice:
            symbol = "¥"
        elif "EUR" in currency_choice:
            symbol = "€"
        self.tree.heading("price", text=f"Price ({symbol})")
        self.filter_table()
        self.draw_chart()

    def setup_table(self, parent):
        self.style_workspace_table("Treeview")
        
        # Columns
        self.columns = ("id", "status", "filename", "supplier", "product", "spec", "color", "elastic", "price", "unit", "moq", "packing", "term", "lead_time", "validity_date", "sourcing_risk")
        self.tree = ttk.Treeview(parent, columns=self.columns, show="headings", style="Treeview")
        
        # Setup column headers with commands for interactive sorting
        self.tree.heading("id", text="ID", command=lambda: self.sort_column("id", False))
        self.tree.heading("status", text="Review", command=lambda: self.sort_column("review_status", False))
        self.tree.heading("filename", text="Source File", command=lambda: self.sort_column("filename", False))
        self.tree.heading("supplier", text="Supplier", command=lambda: self.sort_column("supplier", False))
        self.tree.heading("product", text="Product", command=lambda: self.sort_column("product", False))
        self.tree.heading("spec", text="Specs", command=lambda: self.sort_column("spec", False))
        self.tree.heading("color", text="Color", command=lambda: self.sort_column("color", False))
        self.tree.heading("elastic", text="Elastic", command=lambda: self.sort_column("elastic", False))
        self.tree.heading("price", text="Price", command=lambda: self.sort_column("price", False)) # Dynamic header based on currency
        self.tree.heading("unit", text="Price Unit", command=lambda: self.sort_column("unit", False))
        self.tree.heading("moq", text="MOQ", command=lambda: self.sort_column("moq", False))
        self.tree.heading("packing", text="Packing", command=lambda: self.sort_column("packing", False))
        self.tree.heading("term", text="Price Term", command=lambda: self.sort_column("term", False))
        self.tree.heading("lead_time", text="Lead Time", command=lambda: self.sort_column("lead_time", False))
        self.tree.heading("validity_date", text="Validity", command=lambda: self.sort_column("validity_date", False))
        self.tree.heading("sourcing_risk", text="Risk Alerts", command=lambda: self.sort_column("sourcing_risk", False))

        # Column widths
        self.tree.column("id", width=40, anchor="center")
        self.tree.column("status", width=100, anchor="center")
        self.tree.column("filename", width=120, anchor="w")
        self.tree.column("supplier", width=130, anchor="w")
        self.tree.column("product", width=100, anchor="w")
        self.tree.column("spec", width=110, anchor="w")
        self.tree.column("color", width=60, anchor="center")
        self.tree.column("elastic", width=60, anchor="center")
        self.tree.column("price", width=60, anchor="center")
        self.tree.column("unit", width=60, anchor="center")
        self.tree.column("moq", width=80, anchor="center")
        self.tree.column("packing", width=110, anchor="w")
        self.tree.column("term", width=80, anchor="center")
        self.tree.column("lead_time", width=70, anchor="center")
        self.tree.column("validity_date", width=95, anchor="center")
        self.tree.column("sourcing_risk", width=130, anchor="w")

        # Scrollbars
        ysb = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview)
        xsb = ttk.Scrollbar(parent, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=ysb.set, xscroll=xsb.set)

        # Place Treeview in row 2
        self.tree.grid(row=2, column=0, sticky="nsew", padx=10)
        ysb.grid(row=2, column=1, sticky="ns")
        xsb.grid(row=3, column=0, sticky="ew")

        # Bind events
        self.tree.bind("<Double-1>", lambda event: self.edit_selected_row())
        self.tree.bind("<<TreeviewSelect>>", lambda event: self.on_row_selected(event))

        # Best Deal highlight tag
        self.tree.tag_configure("best_deal", background="#DCFCE7", foreground="#14532D")
        self.tree.tag_configure("needs_review", background="#FEF3C7", foreground="#92400E")
        self.tree.tag_configure("approved", background="#DCFCE7", foreground="#14532D")
        self.tree.tag_configure("rejected", background="#FEE2E2", foreground="#991B1B")

    # --- Direct Path Entry Handler ---
    def on_path_entered(self):
        path = self.folder_entry.get().strip()
        if path.startswith('"') and path.endswith('"'):
            path = path[1:-1]
        
        if path == self.selected_folder:
            self.folder_entry.configure(fg_color=None)
            return

        if path and os.path.exists(path) and os.path.isdir(path):
            self.selected_folder = path
            self.folder_entry.configure(fg_color=None)
            self.save_config()
            self.scan_folder()
        elif path:
            self.folder_entry.configure(fg_color="#5a1f1f")

    # --- Document Preview Event Handlers ---
    def on_row_selected(self, event):
        sel = self.tree.selection()
        if not sel:
            self.clear_preview()
            return
            
        item = sel[0]
        vals = self.tree.item(item, "values")
        
        # 1. Update Preview metrics overlay
        self.update_preview_metrics_overlay(vals)
        
        # 2. Update Gallery selector buttons
        row_id = vals[0]
        self.update_preview_gallery_bar(row_id, vals)
        
        filename = vals[2]
        self.show_preview(filename)

    def clear_preview(self):
        self.preview_filename_lbl.configure(text="Select a row to preview", text_color="grey")
        self.preview_image_lbl.pack_forget()
        self.preview_text_box.pack_forget()
        for widget in self.preview_display_frame.winfo_children():
            if widget not in [self.preview_image_lbl, self.preview_text_box]:
                widget.destroy()
        
        if hasattr(self, 'preview_gallery_bar'):
            for widget in self.preview_gallery_bar.winfo_children():
                widget.destroy()

        self.preview_image_lbl.configure(image=None, text="No document selected")
        self.preview_image_lbl.pack(pady=120, padx=10)
        self.btn_open_external.configure(state="disabled")
        self.current_preview_path = ""
        self.update_preview_metrics_overlay(None)

    def show_preview(self, filename):
        self.preview_filename_lbl.configure(text=filename, text_color="white")
        
        full_path = os.path.join(self.selected_folder, filename)
        
        if filename == "Manually Added" or "Chat (" in filename or not filename:
            self.show_preview_message("Text entry (no document preview)")
            self.btn_open_external.configure(state="disabled")
            return
            
        if not os.path.exists(full_path):
            self.show_preview_message(f"File not found in folder:\n{self.selected_folder}")
            self.btn_open_external.configure(state="disabled")
            return
            
        self.btn_open_external.configure(state="normal")
        self.current_preview_path = full_path
        
        ext = os.path.splitext(filename)[1].lower()
        
        try:
            self.preview_text_box.pack_forget()
            self.preview_image_lbl.pack_forget()
            
            if ext in {".png", ".jpg", ".jpeg"}:
                self.render_image_preview(full_path)
            elif ext == ".pdf":
                self.render_pdf_preview(full_path)
            elif ext == ".txt":
                self.render_text_preview(full_path)
            else:
                self.show_preview_message(f"Preview not supported for {ext} files.\nUse system button below to open.")
        except Exception as e:
            self.show_preview_message(f"Error loading preview:\n{e}")

    def render_image_preview(self, path):
        try:
            pil_img = Image.open(path)
            self.display_image_in_preview(pil_img)
        except Exception as e:
            self.show_preview_message(f"Failed to load image:\n{e}")

    def render_pdf_preview(self, path):
        try:
            import fitz
            doc = fitz.open(path)
            if len(doc) > 0:
                page = doc.load_page(0)
                pix = page.get_pixmap(dpi=120)
                pil_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                self.display_image_in_preview(pil_img)
            else:
                self.show_preview_message("PDF document has no pages")
        except ImportError:
            self.show_preview_message("Install PyMuPDF to view PDF previews")
        except Exception as e:
            self.show_preview_message(f"Failed to render PDF preview:\n{e}")

    def render_text_preview(self, path):
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                text_content = f.read(5000)
            self.preview_text_box.delete("1.0", tk.END)
            self.preview_text_box.insert("1.0", text_content)
            self.preview_text_box.pack(fill="both", expand=True, padx=5, pady=5)
        except Exception as e:
            self.show_preview_message(f"Failed to read text file:\n{e}")

    def display_image_in_preview(self, pil_img):
        width, height = pil_img.size
        max_w, max_h = 330, 480
        
        ratio = min(max_w / width, max_h / height)
        new_w = max(int(width * ratio), 1)
        new_h = max(int(height * ratio), 1)
        
        ctk_img = ctk.CTkImage(light_image=pil_img, dark_image=pil_img, size=(new_w, new_h))
        self.preview_image_lbl.configure(image=ctk_img, text="")
        self.preview_image_lbl.image = ctk_img
        self.preview_image_lbl.pack(pady=10)

    def show_preview_message(self, message):
        self.preview_image_lbl.pack_forget()
        self.preview_text_box.pack_forget()
        self.preview_image_lbl.configure(image=None, text=message)
        self.preview_image_lbl.pack(pady=100)

    def open_file_externally(self):
        if self.current_preview_path:
            try:
                os.startfile(self.current_preview_path)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file: {e}")

    # --- Best Deals Group Key logic ---
    def get_group_key(self, row):
        prod = (row.get("product") or "").lower().strip()
        spec = (row.get("spec") or "").lower()
        
        category = "other"
        if "cap" in prod or "net" in prod or "bouffant" in prod:
            category = "cap"
        elif "apron" in prod:
            category = "apron"
        elif "sleeve" in prod:
            category = "sleeve"
        elif "shoe" in prod:
            category = "shoe_cover"
        elif "mask" in prod:
            category = "mask"
        else:
            category = prod
            
        import re
        numbers = re.findall(r'\d+', prod + " " + spec)
        size_key = numbers[0] if numbers else "default"
        
        return f"{category}_{size_key}"

    def get_decision_quotes(self, rows=None):
        rows = rows if rows is not None else self.extracted_data
        non_rejected = [r for r in rows if (r.get("review_status") or "Needs Review") != "Rejected"]
        approved = [r for r in non_rejected if r.get("review_status") == "Approved"]
        return approved if approved else non_rejected

    # --- Interactive Sorting logic ---
    def sort_column(self, col, reverse):
        def get_sort_val(row_dict):
            val = row_dict.get(col)
            if val is None:
                return ""
            
            if col == "price":
                try:
                    return float(val)
                except ValueError:
                    return 0.0
            if col == "id":
                try:
                    return int(val)
                except ValueError:
                    return 0
            return str(val).lower()

        self.extracted_data.sort(key=get_sort_val, reverse=reverse)
        self.filter_table()
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

    # --- Clear All Data reset ---
    def clear_all_data(self):
        ans = messagebox.askyesno(
            "Clear All Data",
            "⚠️ WARNING: This will permanently delete all quote data, processed files, chat history, and supplier directory records from the database.\n\nAre you sure you want to proceed?"
        )
        if ans:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("DELETE FROM extracted_quotes")
            c.execute("DELETE FROM processed_files")
            c.execute("DELETE FROM supplier_contacts")
            c.execute("DELETE FROM chat_history")
            c.execute("DELETE FROM quote_audit_log")
            c.execute("DELETE FROM master_data_audit_log")
            c.execute("DELETE FROM rfq_register")
            c.execute("DELETE FROM po_register")
            c.execute("DELETE FROM workflow_audit_log")
            conn.commit()
            conn.close()
            
            self.load_all_quotes_from_db()
            self.scan_folder()
            self.clear_preview()
            self.load_chat_history_from_db()
            
            messagebox.showinfo("Database Cleared", "Database reset successfully! All quotes and histories have been cleared.")

    # --- Live Search Filter rendering ---
    def filter_table(self):
        query = self.search_entry.get().strip().lower()
        self.tree.delete(*self.tree.get_children())
        
        # Get active currency conversion specs
        currency_choice = self.currency_cb.get()
        factor = 1.0
        symbol = "$"
        if "CNY" in currency_choice:
            factor = self.exchange_rates["CNY"]
            symbol = "¥"
        elif "EUR" in currency_choice:
            factor = self.exchange_rates["EUR"]
            symbol = "€"
        
        # Recalculate best price comparisons from in-memory quotes (baseline USD)
        groups = {}
        for r in self.get_decision_quotes():
            price = r.get("price")
            try:
                price = float(price)
            except (ValueError, TypeError):
                continue
            if price <= 0:
                continue
            
            g_key = self.get_group_key(r)
            if g_key not in groups:
                groups[g_key] = []
            groups[g_key].append(price)

        best_prices = {}
        for g_key, prices in groups.items():
            if len(prices) > 1:
                best_prices[g_key] = min(prices)

        # Populate tree matching search query
        for row_data in self.extracted_data:
            match = False
            if not query:
                match = True
            else:
                fields = ["review_status", "filename", "supplier", "product", "spec", "color", "elastic", "price", "unit", "moq", "packing", "term", "lead_time", "validity_date", "sourcing_risk"]
                for f in fields:
                    if query in str(row_data.get(f) or "").lower():
                        match = True
                        break
            
            if not match:
                continue

            try:
                price_val = float(row_data["price"])
                converted_price = price_val * factor
                price_display = f"{symbol}{converted_price:.5f}"
            except (ValueError, TypeError):
                price_display = str(row_data["price"])
            
            is_best_deal = False
            g_key = self.get_group_key(row_data)
            try:
                price_val = float(row_data["price"])
                if g_key in best_prices and abs(price_val - best_prices[g_key]) < 1e-7:
                    is_best_deal = True
            except (ValueError, TypeError):
                pass
                
            status = row_data.get("review_status") or "Needs Review"
            if status == "Approved":
                status_display = "Approved"
                status_tag = "approved"
            elif status == "Rejected":
                status_display = "Rejected"
                status_tag = "rejected"
            else:
                status_display = "Needs Review"
                status_tag = "needs_review"

            tags_list = [status_tag]
            if is_best_deal and status == "Approved":
                tags_list.append("best_deal")
            tags = tuple(tags_list)
            
            self.tree.insert("", "end", values=(
                row_data["id"],
                status_display,
                row_data["filename"],
                row_data["supplier"],
                row_data["product"],
                row_data["spec"],
                row_data["color"],
                row_data["elastic"],
                price_display,
                row_data["unit"],
                row_data["moq"],
                row_data["packing"],
                row_data["term"],
                row_data["lead_time"],
                self.get_validity_display(row_data.get("validity_date")),
                self.get_risk_display(row_data.get("sourcing_risk"), supplier_name=row_data.get("supplier"))
            ), tags=tags)

    # --- DB Retrieval ---
    def load_all_quotes_from_db(self):
        self.tree.delete(*self.tree.get_children())
        self.extracted_data = []

        # Load incident status mapping
        self.supplier_incidents_map = {}
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT supplier, severity, incident_type FROM supplier_incidents")
            for sup_n, sev_v, inc_t in c.fetchall():
                sup_clean = self.clean_supplier_name(sup_n)
                if sup_clean not in self.supplier_incidents_map:
                    self.supplier_incidents_map[sup_clean] = []
                self.supplier_incidents_map[sup_clean].append((sev_v, inc_t))
            conn.close()
        except Exception as e:
            print("Error loading incidents:", e)

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT id, filename, supplier, product, spec, color, elastic, price, unit, moq, packing, term, lead_time, validity_date, sourcing_risk, attached_media,
                   review_status, reviewed_by, reviewed_at, review_notes
            FROM extracted_quotes
        """)
        rows = c.fetchall()
        conn.close()

        for row in rows:
            row_data = {
                "id": row[0],
                "filename": row[1],
                "supplier": row[2],
                "product": row[3],
                "spec": row[4],
                "color": row[5],
                "elastic": row[6],
                "price": row[7],
                "unit": row[8],
                "moq": row[9],
                "packing": row[10],
                "term": row[11],
                "lead_time": row[12],
                "validity_date": row[13] if len(row) > 13 else "N/A",
                "sourcing_risk": row[14] if len(row) > 14 else "N/A",
                "attached_media": row[15] if len(row) > 15 else "",
                "review_status": row[16] if len(row) > 16 and row[16] else "Needs Review",
                "reviewed_by": row[17] if len(row) > 17 else "",
                "reviewed_at": row[18] if len(row) > 18 else "",
                "review_notes": row[19] if len(row) > 19 else ""
            }
            self.extracted_data.append(row_data)

            if row_data["id"] >= self.quote_id_counter:
                self.quote_id_counter = row_data["id"] + 1

        self.filter_table()
        self.load_supplier_directory()
        self.update_chart_dropdown()
        self.draw_chart()
        self.update_sourcing_insights()
        self.update_scorecard_tab()
        self.update_timeline_tab()
        if hasattr(self, 'sim_tree'):
            self.update_landed_cost_tab()
        if hasattr(self, 'opt_scroll'):
            self.update_purchase_optimizer_tab()
        if hasattr(self, 'rfq_product_cb'):
            self.update_rfq_generator_tab()
        if hasattr(self, 'profit_tree'):
            self.update_profit_simulator_tab()
        if hasattr(self, 'qc_supplier_cb'):
            self.update_factory_qc_tab()
        if hasattr(self, 'matrix_frame'):
            self.update_sourcing_matrix_tab()
        if hasattr(self, 'uae_category_cb'):
            self.update_uae_customs_tab()
        if hasattr(self, 'packing_product_cb'):
            self.update_container_packing_tab()
        if hasattr(self, 'po_supplier_cb'):
            self.load_po_suppliers()
        if hasattr(self, 'inc_supplier_cb'):
            self.load_incident_logs()
        if hasattr(self, 'cny_slider'):
            self.draw_hedge_chart()
        if hasattr(self, 'hist_supplier_cb'):
            self.load_price_history_dropdowns()
        if hasattr(self, 'neg_supplier_cb'):
            self.load_negotiation_dropdowns()
        if hasattr(self, 'barrier_category_cb'):
            self.load_barrier_categories()
        if hasattr(self, 'dashboard_cards'):
            self.update_dashboard_page()
        if hasattr(self, 'supplier_master_tree'):
            self.load_master_data_tables()
        if hasattr(self, 'rfq_register_tree'):
            self.load_rfq_register()
        if hasattr(self, 'po_register_tree'):
            self.load_po_register()
        self.apply_legacy_light_polish()

    # --- Persistent Chat History logic ---
    def load_chat_history_from_db(self):
        self.chat_log.configure(state="normal")
        self.chat_log.delete("1.0", tk.END)
        self.chat_log.insert("1.0", "AI: Hello! Ask me anything about your current supplier quotes (e.g. comparing prices, payment terms, or drafting negotiation emails).\n")
        
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT sender, message FROM chat_history ORDER BY id ASC")
        rows = c.fetchall()
        conn.close()
        
        for sender, message in rows:
            self.chat_log.insert(tk.END, f"\n{sender}: {message}\n")
            
        self.chat_log.see(tk.END)
        self.chat_log.configure(state="disabled")

    def save_chat_message_to_db(self, sender, message):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("INSERT INTO chat_history (sender, message) VALUES (?, ?)", (sender, message))
        conn.commit()
        conn.close()

    def log_quote_audit(self, quote_id, action, previous_status=None, new_status=None, note=""):
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                INSERT INTO quote_audit_log (quote_id, action, previous_status, new_status, note)
                VALUES (?, ?, ?, ?, ?)
            """, (quote_id, action, previous_status, new_status, note))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Failed to write quote audit log: {e}")

    def log_master_audit(self, entity_type, entity_id, action, note=""):
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                INSERT INTO master_data_audit_log (entity_type, entity_id, action, note)
                VALUES (?, ?, ?, ?)
            """, (entity_type, entity_id, action, note))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Failed to write master data audit log: {e}")

    def seed_master_data_from_quotes(self, show_message=True):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        inserted_suppliers = 0
        inserted_products = 0

        c.execute("""
            SELECT supplier, MIN(term), GROUP_CONCAT(DISTINCT product)
            FROM extracted_quotes
            WHERE supplier IS NOT NULL AND supplier != '' AND supplier != 'Unknown'
            GROUP BY supplier
        """)
        for supplier, payment_terms, categories in c.fetchall():
            display_name = self.clean_supplier_name(supplier)
            if not display_name or display_name == "Unknown":
                continue
            c.execute("SELECT id FROM supplier_master WHERE display_name = ?", (display_name,))
            if c.fetchone():
                continue
            c.execute("""
                INSERT INTO supplier_master (legal_name, display_name, category, status, payment_terms, notes)
                VALUES (?, ?, ?, 'Active', ?, ?)
            """, (
                supplier,
                display_name,
                (categories or "General")[:120],
                payment_terms or "",
                "Seeded from extracted quote history."
            ))
            supplier_id = c.lastrowid
            c.execute("""
                INSERT INTO master_data_audit_log (entity_type, entity_id, action, note)
                VALUES ('Supplier', ?, 'Seeded', ?)
            """, (supplier_id, f"Created supplier master from quote supplier: {supplier}"))
            inserted_suppliers += 1

        c.execute("""
            SELECT product, MIN(spec), MIN(packing), MIN(price)
            FROM extracted_quotes
            WHERE product IS NOT NULL AND product != '' AND product != 'N/A'
            GROUP BY product
        """)
        for product, specs, packing, target_price in c.fetchall():
            product_name = self.clean_product_name(product)
            if not product_name or product_name == "Product":
                continue
            c.execute("SELECT id FROM product_master WHERE product_name = ?", (product_name,))
            if c.fetchone():
                continue
            c.execute("""
                INSERT INTO product_master (product_name, category, standard_specs, packaging, target_price, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                product_name,
                product_name,
                specs or "",
                packing or "",
                target_price if target_price is not None else None,
                "Seeded from extracted quote history."
            ))
            product_id = c.lastrowid
            c.execute("""
                INSERT INTO master_data_audit_log (entity_type, entity_id, action, note)
                VALUES ('Product', ?, 'Seeded', ?)
            """, (product_id, f"Created product master from quote product: {product}"))
            inserted_products += 1

        c.execute("""
            UPDATE extracted_quotes
            SET supplier_master_id = (
                SELECT sm.id
                FROM supplier_master sm
                WHERE sm.display_name = supplier
                   OR sm.legal_name = supplier
                   OR sm.display_name = TRIM(supplier)
                LIMIT 1
            )
            WHERE supplier_master_id IS NULL
        """)
        c.execute("""
            UPDATE extracted_quotes
            SET product_master_id = (
                SELECT pm.id
                FROM product_master pm
                WHERE LOWER(pm.product_name) = LOWER(TRIM(product))
                LIMIT 1
            )
            WHERE product_master_id IS NULL
        """)

        conn.commit()
        conn.close()

        if hasattr(self, "supplier_master_tree"):
            self.load_master_data_tables()
        if show_message:
            messagebox.showinfo(
                "Master Data Seeded",
                f"Created {inserted_suppliers} supplier master record(s) and {inserted_products} product master record(s)."
            )

    def set_selected_quote_status(self, status):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Row", "Please select at least one quote row first.")
            return

        quote_ids = []
        for item in sel:
            vals = self.tree.item(item, "values")
            try:
                quote_ids.append(int(vals[0]))
            except Exception:
                pass
        self.set_quote_status_by_ids(quote_ids, status)

    def set_quote_status_by_ids(self, quote_ids, status):
        if not quote_ids:
            messagebox.showwarning("Select Row", "Please select at least one quote row first.")
            return

        status_notes = {
            "Approved": "Quote approved for sourcing comparison and downstream PO/RFQ use.",
            "Needs Review": "Quote returned to review queue for verification.",
            "Rejected": "Quote rejected from approved sourcing decisions.",
        }
        note = status_notes.get(status, "")
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        updated = 0
        for q_id in quote_ids:
            current = next((r.get("review_status") for r in self.extracted_data if r.get("id") == q_id), None)
            c.execute("""
                UPDATE extracted_quotes
                SET review_status=?, reviewed_by='Local User', reviewed_at=datetime('now'), review_notes=?
                WHERE id=?
            """, (status, note, q_id))
            c.execute("""
                INSERT INTO quote_audit_log (quote_id, action, previous_status, new_status, note)
                VALUES (?, ?, ?, ?, ?)
            """, (q_id, "Review Status Changed", current, status, note))
            updated += 1
        conn.commit()
        conn.close()

        self.load_all_quotes_from_db()
        messagebox.showinfo("Review Status Updated", f"Updated {updated} quote row(s) to: {status}")

    def sync_suppliers_from_quotes(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT DISTINCT supplier FROM extracted_quotes WHERE supplier IS NOT NULL AND supplier != 'Unknown'")
        suppliers = [row[0] for row in c.fetchall()]
        
        synced_count = 0
        for s in suppliers:
            c.execute("SELECT filename FROM extracted_quotes WHERE supplier = ? LIMIT 1", (s,))
            q_row = c.fetchone()
            filename = q_row[0] if q_row else "N/A"
            
            c.execute("SELECT COUNT(*) FROM supplier_contacts WHERE supplier = ?", (s,))
            if c.fetchone()[0] == 0:
                c.execute("INSERT INTO supplier_contacts (supplier, contact_info, source_file) VALUES (?, ?, ?)", 
                          (s, "Contact info placeholder. Click '✏ Edit Info' to fill manually, or process a new quote from this supplier to auto-extract details.", filename))
                synced_count += 1
            else:
                c.execute("UPDATE supplier_contacts SET source_file = ? WHERE supplier = ? AND (source_file IS NULL OR source_file = '')", (filename, s))
                
        conn.commit()
        conn.close()
        
        self.load_supplier_directory()
        messagebox.showinfo("Sync Complete", f"Successfully synced supplier directory!\nAdded {synced_count} new supplier cards.")

    def setup_master_data_tab(self):
        tab_master = self.settings_tabview.tab("Master Data")
        tab_master.grid_columnconfigure(0, weight=1)
        tab_master.grid_columnconfigure(1, weight=1)
        tab_master.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(tab_master, fg_color="transparent")
        header.grid(row=0, column=0, columnspan=2, sticky="ew", padx=16, pady=(14, 8))
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="Enterprise Master Data",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=self.THEME["text"],
        ).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(
            header,
            text="Controlled supplier and product records used by quotes, RFQs, scorecards, costing, and POs.",
            font=ctk.CTkFont(size=12),
            text_color=self.THEME["muted"],
        ).grid(row=1, column=0, sticky="w", pady=(3, 0))

        self.btn_seed_master_data = self.make_button(header, "Seed from Quotes", command=self.seed_master_data_from_quotes, variant="success", width=130)
        self.btn_seed_master_data.grid(row=0, column=1, rowspan=2, sticky="e", padx=(8, 0))
        self.btn_refresh_master_data = self.make_button(header, "Refresh", command=self.load_master_data_tables, variant="secondary", width=90)
        self.btn_refresh_master_data.grid(row=0, column=2, rowspan=2, sticky="e", padx=(8, 0))

        supplier_panel = ctk.CTkFrame(tab_master, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        supplier_panel.grid(row=1, column=0, sticky="nsew", padx=(16, 8), pady=(0, 14))
        supplier_panel.grid_columnconfigure(0, weight=1)
        supplier_panel.grid_rowconfigure(1, weight=1)

        supplier_head = ctk.CTkFrame(supplier_panel, fg_color="transparent")
        supplier_head.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
        supplier_head.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(supplier_head, text="Supplier Master", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w")
        self.btn_edit_supplier_master = self.make_button(supplier_head, "Edit Supplier", command=self.edit_selected_supplier_master, variant="secondary", width=110)
        self.btn_edit_supplier_master.grid(row=0, column=1, sticky="e")

        supplier_cols = ("id", "display", "status", "category", "contact", "phone")
        self.supplier_master_tree = ttk.Treeview(supplier_panel, columns=supplier_cols, show="headings", style="Treeview")
        for col, label, width in [
            ("id", "ID", 45),
            ("display", "Supplier", 160),
            ("status", "Status", 85),
            ("category", "Category", 130),
            ("contact", "Email / Contact", 150),
            ("phone", "Phone", 105),
        ]:
            self.supplier_master_tree.heading(col, text=label)
            self.supplier_master_tree.column(col, width=width, anchor="w")
        self.supplier_master_tree.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))

        product_panel = ctk.CTkFrame(tab_master, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        product_panel.grid(row=1, column=1, sticky="nsew", padx=(8, 16), pady=(0, 14))
        product_panel.grid_columnconfigure(0, weight=1)
        product_panel.grid_rowconfigure(1, weight=1)

        product_head = ctk.CTkFrame(product_panel, fg_color="transparent")
        product_head.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
        product_head.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(product_head, text="Product Master", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w")
        self.btn_edit_product_master = self.make_button(product_head, "Edit Product", command=self.edit_selected_product_master, variant="secondary", width=105)
        self.btn_edit_product_master.grid(row=0, column=1, sticky="e")

        product_cols = ("id", "product", "category", "target", "packaging", "specs")
        self.product_master_tree = ttk.Treeview(product_panel, columns=product_cols, show="headings", style="Treeview")
        for col, label, width in [
            ("id", "ID", 45),
            ("product", "Product", 140),
            ("category", "Category", 100),
            ("target", "Target", 70),
            ("packaging", "Packaging", 120),
            ("specs", "Specs", 160),
        ]:
            self.product_master_tree.heading(col, text=label)
            self.product_master_tree.column(col, width=width, anchor="w")
        self.product_master_tree.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))

    def load_master_data_tables(self):
        if not hasattr(self, "supplier_master_tree"):
            return

        self.supplier_master_tree.delete(*self.supplier_master_tree.get_children())
        self.product_master_tree.delete(*self.product_master_tree.get_children())

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT id, display_name, status, category, COALESCE(email, contact_person, ''), COALESCE(phone, '')
            FROM supplier_master
            ORDER BY display_name
        """)
        for row in c.fetchall():
            self.supplier_master_tree.insert("", tk.END, values=row)

        c.execute("""
            SELECT id, product_name, category, target_price, packaging, standard_specs
            FROM product_master
            ORDER BY product_name
        """)
        for row in c.fetchall():
            row = list(row)
            row[3] = f"${float(row[3]):.5f}" if row[3] is not None else "N/A"
            self.product_master_tree.insert("", tk.END, values=row)
        conn.close()

    def edit_selected_supplier_master(self):
        sel = self.supplier_master_tree.selection()
        if not sel:
            messagebox.showwarning("Select Supplier", "Please select a supplier master record first.")
            return
        supplier_id = int(self.supplier_master_tree.item(sel[0], "values")[0])

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT legal_name, display_name, country, city, contact_person, email, phone, category, status, payment_terms, certifications, notes
            FROM supplier_master WHERE id=?
        """, (supplier_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            return

        fields = [
            ("Legal Name", "legal_name"),
            ("Display Name", "display_name"),
            ("Country", "country"),
            ("City", "city"),
            ("Contact Person", "contact_person"),
            ("Email", "email"),
            ("Phone / WhatsApp", "phone"),
            ("Category", "category"),
            ("Status", "status"),
            ("Payment Terms", "payment_terms"),
            ("Certifications", "certifications"),
            ("Notes", "notes"),
        ]
        self.open_master_edit_dialog(
            "Supplier",
            supplier_id,
            fields,
            dict(zip([f[1] for f in fields], row)),
            self.save_supplier_master_record,
        )

    def edit_selected_product_master(self):
        sel = self.product_master_tree.selection()
        if not sel:
            messagebox.showwarning("Select Product", "Please select a product master record first.")
            return
        product_id = int(self.product_master_tree.item(sel[0], "values")[0])

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT product_name, category, standard_specs, packaging, carton_cbm, compliance_requirements, target_price, notes
            FROM product_master WHERE id=?
        """, (product_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            return

        fields = [
            ("Product Name", "product_name"),
            ("Category", "category"),
            ("Standard Specs", "standard_specs"),
            ("Packaging", "packaging"),
            ("Carton CBM", "carton_cbm"),
            ("Compliance Requirements", "compliance_requirements"),
            ("Target Price", "target_price"),
            ("Notes", "notes"),
        ]
        self.open_master_edit_dialog(
            "Product",
            product_id,
            fields,
            dict(zip([f[1] for f in fields], row)),
            self.save_product_master_record,
        )

    def open_master_edit_dialog(self, entity_type, entity_id, fields, values, save_callback):
        edit_win = ctk.CTkToplevel(self)
        edit_win.title(f"Edit {entity_type} Master #{entity_id}")
        edit_win.geometry("520x620")
        edit_win.resizable(False, False)
        edit_win.attributes("-topmost", True)
        edit_win.grid_columnconfigure(1, weight=1)

        entries = {}
        for idx, (label, field) in enumerate(fields):
            ctk.CTkLabel(edit_win, text=f"{label}:", anchor="w").grid(row=idx, column=0, padx=15, pady=6, sticky="w")
            if field in {"status"}:
                entry = ctk.CTkComboBox(edit_win, values=["Active", "Preferred", "Watchlist", "Blocked"], width=300)
                entry.set(values.get(field) or "Active")
            else:
                entry = ctk.CTkEntry(edit_win, width=300)
                entry.insert(0, "" if values.get(field) is None else str(values.get(field)))
            entry.grid(row=idx, column=1, padx=15, pady=6, sticky="ew")
            entries[field] = entry

        def save():
            payload = {field: widget.get().strip() for field, widget in entries.items()}
            save_callback(entity_id, payload)
            edit_win.destroy()

        self.make_button(edit_win, "Save Master Record", command=save, variant="success", width=180).grid(row=len(fields), column=0, columnspan=2, pady=18)

    def save_supplier_master_record(self, supplier_id, payload):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            UPDATE supplier_master
            SET legal_name=?, display_name=?, country=?, city=?, contact_person=?, email=?, phone=?,
                category=?, status=?, payment_terms=?, certifications=?, notes=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            payload.get("legal_name"),
            payload.get("display_name"),
            payload.get("country"),
            payload.get("city"),
            payload.get("contact_person"),
            payload.get("email"),
            payload.get("phone"),
            payload.get("category"),
            payload.get("status"),
            payload.get("payment_terms"),
            payload.get("certifications"),
            payload.get("notes"),
            supplier_id,
        ))
        c.execute("""
            INSERT INTO master_data_audit_log (entity_type, entity_id, action, note)
            VALUES ('Supplier', ?, 'Updated', ?)
        """, (supplier_id, "Supplier master edited from Master Data screen."))
        conn.commit()
        conn.close()
        self.load_master_data_tables()

    def save_product_master_record(self, product_id, payload):
        try:
            carton_cbm = float(payload.get("carton_cbm")) if payload.get("carton_cbm") else None
        except ValueError:
            messagebox.showerror("Invalid Input", "Carton CBM must be numeric.")
            return
        try:
            target_price = float(str(payload.get("target_price") or "").replace("$", "")) if payload.get("target_price") else None
        except ValueError:
            messagebox.showerror("Invalid Input", "Target price must be numeric.")
            return

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            UPDATE product_master
            SET product_name=?, category=?, standard_specs=?, packaging=?, carton_cbm=?,
                compliance_requirements=?, target_price=?, notes=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            payload.get("product_name"),
            payload.get("category"),
            payload.get("standard_specs"),
            payload.get("packaging"),
            carton_cbm,
            payload.get("compliance_requirements"),
            target_price,
            payload.get("notes"),
            product_id,
        ))
        c.execute("""
            INSERT INTO master_data_audit_log (entity_type, entity_id, action, note)
            VALUES ('Product', ?, 'Updated', ?)
        """, (product_id, "Product master edited from Master Data screen."))
        conn.commit()
        conn.close()
        self.load_master_data_tables()

    def get_product_master_id_by_name(self, product_name):
        if not product_name:
            return None
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT id FROM product_master WHERE LOWER(product_name)=LOWER(?) LIMIT 1", (self.clean_product_name(product_name),))
        row = c.fetchone()
        conn.close()
        return row[0] if row else None

    def get_supplier_master_id_by_name(self, supplier_name):
        if not supplier_name:
            return None
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT id FROM supplier_master
            WHERE LOWER(display_name)=LOWER(?) OR LOWER(legal_name)=LOWER(?)
            LIMIT 1
        """, (self.clean_supplier_name(supplier_name), str(supplier_name).strip()))
        row = c.fetchone()
        conn.close()
        return row[0] if row else None

    def save_rfq_record(self, status="Draft", pdf_path=""):
        product_name = self.rfq_name_entry.get().strip() or self.rfq_product_cb.get().strip()
        if not product_name or product_name == "Custom":
            return None
        import datetime
        rfq_number = f"RFQ-{datetime.date.today().strftime('%Y%m%d')}-{product_name[:4].upper()}"
        product_master_id = self.get_product_master_id_by_name(product_name)
        payload = (
            rfq_number,
            product_master_id,
            product_name,
            self.rfq_qty_entry.get().strip(),
            self.rfq_term_cb.get().strip(),
            self.rfq_lead_entry.get().strip(),
            self.rfq_payment_entry.get().strip(),
            "",
            status,
            pdf_path,
            self.rfq_specs_text.get("1.0", tk.END).strip(),
        )
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT INTO rfq_register (
                rfq_number, product_master_id, product_name, target_quantity, price_terms,
                lead_time, payment_terms, selected_suppliers, status, pdf_path, specs
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(rfq_number) DO UPDATE SET
                product_master_id=excluded.product_master_id,
                product_name=excluded.product_name,
                target_quantity=excluded.target_quantity,
                price_terms=excluded.price_terms,
                lead_time=excluded.lead_time,
                payment_terms=excluded.payment_terms,
                selected_suppliers=excluded.selected_suppliers,
                status=excluded.status,
                pdf_path=COALESCE(NULLIF(excluded.pdf_path, ''), rfq_register.pdf_path),
                specs=excluded.specs,
                updated_at=datetime('now')
        """, payload)
        c.execute("SELECT id FROM rfq_register WHERE rfq_number=?", (rfq_number,))
        rfq_id = c.fetchone()[0]
        c.execute("""
            INSERT INTO workflow_audit_log (workflow_type, workflow_id, action, status, note)
            VALUES ('RFQ', ?, 'Saved', ?, ?)
        """, (rfq_id, status, f"RFQ record saved for {product_name}."))
        conn.commit()
        conn.close()
        if hasattr(self, "rfq_register_tree"):
            self.load_rfq_register()
        if hasattr(self, "dashboard_cards"):
            self.update_dashboard_page()
        return rfq_id

    def save_po_record(self, status="Issued", pdf_path=""):
        supplier = self.po_supplier_cb.get()
        product = self.po_product_cb.get()
        po_num = self.po_number_entry.get().strip()
        try:
            qty = int(self.po_qty_entry.get().strip().replace(",", ""))
        except Exception:
            qty = 0
        unit_price = getattr(self, 'po_active_price', 0.0)
        total_cost = qty * unit_price
        quote = getattr(self, "po_active_quote", None) or {}
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT INTO po_register (
                po_number, supplier_master_id, product_master_id, quote_id, supplier_name, product_name,
                quantity, unit_cost, total_value, payment_terms, delivery_address, status, pdf_path
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(po_number) DO UPDATE SET
                supplier_master_id=excluded.supplier_master_id,
                product_master_id=excluded.product_master_id,
                quote_id=excluded.quote_id,
                supplier_name=excluded.supplier_name,
                product_name=excluded.product_name,
                quantity=excluded.quantity,
                unit_cost=excluded.unit_cost,
                total_value=excluded.total_value,
                payment_terms=excluded.payment_terms,
                delivery_address=excluded.delivery_address,
                status=excluded.status,
                pdf_path=COALESCE(NULLIF(excluded.pdf_path, ''), po_register.pdf_path),
                updated_at=datetime('now')
        """, (
            po_num,
            quote.get("supplier_master_id") or self.get_supplier_master_id_by_name(supplier),
            quote.get("product_master_id") or self.get_product_master_id_by_name(product),
            quote.get("id"),
            supplier,
            product,
            qty,
            unit_price,
            total_cost,
            self.po_payment_entry.get().strip(),
            self.po_address_entry.get().strip(),
            status,
            pdf_path,
        ))
        c.execute("SELECT id FROM po_register WHERE po_number=?", (po_num,))
        po_id = c.fetchone()[0]
        c.execute("""
            INSERT INTO workflow_audit_log (workflow_type, workflow_id, action, status, note)
            VALUES ('PO', ?, 'Saved', ?, ?)
        """, (po_id, status, f"PO record saved for {supplier} / {product}."))
        conn.commit()
        conn.close()
        if hasattr(self, "po_register_tree"):
            self.load_po_register()
        if hasattr(self, "dashboard_cards"):
            self.update_dashboard_page()
        return po_id

    def log_workflow_blocked_attempt(self, workflow_type, action, note, workflow_id=0):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT INTO workflow_audit_log (workflow_type, workflow_id, action, status, note)
            VALUES (?, ?, ?, 'Blocked', ?)
        """, (workflow_type, workflow_id or 0, action, note))
        conn.commit()
        conn.close()
        if hasattr(self, "dashboard_cards"):
            self.update_dashboard_page()

    def parse_positive_int(self, value):
        try:
            parsed = int(str(value).replace(",", "").strip())
            return parsed if parsed > 0 else None
        except Exception:
            return None

    def parse_positive_float(self, value):
        try:
            parsed = float(str(value).replace(",", "").strip())
            return parsed if parsed > 0 else None
        except Exception:
            return None

    def is_quote_expired(self, quote):
        validity = str((quote or {}).get("validity_date") or "").strip()
        if not validity or validity in {"N/A", "Unknown"}:
            return False
        try:
            import datetime
            return datetime.datetime.strptime(validity[:10], "%Y-%m-%d").date() < datetime.date.today()
        except Exception:
            return False

    def validate_rfq_ready(self):
        issues = []
        warnings = []
        product_name = self.rfq_name_entry.get().strip() or self.rfq_product_cb.get().strip()
        specs = self.rfq_specs_text.get("1.0", tk.END).strip()

        if not product_name or product_name == "Custom":
            issues.append("Product name is required.")
        if not self.parse_positive_int(self.rfq_qty_entry.get()):
            issues.append("Target quantity must be a positive whole number.")
        if not self.rfq_term_cb.get().strip():
            issues.append("Price terms are required.")
        if not self.rfq_lead_entry.get().strip():
            issues.append("Target lead time is required.")
        if not self.rfq_payment_entry.get().strip():
            issues.append("Payment terms are required.")
        if not specs:
            issues.append("Product specifications are required.")
        elif len(specs) < 40:
            warnings.append("Specifications are very short; add material, size, packing, and quality requirements for a stronger RFQ.")
        return issues, warnings

    def validate_po_ready(self):
        issues = []
        warnings = []
        supplier = self.po_supplier_cb.get().strip()
        product = self.po_product_cb.get().strip()
        quote = getattr(self, "po_active_quote", None)

        if not quote or quote.get("review_status") != "Approved":
            issues.append("An approved quote must be selected before issuing a PO.")
        elif self.is_quote_expired(quote):
            issues.append(f"The selected approved quote expired on {quote.get('validity_date')}; refresh or re-approve before issuing.")

        if not supplier or supplier in {"Select Supplier", "No approved quotes"}:
            issues.append("Supplier is required.")
        if not product or product in {"Select Product", "Approve quotes first"}:
            issues.append("Product is required.")
        if not self.po_number_entry.get().strip():
            issues.append("PO number is required.")
        if not self.parse_positive_int(self.po_qty_entry.get()):
            issues.append("Order quantity must be a positive whole number.")
        if not self.parse_positive_float(getattr(self, "po_active_price", 0.0)):
            issues.append("Unit cost must be greater than zero.")
        if not self.po_payment_entry.get().strip():
            issues.append("Payment terms are required.")
        if not self.po_address_entry.get().strip():
            issues.append("Delivery address is required.")

        if quote:
            risk = str(quote.get("risk") or quote.get("sourcing_risk") or "").lower()
            if "high" in risk:
                warnings.append("Selected quote has a high-risk alert; confirm commercial approval before issuing.")
            for field, label in [("moq", "MOQ"), ("lead_time", "lead time"), ("packing", "packing")]:
                value = str(quote.get(field) or "").strip()
                if not value or value in {"N/A", "Unknown"}:
                    warnings.append(f"Selected quote is missing {label}; confirm with supplier before shipment commitment.")
        return issues, warnings

    def format_readiness_text(self, title, issues, warnings):
        if not issues and not warnings:
            return f"{title}\nReady for controlled workflow action."
        lines = [title]
        if issues:
            lines.append("\nRequired before proceeding:")
            lines.extend([f"- {issue}" for issue in issues])
        if warnings:
            lines.append("\nCommercial warnings:")
            lines.extend([f"- {warning}" for warning in warnings])
        return "\n".join(lines)

    def set_readiness_box(self, box, title, issues, warnings):
        if not box:
            return
        box.configure(state="normal")
        box.delete("1.0", tk.END)
        box.insert("1.0", self.format_readiness_text(title, issues, warnings))
        box.configure(state="disabled")

    def update_rfq_readiness_panel(self):
        if not hasattr(self, "rfq_readiness_box"):
            return
        issues, warnings = self.validate_rfq_ready()
        self.set_readiness_box(self.rfq_readiness_box, "RFQ Readiness", issues, warnings)

    def update_po_readiness_panel(self):
        if not hasattr(self, "po_readiness_box"):
            return
        issues, warnings = self.validate_po_ready()
        self.set_readiness_box(self.po_readiness_box, "PO Issuance Gate", issues, warnings)

    def start_contact_extraction_thread(self):
        if not self.api_key:
            messagebox.showerror("API Key Missing", "Please enter and save your Gemini API Key first.")
            return
        if not self.selected_folder:
            messagebox.showerror("Folder Missing", "Please select a folder containing the quotes first.")
            return
            
        self.btn_sync_dir.configure(state="disabled")
        self.btn_extract_contacts.configure(state="disabled", text="Extracting...")
        
        threading.Thread(target=self.run_contact_extraction, daemon=True).start()

    def run_contact_extraction(self):
        try:
            prompt = """
            You are an expert contact card extractor. Look at this document or image and extract:
            1. The supplier company name.
            2. All contact details (phone numbers, email addresses, WhatsApp numbers, websites, address) formatted as a clean, concise, multi-line text block.
            
            Return the results strictly matching this JSON schema:
            {
              "supplier_name": "Name of the supplier company",
              "contact_info": "Phone: ...\\nEmail: ...\\nWhatsApp: ...\\nWebsite: ..."
            }
            Output ONLY raw JSON. No markdown backticks.
            """
            
            # Read files in selected folder
            valid_exts = {".pdf", ".png", ".jpg", ".jpeg", ".txt"}
            files_to_process = []
            if os.path.exists(self.selected_folder):
                for file in os.listdir(self.selected_folder):
                    ext = os.path.splitext(file)[1].lower()
                    if ext in valid_exts:
                        files_to_process.append(os.path.join(self.selected_folder, file))
                        
            if not files_to_process:
                self.after(0, lambda: messagebox.showinfo("No Files", "No printable quote files found in the selected folder."))
                self.after(0, self.finish_contact_extraction)
                return
                
            for path in files_to_process:
                try:
                    import mimetypes
                    mime_type, _ = mimetypes.guess_type(path)
                    if not mime_type:
                        ext = os.path.splitext(path)[1].lower()
                        if ext == ".pdf":
                            mime_type = "application/pdf"
                        elif ext == ".png":
                            mime_type = "image/png"
                        elif ext in {".jpg", ".jpeg"}:
                            mime_type = "image/jpeg"
                        elif ext == ".txt":
                            mime_type = "text/plain"
                        else:
                            mime_type = "application/octet-stream"

                    with open(path, "rb") as f:
                        file_bytes = f.read()

                    response_text = self.generate_with_fallback(
                        [{"mime_type": mime_type, "data": file_bytes}],
                        prompt
                    )
                    
                    result = json.loads(response_text)
                    supplier = result.get("supplier_name") or "Unknown"
                    contact_info = result.get("contact_info") or "N/A"
                    
                    if supplier != "Unknown" and contact_info != "N/A":
                        conn = sqlite3.connect(DB_FILE)
                        c = conn.cursor()
                        c.execute("INSERT OR REPLACE INTO supplier_contacts (supplier, contact_info, source_file) VALUES (?, ?, ?)", (supplier, contact_info, os.path.basename(path)))
                        conn.commit()
                        conn.close()
                        
                        # Refresh UI live
                        self.after(0, self.load_supplier_directory)
                        
                    import time
                    time.sleep(2.0)
                    
                except Exception as e:
                    print(f"Error extracting contact from {os.path.basename(path)}: {e}")
                    
            self.after(0, lambda: messagebox.showinfo("Success", "Finished extracting supplier contacts from files!"))
        except Exception as e:
            self.after(0, lambda e_err=e: messagebox.showerror("Error", f"An error occurred: {e_err}"))
        finally:
            self.after(0, self.finish_contact_extraction)

    def finish_contact_extraction(self):
        self.btn_sync_dir.configure(state="normal")
        self.btn_extract_contacts.configure(state="normal", text="🔍 Auto-Extract Contacts")

    def start_one_time_risk_extraction_thread(self):
        if not self.api_key:
            messagebox.showerror("API Key Missing", "Please enter and save your Gemini API Key first.")
            return
        if not self.selected_folder:
            messagebox.showerror("Folder Missing", "Please select a folder containing the quotes first.")
            return
            
        self.btn_start.configure(state="disabled")
        
        threading.Thread(target=self.run_one_time_risk_extraction, daemon=True).start()

    def run_one_time_risk_extraction(self):
        try:
            prompt = """
            You are an expert quote validity and risk auditor. Look at this quotation document and extract:
            1. The quotation validity date or expiration date of this quote. Return in YYYY-MM-DD format if possible.
               If not explicitly mentioned, look for quote date and validity period (e.g. quote dated 2026-05-10, valid for 30 days -> return 2026-06-09). If not found, return "N/A".
            2. Sourcing risk details (e.g., payment term risk, extremely long lead times, shipping port risks, or lack of certification details). Keep it short (under 15 words).
            
            Return the results strictly matching this JSON schema:
            {
              "validity_date": "YYYY-MM-DD or N/A",
              "sourcing_risk": "Low risk / Medium risk: ... / High risk: ..."
            }
            Output ONLY raw JSON. No markdown backticks.
            """
            
            # Get list of unique files already processed in quotes table
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT DISTINCT filename FROM extracted_quotes WHERE filename != 'Manually Added' AND filename NOT LIKE 'Chat (%'")
            files = [row[0] for row in c.fetchall()]
            conn.close()
            
            if not files:
                self.after(0, lambda: messagebox.showinfo("No Records", "No processed quote files found in the database to audit."))
                self.after(0, self.finish_one_time_risk_extraction)
                return
                
            audited_count = 0
            for filename in files:
                full_path = os.path.join(self.selected_folder, filename)
                if not os.path.exists(full_path):
                    continue
                    
                try:
                    import mimetypes
                    mime_type, _ = mimetypes.guess_type(full_path)
                    if not mime_type:
                        ext = os.path.splitext(full_path)[1].lower()
                        if ext == ".pdf":
                            mime_type = "application/pdf"
                        elif ext == ".png":
                            mime_type = "image/png"
                        elif ext in {".jpg", ".jpeg"}:
                            mime_type = "image/jpeg"
                        elif ext == ".txt":
                            mime_type = "text/plain"
                        else:
                            mime_type = "application/octet-stream"

                    with open(full_path, "rb") as f:
                        file_bytes = f.read()

                    response_text = self.generate_with_fallback(
                        [{"mime_type": mime_type, "data": file_bytes}],
                        prompt
                    )
                    
                    result = json.loads(response_text)
                    validity_date = result.get("validity_date") or "N/A"
                    sourcing_risk = result.get("sourcing_risk") or "N/A"
                    
                    conn = sqlite3.connect(DB_FILE)
                    c = conn.cursor()
                    c.execute("""
                        UPDATE extracted_quotes 
                        SET validity_date = ?, sourcing_risk = ?
                        WHERE filename = ?
                    """, (validity_date, sourcing_risk, filename))
                    conn.commit()
                    conn.close()
                    
                    audited_count += 1
                    
                    import time
                    time.sleep(4.0)
                    
                except Exception as e:
                    print(f"Error auditing {filename}: {e}")
                    
            self.after(0, lambda count=audited_count: messagebox.showinfo("Audit Complete", f"Successfully audited existing files!\nUpdated validity dates and risk alerts for {count} quotes in the grid without changing any other price or spec details."))
            
        except Exception as e:
            self.after(0, lambda e_err=e: messagebox.showerror("Error", f"Audit failed: {e_err}"))
        finally:
            self.after(0, self.finish_one_time_risk_extraction)

    def finish_one_time_risk_extraction(self):
        self.btn_start.configure(state="normal")
        self.load_all_quotes_from_db()

    # --- Supplier Directory rendering logic ---
    def load_supplier_directory(self):
        for widget in self.directory_scroll_frame.winfo_children():
            widget.destroy()
            
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT supplier, contact_info, source_file FROM supplier_contacts")
        rows = c.fetchall()
        conn.close()
        
        if not rows:
            lbl = ctk.CTkLabel(self.directory_scroll_frame, text="No suppliers found in directory.\nProcess some files or paste chats to populate cards!", text_color="grey")
            lbl.pack(pady=50)
            return
            
        for supplier, contact_info, source_file in rows:
            card = ctk.CTkFrame(self.directory_scroll_frame, fg_color=self.THEME["surface"], corner_radius=8, border_width=1, border_color=self.THEME["border"])
            card.pack(fill="x", padx=12, pady=7)
            
            title = ctk.CTkLabel(card, text=supplier, font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"])
            title.pack(anchor="w", padx=16, pady=(12, 4))
            
            info_lbl = ctk.CTkLabel(card, text=contact_info, text_color=self.THEME["muted"], justify="left")
            info_lbl.pack(anchor="w", padx=16, pady=(2, 2))
            
            src_val = source_file or "N/A"
            src_lbl = ctk.CTkLabel(card, text=f"Source: {src_val}", text_color=self.THEME["muted"], font=ctk.CTkFont(size=10, slant="italic"))
            src_lbl.pack(anchor="w", padx=16, pady=(2, 10))
            
            btn_frame = ctk.CTkFrame(card, fg_color="transparent")
            btn_frame.pack(fill="x", padx=15, pady=(0, 10), anchor="e")
            
            email = self.extract_email_from_text(contact_info)
            phone = self.extract_phone_from_text(contact_info)
            
            if email:
                btn_email = ctk.CTkButton(btn_frame, text="📧 Email", width=70, height=22, command=lambda e=email: self.open_email(e))
                btn_email.configure(text="Email", fg_color=self.THEME["primary"], hover_color=self.THEME["primary_hover"], corner_radius=6, height=24)
                btn_email.pack(side="left", padx=2)
            if phone:
                btn_wa = ctk.CTkButton(btn_frame, text="💬 WhatsApp", width=80, height=22, fg_color="#1f7d44", hover_color="#15592e", command=lambda p=phone: self.open_whatsapp(p))
                btn_wa.configure(text="WhatsApp", fg_color=self.THEME["success"], hover_color=self.THEME["success_hover"], corner_radius=6, height=24)
                btn_wa.pack(side="left", padx=2)
                
            btn_edit = ctk.CTkButton(btn_frame, text="✏ Edit Info", width=70, height=22, command=lambda s=supplier, c_info=contact_info: self.edit_supplier_contact(s, c_info))
            btn_edit.configure(text="Edit Info", fg_color="#3B82C4", hover_color="#2C6DA6", corner_radius=6, height=24)
            btn_edit.pack(side="right", padx=2)

        self.apply_legacy_light_polish(self.directory_scroll_frame)

    def extract_email_from_text(self, text):
        import re
        match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', text)
        return match.group(0) if match else None

    def extract_phone_from_text(self, text):
        import re
        cleaned_text = re.sub(r'[^\w\s\+]', '', text)
        match = re.search(r'\+?\d{8,15}', cleaned_text.replace(" ", ""))
        return match.group(0) if match else None

    def open_email(self, email):
        try:
            webbrowser.open(f"mailto:{email}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open mail client:\n{e}")

    def open_whatsapp(self, phone):
        cleaned = phone.replace("+", "").replace(" ", "").strip()
        try:
            webbrowser.open(f"https://wa.me/{cleaned}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open web browser:\n{e}")

    def edit_supplier_contact(self, supplier, current_info):
        edit_win = ctk.CTkToplevel(self)
        edit_win.title(f"Edit Info: {supplier}")
        edit_win.geometry("400x250")
        edit_win.resizable(False, False)
        edit_win.attributes("-topmost", True)
        
        lbl = ctk.CTkLabel(edit_win, text=f"Edit Contact Details for {supplier}:", font=ctk.CTkFont(size=13, weight="bold"))
        lbl.pack(pady=10, padx=15, anchor="w")
        
        textbox = ctk.CTkTextbox(edit_win, height=100)
        textbox.pack(fill="both", expand=True, padx=15, pady=5)
        textbox.insert("1.0", current_info)
        
        def save_contact():
            new_info = textbox.get("1.0", tk.END).strip()
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO supplier_contacts (supplier, contact_info) VALUES (?, ?)", (supplier, new_info))
            conn.commit()
            conn.close()
            
            self.load_supplier_directory()
            edit_win.destroy()
            
        btn_save = ctk.CTkButton(edit_win, text="Save Contact Details", fg_color="#1f7d44", hover_color="#15592e", command=save_contact)
        btn_save.pack(pady=15, padx=15, fill="x")

    # --- Matplotlib Visual Charts Dashboard ---
    def clean_supplier_name(self, name):
        if not name:
            return "Unknown"
        name = name.strip().title()
        # Clean common corporate suffix tokens
        for sfx in ["Nonwoven", "Protective", "Products", "Co.", "Ltd", "Co.,Ltd", "Corporation", "Factory", "Manufacturing", "Industry", "Limited", "Ltd.", "Co", "Plastic"]:
            name = name.replace(sfx.title(), "")
        words = name.split()
        return " ".join(words[:2]).strip()

    def clean_product_name(self, product):
        if not product:
            return "Product"
        p = product.split("(")[0].strip()
        p = p.split("-")[0].strip()
        words = p.split()
        res = " ".join(words[:3]).strip()
        if len(res) > 16:
            res = res[:14] + "..."
        return res

    def update_chart_dropdown(self):
        products = set()
        for r in self.extracted_data:
            prod = (r.get("product") or "").strip().title()
            if prod:
                products.add(prod)
                
        sorted_prods = ["All"] + sorted(list(products))
        
        # 1. Update charts dropdown
        current_chart = self.chart_category_cb.get()
        self.chart_category_cb.configure(values=sorted_prods)
        if current_chart in sorted_prods:
            self.chart_category_cb.set(current_chart)
        else:
            self.chart_category_cb.set("All")
            
        # 2. Update insights dropdown
        if hasattr(self, 'insights_category_cb'):
            current_insight = self.insights_category_cb.get()
            self.insights_category_cb.configure(values=sorted_prods)
            if current_insight in sorted_prods:
                self.insights_category_cb.set(current_insight)
            else:
                self.insights_category_cb.set("All")

        # 3. Update scorecard dropdown
        if hasattr(self, 'scorecard_category_cb'):
            current_scorecard = self.scorecard_category_cb.get()
            self.scorecard_category_cb.configure(values=sorted_prods)
            if current_scorecard in sorted_prods:
                self.scorecard_category_cb.set(current_scorecard)
            else:
                self.scorecard_category_cb.set("All")

    def draw_chart(self):
        for widget in self.chart_display_frame.winfo_children():
            widget.destroy()
            
        category = self.chart_category_cb.get().lower()
        
        # Get active currency conversion specs for visual charts
        currency_choice = self.currency_cb.get()
        factor = 1.0
        symbol = "$"
        currency_name = "USD"
        if "CNY" in currency_choice:
            factor = self.exchange_rates["CNY"]
            symbol = "¥"
            currency_name = "CNY"
        elif "EUR" in currency_choice:
            factor = self.exchange_rates["EUR"]
            symbol = "€"
            currency_name = "EUR"

        data = []
        for r in self.extracted_data:
            prod_name = (r.get("product") or "").strip().lower()
            
            if category != "all" and prod_name != category:
                continue
                
            price = r.get("price")
            try:
                price = float(price)
            except (ValueError, TypeError):
                continue
                
            if price > 0:
                data.append({
                    "supplier": r.get("supplier", "Unknown"),
                    "product": r.get("product", "Product"),
                    "price": price * factor
                })
                
        if not data:
            lbl = ctk.CTkLabel(self.chart_display_frame, text="No price data available for category to display charts.", text_color="grey")
            lbl.pack(pady=100)
            return
            
        df = pd.DataFrame(data)
        # Sort and select Top 12 cheapest quotes to prevent overcrowding
        df = df.sort_values("price", ascending=True)
        if len(df) > 12:
            df = df.head(12)

        df["label"] = df["supplier"].apply(self.clean_supplier_name) + "\n(" + df["product"].apply(self.clean_product_name) + ")"
        
        fig, ax = plt.subplots(figsize=(8, 4.5), dpi=100)
        fig.patch.set_facecolor(self.THEME["surface"])
        ax.set_facecolor(self.THEME["surface"])
        
        bars = ax.bar(df["label"], df["price"], color="#1f538d", edgecolor="#1a4473")
        
        ax.set_title(f"Price Comparison: {category.capitalize()}", color=self.THEME["text"], fontsize=12, pad=15)
        ax.set_ylabel(f"Unit Price ({currency_name})", color=self.THEME["muted"], fontsize=10)
        ax.tick_params(colors=self.THEME["muted"], labelsize=8)
        
        plt.xticks(rotation=15, ha="right")
        ax.yaxis.grid(True, linestyle="--", alpha=0.9, color=self.THEME["border"])
        ax.set_axisbelow(True)
        
        for spine in ax.spines.values():
            spine.set_edgecolor(self.THEME["border"])
            
        for bar in bars:
            yval = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2.0, yval, f"{symbol}{yval:.5f}", ha='center', va='bottom', color=self.THEME["text"], fontsize=7.5)
            
        plt.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.chart_display_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    # --- Config Management ---
    def load_config(self):
        self.api_provider = "Google Gemini"
        self.custom_base_url = "https://api.openai.com/v1"
        self.custom_model = "gpt-5.6-luna"
        
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f:
                    config = json.load(f)
                    self.api_key = config.get("api_key", "")
                    self.selected_folder = config.get("last_folder", "")
                    self.api_provider = config.get("api_provider", "Google Gemini")
                    self.custom_base_url = config.get("custom_base_url", "https://api.openai.com/v1")
                    self.custom_model = config.get("custom_model", "gpt-5.6-luna")
            except Exception:
                pass
        
        if hasattr(self, 'provider_cb'):
            self.provider_cb.set(self.api_provider)
            self.on_provider_changed(self.api_provider)
        if hasattr(self, 'api_entry') and self.api_key:
            self.api_entry.delete(0, tk.END)
            self.api_entry.insert(0, self.api_key)
        if hasattr(self, 'base_url_entry'):
            self.base_url_entry.delete(0, tk.END)
            self.base_url_entry.insert(0, self.custom_base_url)
        if hasattr(self, 'model_entry'):
            self.model_entry.delete(0, tk.END)
            self.model_entry.insert(0, self.custom_model)

        if self.selected_folder:
            self.folder_entry.delete(0, tk.END)
            self.folder_entry.insert(0, self.selected_folder)
            if os.path.exists(self.selected_folder):
                self.scan_folder()

    def save_config(self):
        config = {
            "api_key": self.api_key,
            "last_folder": self.selected_folder,
            "api_provider": self.provider_cb.get(),
            "custom_base_url": self.base_url_entry.get().strip(),
            "custom_model": self.model_entry.get().strip()
        }
        try:
            with open(CONFIG_FILE, "w") as f:
                json.dump(config, f)
        except Exception:
            pass

    def save_and_test_key(self):
        key = self.api_entry.get().strip()
        if not key:
            messagebox.showerror("Error", "Please enter a valid API Key")
            return
        
        self.api_key = key
        self.api_provider = self.provider_cb.get()
        self.custom_base_url = self.base_url_entry.get().strip()
        self.custom_model = self.model_entry.get().strip()
        self.save_config()

        threading.Thread(target=self.async_test_key, daemon=True).start()

    def async_test_key(self):
        try:
            self.generate_with_fallback([], "Ping", json_response=False)
            self.api_entry.configure(fg_color="#1f5a34")
            messagebox.showinfo("Success", f"{self.api_provider} API Key is valid and working!")
        except Exception as e:
            self.api_entry.configure(fg_color="#5a1f1f")
            messagebox.showerror("Error", f"Failed to connect to API:\n{e}")

    def on_provider_changed(self, choice):
        if choice == "Custom OpenAI/Luna":
            self.custom_api_frame.pack(fill="x", pady=5)
            self.api_lbl.configure(text="Custom API Key:")
        else:
            self.custom_api_frame.pack_forget()
            self.api_lbl.configure(text="Gemini API Key:")

    def show_page(self, page_name):
        # Hide all pages
        for name, frame in self.pages.items():
            frame.grid_forget()
            
        # Show selected page
        self.pages[page_name].grid(row=0, column=0, sticky="nsew")
        self.active_page = page_name
        if page_name == "Dashboard":
            self.update_dashboard_page()
        self.apply_legacy_light_polish(self.pages[page_name])
        
        # Highlight active sidebar button
        for name, btn in self.sidebar_buttons.items():
            if name == page_name:
                btn.configure(fg_color=self.THEME["primary"], text_color="white")
            else:
                btn.configure(fg_color="transparent", text_color="#D1D5DB")
                
        # Toggle document preview frame visibility based on page and user visibility preference
        if page_name in ["Sourcing Analysis", "Settings Directory"] and self.document_preview_visible:
            self.preview_frame.grid(row=0, column=2, sticky="nsew", padx=(0, 14), pady=14)
            self.grid_columnconfigure(2, minsize=360, weight=0)
        else:
            self.preview_frame.grid_forget()
            self.grid_columnconfigure(2, minsize=0, weight=0)

        # Show/Hide Toggle Buttons dynamically based on page context
        if page_name == "Sourcing Analysis":
            self.btn_toggle_files.pack(side="left", padx=5)
        else:
            self.btn_toggle_files.pack_forget()

        if page_name in ["Sourcing Analysis", "Settings Directory"]:
            self.btn_toggle_preview.pack(side="left", padx=5)
        else:
            self.btn_toggle_preview.pack_forget()

    # --- Folder Logic ---
    def select_folder(self):
        folder = filedialog.askdirectory(initialdir=self.selected_folder)
        if folder:
            self.selected_folder = folder
            self.folder_entry.delete(0, tk.END)
            self.folder_entry.insert(0, folder)
            self.folder_entry.configure(fg_color=None)
            self.save_config()
            self.scan_folder()

    def scan_folder(self):
        self.files_box_unsynced.delete(0, tk.END)
        self.files_box_synced.delete(0, tk.END)
        self.files_list = []
        
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT filename FROM processed_files")
        processed = {row[0] for row in c.fetchall()}
        conn.close()
        
        valid_exts = {".pdf", ".png", ".jpg", ".jpeg", ".txt", ".xlsx", ".xls"}
        if os.path.exists(self.selected_folder):
            for file in os.listdir(self.selected_folder):
                ext = os.path.splitext(file)[1].lower()
                if ext in valid_exts:
                    full_path = os.path.join(self.selected_folder, file)
                    self.files_list.append((file, full_path))
                    
                    if file in processed:
                        self.files_box_synced.insert(tk.END, f"✅ {file}")
                    else:
                        self.files_box_unsynced.insert(tk.END, f"⏳ {file}")
            
            # Enable Start button if there are any unsynced files
            has_unsynced = any(name not in processed for (name, path) in self.files_list)
            if has_unsynced:
                self.btn_start.configure(state="normal")
            else:
                self.btn_start.configure(state="disabled")

    # --- Extraction Engine ---
    def animate_spinner(self):
        if self.is_extracting:
            spinners = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
            current_char = spinners[self.spinner_idx % len(spinners)]
            self.spinner_idx += 1
            
            self.btn_start.configure(text=f"Extracting {current_char}")
            self.after(100, self.animate_spinner)

    def start_extraction_thread(self):
        if not self.api_key:
            messagebox.showerror("API Key Missing", "Please enter and save your Gemini API Key before starting.")
            return
        self.btn_start.configure(state="disabled")
        self.btn_select_folder.configure(state="disabled")
        
        self.is_extracting = True
        self.spinner_idx = 0
        self.animate_spinner()
        self.progress_bar.configure(mode="indeterminate")
        self.progress_bar.start()
        
        threading.Thread(target=self.run_extraction, daemon=True).start()

    def save_extracted_data_to_db(self, filename, data):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        
        try:
            c.execute("INSERT OR REPLACE INTO processed_files (filename) VALUES (?)", (filename,))
        except Exception as e:
            print(f"Error marking file in DB: {e}")

        supplier = data.get("supplier_name") or "Unknown"
        term = data.get("price_term") or "N/A"
        payment = data.get("payment_terms") or "N/A"
        lead_time = data.get("delivery_lead_time") or "N/A"
        
        validity_date = data.get("validity_date") or "N/A"
        sourcing_risk = data.get("sourcing_risk") or "N/A"
        
        # Save contact details to db
        contact_info = data.get("contact_info") or "N/A"
        if supplier != "Unknown":
            c.execute("INSERT OR REPLACE INTO supplier_contacts (supplier, contact_info, source_file) VALUES (?, ?, ?)", (supplier, contact_info, filename))

        quotes = data.get("quotes") or []
        for quote in quotes:
            c.execute("""
                INSERT INTO extracted_quotes (filename, supplier, product, spec, color, elastic, price, unit, moq, packing, term, lead_time, validity_date, sourcing_risk, review_status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                filename,
                supplier,
                quote.get("product_name") or "N/A",
                quote.get("specifications") or "N/A",
                quote.get("color") or "N/A",
                quote.get("elastic_type") or "N/A",
                quote.get("unit_price_usd") or 0.0,
                quote.get("price_unit") or "piece",
                quote.get("moq") or "N/A",
                quote.get("packing_details") or "N/A",
                term,
                lead_time,
                validity_date,
                sourcing_risk,
                "Needs Review"
            ))
        conn.commit()
        conn.close()

    def run_extraction(self):
        prompt = """
        You are an expert procurement assistant. Extract all quotation details from the provided document or image.
        
        Return the results strictly matching this JSON schema:
        {
          "supplier_name": "Name of the supplier (e.g. Hubei Ruichen, Xiantao Topmed, etc.)",
          "contact_info": "email, phone number, website if visible",
          "price_term": "FOB Wuhan, FOB Shanghai, EXW, etc. (default to FOB if not mentioned but specified near price)",
          "payment_terms": "e.g. 30% deposit, 70% balance",
          "delivery_lead_time": "e.g. 30 days",
          "validity_date": "YYYY-MM-DD format (extract validity or expiration date of this quote. E.g. Valid until 2026-10-31 or quote date + validity duration)",
          "sourcing_risk": "e.g. 'Low: standard terms', 'Medium: long lead time (45 days)', 'High: 50% upfront prepayment required', etc.",
          "quotes": [
            {
              "product_name": "Surgical Cap, Clip Cap, Apron, Poncho, etc.",
              "specifications": "size, weight, gsm, elastic specs (e.g. 10g, 21 inch)",
              "color": "White, Blue, Black, etc.",
              "elastic_type": "Single, Double, or None",
              "unit_price_usd": 0.0059, // float unit price. If quoted per 1000pcs or bag, calculate unit price (divide by quantity). If quoted in CNY/RMB, convert to USD (use 1 USD = 7.2 CNY) and note in specifications.
              "price_unit": "piece", // strictly piece
              "packing_details": "e.g. 100pcs/bag, 1000pcs/carton",
              "moq": "e.g. 100,000 pcs"
            }
          ]
        }

        Make sure:
        - If a value is missing or not applicable, return null or empty string.
        - Calculate the single piece unit price in USD even if they quote per bag, carton, or per 1000pcs.
        - Output ONLY raw JSON. No markdown backticks.
        """

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT filename FROM processed_files")
        processed = {row[0] for row in c.fetchall()}
        conn.close()

        unsynced_files = [(name, path) for (name, path) in self.files_list if name not in processed]
        if not unsynced_files:
            self.is_extracting = False
            self.progress_bar.stop()
            self.progress_bar.configure(mode="determinate")
            self.progress_bar.set(1.0)
            self.btn_start.configure(state="disabled", text="Start Extraction")
            messagebox.showinfo("Extraction Info", "All files in the queue are already synced!")
            return

        total_files = len(unsynced_files)
        for i, (name, path) in enumerate(unsynced_files):
            idx = -1
            for box_idx in range(self.files_box_unsynced.size()):
                if name in self.files_box_unsynced.get(box_idx):
                    idx = box_idx
                    break
            
            if idx != -1:
                self.files_box_unsynced.delete(idx)
                self.files_box_unsynced.insert(idx, f"🔄 {name}")
                
            self.progress_bar.set(i / total_files)
            
            success = False
            retries = 3
            while not success and retries > 0:
                try:
                    ext = os.path.splitext(path)[1].lower()
                    if ext in {".xlsx", ".xls"}:
                        import pandas as pd
                        df = pd.read_excel(path)
                        csv_text = df.to_csv(index=False)
                        content_list = [f"EXCEL SPREADSHEET CONTENT (CSV format):\n\n{csv_text}"]
                    else:
                        import mimetypes
                        mime_type, _ = mimetypes.guess_type(path)
                        if not mime_type:
                            if ext == ".pdf":
                                mime_type = "application/pdf"
                            elif ext == ".png":
                                mime_type = "image/png"
                            elif ext in {".jpg", ".jpeg"}:
                                mime_type = "image/jpeg"
                            elif ext == ".txt":
                                mime_type = "text/plain"
                            else:
                                mime_type = "application/octet-stream"

                        with open(path, "rb") as f:
                            file_bytes = f.read()
                        content_list = [{"mime_type": mime_type, "data": file_bytes}]

                    response_text = self.generate_with_fallback(
                        content_list,
                        prompt
                    )
                    
                    result = json.loads(response_text)
                    self.save_extracted_data_to_db(name, result)
                    
                    self.load_all_quotes_from_db()
                    
                    # Move UI item
                    self.after(0, lambda n=name: self.move_file_to_synced(n))
                    success = True
                    
                    import time
                    time.sleep(6.0)
                    
                except Exception as e:
                    if "ResourceExhausted" in str(type(e)) or "429" in str(e):
                        print(f"Rate limit hit for {name}. Waiting 15s before retry... ({retries} left)")
                        for box_idx in range(self.files_box_unsynced.size()):
                            if name in self.files_box_unsynced.get(box_idx):
                                self.files_box_unsynced.delete(box_idx)
                                self.files_box_unsynced.insert(box_idx, f"⏳ {name} (Rate Limit Wait)")
                                break
                        import time
                        time.sleep(15.0)
                        retries -= 1
                    else:
                        for box_idx in range(self.files_box_unsynced.size()):
                            if name in self.files_box_unsynced.get(box_idx):
                                self.files_box_unsynced.delete(box_idx)
                                self.files_box_unsynced.insert(box_idx, f"❌ {name} (Error)")
                                break
                        print(f"Error processing {name}: {e}")
                        break
            
            if not success and retries == 0:
                for box_idx in range(self.files_box_unsynced.size()):
                    if name in self.files_box_unsynced.get(box_idx):
                        self.files_box_unsynced.delete(box_idx)
                        self.files_box_unsynced.insert(box_idx, f"❌ {name} (Limit Exceeded)")
                        break
                
        self.is_extracting = False
        self.progress_bar.stop()
        self.progress_bar.configure(mode="determinate")
        self.progress_bar.set(1.0)
        self.btn_start.configure(state="disabled", text="Start Extraction")
        self.btn_select_folder.configure(state="normal")
        messagebox.showinfo("Processing Completed", "Finished processing all files in the queue!")

    # --- Chat Extraction Engine ---
    def open_paste_chat_window(self):
        if not self.api_key:
            messagebox.showerror("API Key Missing", "Please enter and save your Gemini API Key first.")
            return

        chat_win = ctk.CTkToplevel(self)
        chat_win.title("Extract Quotes from Chat Text")
        chat_win.geometry("500x550")
        chat_win.resizable(False, False)
        chat_win.attributes("-topmost", True)

        lbl = ctk.CTkLabel(chat_win, text="Paste WhatsApp or Made-in-China Chat Text:", font=ctk.CTkFont(size=14, weight="bold"))
        lbl.pack(pady=10, padx=15, anchor="w")

        textbox = ctk.CTkTextbox(chat_win, height=350)
        textbox.pack(fill="both", expand=True, padx=15, pady=5)
        textbox.insert("1.0", "Paste chat content here...\n\nExample:\nSupplier: We can offer Surgical Cap 21 inch double elastic blue color at $0.0055/pc FOB Wuhan.\nBuyer: What is the MOQ?\nSupplier: 100,000pcs.")

        self.chat_is_extracting = False
        self.chat_spinner_idx = 0

        def animate_chat_spinner(btn):
            if self.chat_is_extracting:
                spinners = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
                char = spinners[self.chat_spinner_idx % len(spinners)]
                self.chat_spinner_idx += 1
                btn.configure(text=f"Extracting {char}")
                chat_win.after(100, lambda: animate_chat_spinner(btn))

        def process_chat():
            raw_text = textbox.get("1.0", tk.END).strip()
            if not raw_text or raw_text.startswith("Paste chat content"):
                messagebox.showwarning("Empty Content", "Please paste chat text before extracting.", parent=chat_win)
                return

            btn_extract.configure(state="disabled")
            self.chat_is_extracting = True
            self.chat_spinner_idx = 0
            animate_chat_spinner(btn_extract)

            threading.Thread(target=self.run_chat_extraction, args=(raw_text, chat_win), daemon=True).start()

        btn_extract = ctk.CTkButton(chat_win, text="Extract & Add Quotes", fg_color="#1f7d44", hover_color="#15592e", command=process_chat)
        btn_extract.pack(pady=15, padx=15, fill="x")

    def run_chat_extraction(self, text, window):
        try:
            prompt = """
            You are an expert procurement assistant. Extract all quotation details from the provided raw chat text conversation (from WhatsApp or Made-in-China).
            Analyze the chat and extract the final or best quoted prices, specifications, and terms.
            
            Return the results strictly matching this JSON schema:
            {
              "supplier_name": "Name of the supplier if mentioned in the chat text (otherwise set as Unknown)",
              "contact_info": "email, phone number, website if visible in the chat text",
              "price_term": "FOB Wuhan, FOB Shanghai, EXW, etc. (default to FOB if not mentioned but specified near price)",
              "payment_terms": "e.g. 30% deposit, 70% balance",
              "delivery_lead_time": "e.g. 30 days",
              "validity_date": "YYYY-MM-DD format (extract validity or expiration date of this quote if mentioned, otherwise null)",
              "sourcing_risk": "e.g. 'Low: standard terms', 'Medium: long lead time', 'High: 50% upfront prepayment required', etc.",
              "quotes": [
                {
                  "product_name": "Surgical Cap, Clip Cap, Apron, Poncho, etc.",
                  "specifications": "size, weight, gsm, elastic specs (e.g. 10g, 21 inch)",
                  "color": "White, Blue, Black, etc.",
                  "elastic_type": "Single, Double, or None",
                  "unit_price_usd": 0.0059, // float unit price. If quoted per 1000pcs or bag, calculate unit price (divide by quantity). If quoted in CNY/RMB, convert to USD (use 1 USD = 7.2 CNY).
                  "price_unit": "piece", // strictly piece
                  "packing_details": "e.g. 100pcs/bag, 1000pcs/carton",
                  "moq": "e.g. 100,000 pcs"
                }
              ]
            }
            Output ONLY raw JSON. No markdown backticks.
            """

            response_text = self.generate_with_fallback(
                [f"CHAT CONVERSATION TEXT:\n\n{text}"],
                prompt
            )

            result = json.loads(response_text)

            import datetime
            now_str = datetime.datetime.now().strftime("%H:%M:%S")
            source_name = f"Chat ({now_str})"

            self.save_extracted_data_to_db(source_name, result)
            self.after(0, lambda: self.finish_chat_extraction(source_name, window))

        except Exception as e:
            self.chat_is_extracting = False
            self.after(0, lambda: messagebox.showerror("Extraction Error", f"Failed to extract chat quotes:\n{e}", parent=window))
            self.after(0, lambda: window.destroy())

    def finish_chat_extraction(self, source_name, window):
        self.chat_is_extracting = False
        self.load_all_quotes_from_db()
        messagebox.showinfo("Extraction Success", f"Successfully extracted and saved quotes from chat text under source: {source_name}!")
        window.destroy()

    def send_chat_message(self):
        msg = self.chat_entry.get().strip()
        if not msg:
            return
            
        self.chat_entry.delete(0, tk.END)
        
        # Enable textbox to insert user message
        self.chat_log.configure(state="normal")
        self.chat_log.insert(tk.END, f"\nYou: {msg}\n")
        self.chat_log.see(tk.END)
        self.chat_log.configure(state="disabled")
        
        # Save user message to database
        self.save_chat_message_to_db("You", msg)
        
        # Disable inputs
        self.btn_chat_send.configure(state="disabled")
        self.chat_entry.configure(state="disabled")
        
        # Start AI thread
        threading.Thread(target=self.run_chat_ai, args=(msg,), daemon=True).start()

    def run_chat_ai(self, user_msg):
        try:
            if not self.api_key:
                self.append_ai_response("Error: Please save your Gemini API key on the left panel first.")
                return
                
            # Format context from comparison table
            quotes_context = ""
            if self.extracted_data:
                quotes_context = json.dumps(self.extracted_data, indent=2)
            else:
                quotes_context = "The database is currently empty. No supplier quotes have been loaded."
                
            system_prompt = f"""
            You are an expert AI Procurement Assistant helping the user analyze their supplier quotations.
            Here is the current database of extracted supplier quotes:
            {quotes_context}
            
            Guidelines:
            - Provide direct, concise, and helpful answers.
            - Help compare prices, lead times, payment terms, and MOQs.
            - If asked to write a negotiation email, write a professional email tailored to the chosen supplier.
            - Keep answers friendly, highly professional, and focused on the data.
            """
            
            ai_response = self.generate_with_fallback(
                [],
                f"{system_prompt}\n\nUser Message: {user_msg}",
                json_response=False
            )
            
            self.append_ai_response(ai_response)
        except Exception as e:
            self.append_ai_response(f"Error generating response: {e}")

    def append_ai_response(self, response_text):
        self.after(0, lambda: self.finish_ai_response(response_text))

    def finish_ai_response(self, text):
        self.chat_log.configure(state="normal")
        self.chat_log.insert(tk.END, f"\nAI: {text}\n")
        self.chat_log.see(tk.END)
        self.chat_log.configure(state="disabled")
        
        # Save response message to SQLite
        self.save_chat_message_to_db("AI", text)
        
        self.btn_chat_send.configure(state="normal")
        self.chat_entry.configure(state="normal")
        self.chat_entry.focus()

    # --- Manual Edits logic ---
    def add_empty_row(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT INTO extracted_quotes (filename, supplier, product, spec, color, elastic, price, unit, moq, packing, term, lead_time, review_status, review_notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Manually Added", 
            "New Supplier", 
            "Product", 
            "Size / Gsm", 
            "Color", 
            "Single/Double", 
            0.0, 
            "piece", 
            "N/A", 
            "N/A", 
            "EXW/FOB", 
            "N/A",
            "Needs Review",
            "Manually added quote awaiting verification."
        ))
        conn.commit()
        last_id = c.lastrowid
        conn.close()
        self.log_quote_audit(last_id, "Manual Quote Added", None, "Needs Review", "Manual quote created from the comparison grid.")
        
        self.load_all_quotes_from_db()
        
        for item in self.tree.get_children():
            vals = self.tree.item(item, "values")
            if int(vals[0]) == last_id:
                self.tree.selection_set(item)
                self.edit_selected_row()
                break

    def delete_selected_row(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Row", "Please select a row to delete first.")
            return
            
        ans = messagebox.askyesno("Delete", "Are you sure you want to delete the selected row?")
        if ans:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            for item in sel:
                vals = self.tree.item(item, "values")
                q_id = int(vals[0])
                current = next((r for r in self.extracted_data if r.get("id") == q_id), {})
                c.execute("DELETE FROM extracted_quotes WHERE id = ?", (q_id,))
                c.execute("""
                    INSERT INTO quote_audit_log (quote_id, action, previous_status, new_status, note)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    q_id,
                    "Quote Deleted",
                    current.get("review_status"),
                    None,
                    f"Deleted quote for {current.get('supplier', 'Unknown')} / {current.get('product', 'Unknown')}"
                ))
            conn.commit()
            conn.close()
            self.load_all_quotes_from_db()

    def edit_selected_row(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Row", "Please select a row to edit first.")
            return
            
        item = sel[0]
        vals = self.tree.item(item, "values")
        q_id = int(vals[0])
        
        row_dict = None
        for r in self.extracted_data:
            if r["id"] == q_id:
                row_dict = r
                break
                
        if not row_dict:
            return
            
        edit_win = ctk.CTkToplevel(self)
        edit_win.title(f"Edit Quote Details (ID: {q_id})")
        edit_win.geometry("450x650")
        edit_win.resizable(False, False)
        edit_win.attributes("-topmost", True)
        
        labels = [
            ("Supplier Name:", "supplier"),
            ("Product Name:", "product"),
            ("Specifications:", "spec"),
            ("Color:", "color"),
            ("Elastic (Single/Double):", "elastic"),
            ("Unit Price:", "price"),
            ("Price Unit:", "unit"),
            ("MOQ:", "moq"),
            ("Packing Details:", "packing"),
            ("Price Term:", "term"),
            ("Delivery Lead Time:", "lead_time"),
            ("Validity Date:", "validity_date"),
            ("Risk Alerts:", "sourcing_risk")
        ]
        
        entries = {}
        for idx, (lbl_txt, field) in enumerate(labels):
            tk_lbl = ctk.CTkLabel(edit_win, text=lbl_txt, anchor="w")
            tk_lbl.grid(row=idx, column=0, padx=15, pady=5, sticky="ew")
            
            val = str(row_dict[field])
            tk_ent = ctk.CTkEntry(edit_win, width=250)
            tk_ent.insert(0, val)
            tk_ent.grid(row=idx, column=1, padx=15, pady=5, sticky="ew")
            entries[field] = tk_ent
            
        def save_changes():
            try:
                price_val = float(entries["price"].get().replace("$", "").replace("¥", "").replace("€", ""))
            except ValueError:
                messagebox.showerror("Invalid Input", "Price must be a numeric value.", parent=edit_win)
                return
                
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                UPDATE extracted_quotes 
                SET supplier=?, product=?, spec=?, color=?, elastic=?, price=?, unit=?, moq=?, packing=?, term=?, lead_time=?, validity_date=?, sourcing_risk=?, review_status='Needs Review', reviewed_by='Local User', reviewed_at=datetime('now'), review_notes='Edited quote requires re-approval.'
                WHERE id=?
            """, (
                entries["supplier"].get(),
                entries["product"].get(),
                entries["spec"].get(),
                entries["color"].get(),
                entries["elastic"].get(),
                price_val,
                entries["unit"].get(),
                entries["moq"].get(),
                entries["packing"].get(),
                entries["term"].get(),
                entries["lead_time"].get(),
                entries["validity_date"].get(),
                entries["sourcing_risk"].get(),
                q_id
            ))
            c.execute("""
                INSERT INTO quote_audit_log (quote_id, action, previous_status, new_status, note)
                VALUES (?, ?, ?, ?, ?)
            """, (
                q_id,
                "Quote Edited",
                row_dict.get("review_status"),
                "Needs Review",
                "Manual edit saved; quote returned to review queue."
            ))
            conn.commit()
            conn.close()
            
            self.load_all_quotes_from_db()
            edit_win.destroy()
            
        btn_save = ctk.CTkButton(edit_win, text="Save Changes", fg_color="#1f7d44", hover_color="#15592e", command=save_changes)
        btn_save.grid(row=len(labels), column=0, columnspan=2, pady=15)

    # --- Export Logic (Beautiful Excel Report Designer) ---
    def export_to_excel(self):
        if not self.extracted_data:
            messagebox.showwarning("No Data", "There is no extracted data to export.")
            return
            
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                df = pd.DataFrame(self.extracted_data)
                df = df.drop(columns=["id"])
                
                # Retrieve active currency choices
                currency_choice = self.currency_cb.get()
                factor = 1.0
                price_col_header = "Unit Price (USD)"
                excel_format = '$#,##0.00000'
                
                if "CNY" in currency_choice:
                    factor = self.exchange_rates["CNY"]
                    price_col_header = "Unit Price (CNY)"
                    excel_format = '[$¥-804]#,##0.00000'
                elif "EUR" in currency_choice:
                    factor = self.exchange_rates["EUR"]
                    price_col_header = "Unit Price (EUR)"
                    excel_format = '[$€-2] #,##0.00000'

                # Apply active conversion to df export
                df["price"] = df["price"].apply(lambda p: float(p) * factor if isinstance(p, (int, float)) else p)
                df.columns = ["Source File", "Supplier", "Product", "Specifications", "Color", "Elastic", price_col_header, "Price Unit", "MOQ", "Packing Details", "Price Term", "Lead Time", "Validity Date", "Risk Alerts"]
                
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Quotes Comparison")
                    workbook = writer.book
                    worksheet = writer.sheets["Quotes Comparison"]
                    
                    # Styles definitions
                    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Navy Blue
                    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    
                    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Grey
                    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
                    
                    normal_font = Font(name="Segoe UI", size=10)
                    best_font = Font(name="Segoe UI", size=10, bold=True, color="375623") # Dark Green
                    
                    thin_side = Side(border_style="thin", color="D9D9D9")
                    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                    
                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")
                    
                    # Format Header Row
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = align_center
                        cell.border = thin_border
                    
                    # Format Data Rows & highlights
                    groups = {}
                    for r in self.extracted_data:
                        price = r.get("price")
                        try:
                            price = float(price)
                        except (ValueError, TypeError):
                            continue
                        if price <= 0:
                            continue
                        g_key = self.get_group_key(r)
                        if g_key not in groups:
                            groups[g_key] = []
                        groups[g_key].append(price)

                    best_prices = {}
                    for g_key, prices in groups.items():
                        if len(prices) > 1:
                            best_prices[g_key] = min(prices)

                    for row_idx, row_data in enumerate(self.extracted_data, start=2):
                        is_best = False
                        g_key = self.get_group_key(row_data)
                        try:
                            price_val = float(row_data["price"])
                            if g_key in best_prices and abs(price_val - best_prices[g_key]) < 1e-7:
                                is_best = True
                        except (ValueError, TypeError):
                            pass
                            
                        row_fill = best_fill if is_best else (alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None))
                        row_font = best_font if is_best else normal_font
                        
                        for col_idx in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = row_fill
                            cell.font = row_font
                            cell.border = thin_border
                            
                            # Alignments
                            if col_idx in [1, 2, 3, 4, 10]:  # Source, Supplier, Product, Specs, Packing
                                cell.alignment = align_left
                            else:
                                cell.alignment = align_center
                                
                            # Price custom formatting matching currency selected
                            if col_idx == 7:
                                cell.number_format = excel_format
                                
                    # Set custom row heights
                    worksheet.row_dimensions[1].height = 28
                    for r_idx in range(2, len(self.extracted_data) + 2):
                        worksheet.row_dimensions[r_idx].height = 20
                        
                    # Auto-fit columns
                    for col in worksheet.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = col[0].column_letter
                        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                        
                messagebox.showinfo("Export Successful", f"Excel sheet exported successfully to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not write Excel file:\n{e}")

    def export_to_csv(self):
        if not self.extracted_data:
            messagebox.showwarning("No Data", "There is no extracted data to export.")
            return
            
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if file_path:
            try:
                df = pd.DataFrame(self.extracted_data)
                df = df.drop(columns=["id"])
                df.columns = ["Source File", "Supplier", "Product", "Specifications", "Color", "Elastic", "Unit Price (USD)", "Price Unit", "MOQ", "Packing Details", "Price Term", "Lead Time", "Validity Date", "Risk Alerts"]
                df.to_csv(file_path, index=False)
                messagebox.showinfo("Export Successful", f"CSV file exported successfully to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not write CSV file:\n{e}")

    def export_to_pdf(self):
        if not self.extracted_data:
            messagebox.showwarning("No Data", "There is no extracted data to generate a report.")
            return
            
        category_filter = self.insights_category_cb.get()
        
        # Filter quotes by selected category
        filtered_data = self.extracted_data
        if category_filter and category_filter != "All":
            filtered_data = [r for r in self.extracted_data if (r.get("product") or "").strip().lower() == category_filter.lower()]
            
        if not filtered_data:
            messagebox.showwarning("No Data", f"No quote data found for category '{category_filter}' to export.")
            return
            
        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not file_path:
            return
            
        # Run in a thread with a loading dialog since it asks Gemini for the executive summary
        loading_win = ctk.CTkToplevel(self)
        loading_win.title("Generating PDF Report")
        loading_win.geometry("350x120")
        loading_win.resizable(False, False)
        loading_win.attributes("-topmost", True)
        
        lbl = ctk.CTkLabel(loading_win, text="AI is writing executive summary...\nPlease wait...", font=ctk.CTkFont(size=12))
        lbl.pack(pady=30)
        
        def run_pdf_gen():
            try:
                # 1. Ask Gemini for executive summary
                quotes_summary_str = ""
                for idx, r in enumerate(filtered_data):
                    quotes_summary_str += f"- Supplier: {r.get('supplier')}, Product: {r.get('product')}, Spec: {r.get('spec')}, Price: {r.get('price')} USD, Lead Time: {r.get('lead_time')}, Validity: {r.get('validity_date')}, Sourcing Risk: {r.get('sourcing_risk')}\n"
                
                category_desc = f"category '{category_filter}'" if category_filter != "All" else "all categories"
                ai_prompt = f"""
                You are a senior global sourcing manager. Write a concise executive summary and recommendation report based on these supplier quotes for {category_desc}:
                {quotes_summary_str}
                
                Keep the summary strictly under 150 words. Focus on:
                - Who offers the best price.
                - Who is the fastest/safest option.
                - Sourcing warnings or risks to be aware of.
                Write in highly professional, executive language. Do not use markdown syntax.
                """
                
                # Call fallback generator
                try:
                    summary_text = self.generate_with_fallback([], ai_prompt, json_response=False)
                except Exception:
                    summary_text = f"Standard quote comparison report generated for {category_desc}. Please inspect details in the table below to determine supplier suitability based on pricing, lead time, and risk considerations."
                
                # 2. Build PDF Document using ReportLab
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                
                doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
                story = []
                
                styles = getSampleStyleSheet()
                
                # Custom Styles
                title_style = ParagraphStyle(
                    'DocTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    leading=22,
                    textColor=colors.HexColor('#1f538d'),
                    spaceAfter=15
                )
                
                section_heading = ParagraphStyle(
                    'SectionHeading',
                    parent=styles['Heading2'],
                    fontSize=12,
                    leading=15,
                    textColor=colors.HexColor('#1f538d'),
                    spaceBefore=12,
                    spaceAfter=6,
                    keepWithNext=True
                )
                
                body_style = ParagraphStyle(
                    'BodyTextCustom',
                    parent=styles['BodyText'],
                    fontSize=9,
                    leading=12,
                    textColor=colors.HexColor('#2c3e50')
                )
                
                header_style = ParagraphStyle(
                    'TableHeader',
                    parent=body_style,
                    textColor=colors.white,
                    fontSize=8,
                    leading=10,
                    fontName='Helvetica-Bold'
                )
                
                cell_style = ParagraphStyle(
                    'TableCell',
                    parent=body_style,
                    fontSize=7.5,
                    leading=9
                )
                
                # Document Header
                story.append(Paragraph(f"SUPPLIER COMPARISON & SOURCING RISK REPORT ({category_filter.upper()})", title_style))
                story.append(Paragraph(f"<b>Date:</b> August 1, 2026 &nbsp;&nbsp;|&nbsp;&nbsp; <b>Total Quotes:</b> {len(filtered_data)}", body_style))
                story.append(Spacer(1, 10))
                
                # Divider Line
                d_table = Table([[""]], colWidths=[540], rowHeights=[2])
                d_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#1f538d')),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                    ('TOPPADDING', (0,0), (-1,-1), 0),
                ]))
                story.append(d_table)
                story.append(Spacer(1, 10))
                
                # AI Executive Summary Section
                story.append(Paragraph("AI Executive Summary & Sourcing Advice", section_heading))
                story.append(Paragraph(summary_text.replace("\n", "<br/>"), body_style))
                story.append(Spacer(1, 15))
                
                # Comparison Table Section
                story.append(Paragraph("Supplier Quotation Grid Details", section_heading))
                
                # Build table headers & rows
                headers = [
                    Paragraph("Supplier", header_style),
                    Paragraph("Product", header_style),
                    Paragraph("Specs", header_style),
                    Paragraph("Price", header_style),
                    Paragraph("Lead Time", header_style),
                    Paragraph("Validity", header_style),
                    Paragraph("Risk Alerts", header_style)
                ]
                
                table_data = [headers]
                for r in filtered_data:
                    # Clean currency display for PDF table
                    currency_choice = self.currency_cb.get()
                    symbol = "$"
                    factor = 1.0
                    if "CNY" in currency_choice:
                         symbol = "¥"
                         factor = self.exchange_rates["CNY"]
                    elif "EUR" in currency_choice:
                         symbol = "€"
                         factor = self.exchange_rates["EUR"]
                        
                    try:
                        p_val = float(r["price"])
                        price_display = f"{symbol}{p_val * factor:.5f}"
                    except Exception:
                        price_display = str(r["price"])
                        
                    val_display = self.get_validity_display(r.get("validity_date"))
                    risk_display = self.get_risk_display(r.get("sourcing_risk"))
                    
                    row = [
                        Paragraph(self.clean_supplier_name(r.get("supplier")), cell_style),
                        Paragraph(r.get("product"), cell_style),
                        Paragraph(r.get("spec"), cell_style),
                        Paragraph(price_display, cell_style),
                        Paragraph(r.get("lead_time"), cell_style),
                        Paragraph(val_display, cell_style),
                        Paragraph(risk_display, cell_style)
                    ]
                    table_data.append(row)
                    
                # Setup column widths (Total width 540)
                col_widths = [80, 75, 90, 50, 50, 85, 110]
                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                # Table style
                t_style = TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f538d')),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                ])
                
                # Alternating row colors
                for r_idx in range(1, len(table_data)):
                    bg = colors.HexColor('#ffffff') if r_idx % 2 == 1 else colors.HexColor('#f9f9f9')
                    t_style.add('BACKGROUND', (0, r_idx), (-1, r_idx), bg)
                    
                t.setStyle(t_style)
                story.append(t)
                story.append(Spacer(1, 20))
                
                # 3. Add Sourcing Dossier Analytics (Landed Cost & Profit Simulator overview)
                story.append(Paragraph("Sourcing Dossier & Business Profitability Projections", section_heading))
                story.append(Paragraph("Below is a simulated projection of DDP landed costs and net operating profitability margins for your active quotes (assuming a target retail price of $1.50 per unit, $120/CBM LCL rate, and 6.5% duties):", body_style))
                story.append(Spacer(1, 10))

                # Build Dossier table
                dossier_headers = [
                    Paragraph("Supplier", header_style),
                    Paragraph("Product", header_style),
                    Paragraph("FOB ($)", header_style),
                    Paragraph("Landed ($)", header_style),
                    Paragraph("Net Profit ($)", header_style),
                    Paragraph("ROI (%)", header_style)
                ]
                dossier_data = [dossier_headers]

                import math
                for r_d in filtered_data:
                    try:
                        fob = float(r_d.get("price"))
                    except:
                        continue
                    s_name = self.clean_supplier_name(r_d.get("supplier"))
                    p_name = r_d.get("product") or "Product"
                    packing = r_d.get("packing_details") or "100/ctn, 0.1cbm"
                    
                    # Basic parser for pack metrics (Simplified)
                    parts = packing.lower().replace(",", " ").split()
                    pcs_per_ctn = 100
                    unit_cbm = 0.001
                    for p in parts:
                        if "ctn" in p: pcs_per_ctn = float(p.replace("ctn", "").strip("/")) or 100
                        if "cbm" in p: unit_cbm = float(p.replace("cbm", "").strip("/")) or 0.001
                    
                    total_ctns = math.ceil(100000 / pcs_per_ctn)
                    total_cbm = total_ctns * unit_cbm
                    total_fob = 100000 * fob
                    total_freight = total_cbm * 120.0
                    total_duty = total_fob * 0.065
                    total_local = 350.0
                    total_landed = total_fob + total_freight + total_duty + total_local
                    landed_pc = total_landed / 100000
                    
                    opex_total_pc = 0.50
                    unit_total_cost = landed_pc + opex_total_pc
                    net_profit = 1.50 - unit_total_cost
                    roi = (net_profit / landed_pc) * 100.0 if landed_pc > 0 else 0.0
                    
                    dossier_data.append([
                        Paragraph(s_name, cell_style),
                        Paragraph(p_name, cell_style),
                        Paragraph(f"${fob:.2f}", cell_style),
                        Paragraph(f"${landed_pc:.4f}", cell_style),
                        Paragraph(f"${net_profit:.4f}", cell_style),
                        Paragraph(f"{roi:.1f}%", cell_style)
                    ])

                dossier_table = Table(dossier_data, colWidths=[120, 110, 75, 75, 80, 80])
                dossier_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2b3e50')),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                ]))
                for r_idx in range(1, len(dossier_data)):
                    bg = colors.HexColor('#ffffff') if r_idx % 2 == 1 else colors.HexColor('#f2f5f9')
                    dossier_table.setStyle(TableStyle([('BACKGROUND', (0, r_idx), (-1, r_idx), bg)]))
                
                story.append(dossier_table)
                story.append(Spacer(1, 15))
                
                # Footer note
                story.append(Spacer(1, 10))
                story.append(Paragraph("<i>Disclaimer: This report was generated automatically via the AI Supplier Quote Extractor based on extracted quotation records in your database. Please confirm terms directly with suppliers before committing.</i>", cell_style))
                
                doc.build(story)
                
                loading_win.after(0, loading_win.destroy)
                loading_win.after(0, lambda: messagebox.showinfo("Export Successful", f"PDF Sourcing Report generated successfully:\n{file_path}"))
                
            except Exception as e:
                loading_win.after(0, loading_win.destroy)
                import traceback
                traceback.print_exc()
                loading_win.after(0, lambda e_err=e: messagebox.showerror("Export Failed", f"Could not write PDF file:\n{e_err}"))
                
        threading.Thread(target=run_pdf_gen, daemon=True).start()

    def setup_sourcing_insights_tab(self):
        tab_insights = self.sourcing_tabview.tab("💡 AI Sourcing Insights")
        tab_insights.grid_columnconfigure(0, weight=1)
        tab_insights.grid_columnconfigure(1, weight=1)
        tab_insights.grid_rowconfigure(1, weight=1)

        # Header Row Container
        header_row = ctk.CTkFrame(tab_insights, fg_color="transparent")
        header_row.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 5), sticky="ew")

        title_lbl = ctk.CTkLabel(header_row, text="AI Sourcing Insights & Negotiator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.pack(side="left")

        # Category Filter Combobox
        self.insights_category_cb = ctk.CTkComboBox(header_row, values=["All"], command=lambda choice: self.update_sourcing_insights(), width=150)
        self.insights_category_cb.pack(side="right", padx=10)
        self.insights_category_cb.set("All")

        insights_lbl = ctk.CTkLabel(header_row, text="Filter Product Category:")
        insights_lbl.pack(side="right", padx=5)

        # --- LEFT PANEL: AI Analysis Cards ---
        left_frame = ctk.CTkFrame(tab_insights)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(1, weight=1)
        left_frame.grid_rowconfigure(2, weight=1)
        left_frame.grid_rowconfigure(3, weight=1)

        ctk.CTkLabel(left_frame, text="🏆 AI Sourcing Matrix Highlights", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        # Cards for Cheapest, Fastest, Safest
        self.cheapest_card = ctk.CTkTextbox(left_frame, height=90, activate_scrollbars=False, wrap="word")
        self.cheapest_card.grid(row=1, column=0, padx=15, pady=5, sticky="ew")
        
        self.fastest_card = ctk.CTkTextbox(left_frame, height=90, activate_scrollbars=False, wrap="word")
        self.fastest_card.grid(row=2, column=0, padx=15, pady=5, sticky="ew")
        
        self.safest_card = ctk.CTkTextbox(left_frame, height=90, activate_scrollbars=False, wrap="word")
        self.safest_card.grid(row=3, column=0, padx=15, pady=5, sticky="ew")

        # Build PDF Button inside left_frame under cards
        self.btn_export_pdf = ctk.CTkButton(left_frame, text="📄 Build PDF Report", fg_color="#8c1c1c", hover_color="#6e1313", command=self.export_to_pdf)
        self.btn_export_pdf.grid(row=4, column=0, padx=15, pady=(10, 15), sticky="ew")

        # --- RIGHT PANEL: Counter-Offer Email Writer ---
        right_frame = ctk.CTkFrame(tab_insights)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(right_frame, text="✉ One-Click Negotiation Email Generator", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        ctrl_row = ctk.CTkFrame(right_frame, fg_color="transparent")
        ctrl_row.grid(row=1, column=0, padx=15, pady=5, sticky="ew")

        # Select Supplier
        ctk.CTkLabel(ctrl_row, text="Supplier:").pack(side="left", padx=2)
        self.negotiate_supplier_cb = ctk.CTkComboBox(ctrl_row, values=["No suppliers loaded"], width=120)
        self.negotiate_supplier_cb.pack(side="left", padx=2)

        # Select Goal
        ctk.CTkLabel(ctrl_row, text="Goal:").pack(side="left", padx=2)
        self.negotiate_goal_cb = ctk.CTkComboBox(ctrl_row, values=["Price Match", "Better Terms", "Fast Delivery", "Lower MOQ"], width=110)
        self.negotiate_goal_cb.pack(side="left", padx=2)

        # Select Channel
        ctk.CTkLabel(ctrl_row, text="Channel:").pack(side="left", padx=2)
        self.negotiate_channel_btn = ctk.CTkSegmentedButton(ctrl_row, values=["Email", "Chat"], width=110)
        self.negotiate_channel_btn.pack(side="left", padx=2)
        self.negotiate_channel_btn.set("Email")

        self.btn_gen_email = ctk.CTkButton(ctrl_row, text="✍ Generate", width=70, fg_color="#1f538d", command=self.generate_negotiation_email)
        self.btn_gen_email.pack(side="left", padx=5)

        # Textbox for output email
        self.email_text_box = ctk.CTkTextbox(right_frame, wrap="word")
        self.email_text_box.grid(row=2, column=0, padx=15, pady=10, sticky="nsew")

        # Copy and Launch buttons row
        action_row = ctk.CTkFrame(right_frame, fg_color="transparent")
        action_row.grid(row=3, column=0, padx=15, pady=(5, 15), sticky="ew")

        self.btn_copy_email = ctk.CTkButton(action_row, text="📋 Copy Text", width=130, fg_color="#1f7d44", hover_color="#15592e", command=self.copy_email_to_clipboard)
        self.btn_copy_email.pack(side="left", padx=(0, 5))

        self.btn_launch_email = ctk.CTkButton(action_row, text="📧 Launch Email", width=140, fg_color="#1f538d", command=self.launch_email_client)
        self.btn_launch_email.pack(side="left", padx=(5, 0))

    def update_sourcing_insights(self):
        if not self.extracted_data:
            self.set_card_text(self.cheapest_card, "💵 Cheapest Option\nNo quotes data loaded in database.")
            self.set_card_text(self.fastest_card, "⚡ Fastest Option\nNo quotes data loaded in database.")
            self.set_card_text(self.safest_card, "🛡 Safest Option (Lowest Risk)\nNo quotes data loaded in database.")
            self.negotiate_supplier_cb.configure(values=["No suppliers loaded"])
            self.negotiate_supplier_cb.set("No suppliers loaded")
            return

        category_filter = self.insights_category_cb.get()
        
        # Filter quotes by selected category
        filtered_data = self.extracted_data
        if category_filter and category_filter != "All":
            filtered_data = [r for r in self.extracted_data if (r.get("product") or "").strip().lower() == category_filter.lower()]

        if not filtered_data:
            self.set_card_text(self.cheapest_card, f"💵 Cheapest Option\nNo quotes found for category '{category_filter}'.")
            self.set_card_text(self.fastest_card, f"⚡ Fastest Option\nNo quotes found for category '{category_filter}'.")
            self.set_card_text(self.safest_card, f"🛡 Safest Option (Lowest Risk)\nNo quotes found for category '{category_filter}'.")
            self.negotiate_supplier_cb.configure(values=["No suppliers loaded"])
            self.negotiate_supplier_cb.set("No suppliers loaded")
            return

        # 1. Update dropdown values to only show suppliers who offer the selected product category
        suppliers = sorted(list({r.get("supplier") for r in filtered_data if r.get("supplier") and r.get("supplier") != "Unknown"}))
        if suppliers:
            self.negotiate_supplier_cb.configure(values=suppliers)
            self.negotiate_supplier_cb.set(suppliers[0])
        else:
            self.negotiate_supplier_cb.configure(values=["No suppliers loaded"])
            self.negotiate_supplier_cb.set("No suppliers loaded")

        # 2. Extract Cheapest Option
        cheapest_row = None
        min_price = float('inf')
        for r in filtered_data:
            try:
                p = float(r["price"])
                if p < min_price:
                    min_price = p
                    cheapest_row = r
            except Exception:
                pass
        
        if cheapest_row:
            self.set_card_text(self.cheapest_card, f"💵 CHEAPEST OPTION:\nSupplier: {self.clean_supplier_name(cheapest_row['supplier'])}\nProduct: {cheapest_row['product']} - Price: ${cheapest_row['price']:.5f} / pc\nSpecs: {cheapest_row['spec']}")
        else:
            self.set_card_text(self.cheapest_card, "💵 Cheapest Option\nNo numeric price quotes found.")

        # 3. Extract Fastest Option (Lowest lead time)
        fastest_row = None
        min_days = float('inf')
        for r in filtered_data:
            lt_str = r.get("lead_time") or ""
            import re
            match = re.search(r'\d+', lt_str)
            if match:
                days = int(match.group(0))
                if days < min_days:
                    min_days = days
                    fastest_row = r
        
        if fastest_row:
            self.set_card_text(self.fastest_card, f"⚡ FASTEST OPTION:\nSupplier: {self.clean_supplier_name(fastest_row['supplier'])}\nProduct: {fastest_row['product']} - Lead Time: {fastest_row['lead_time']}\nPrice: ${fastest_row['price']} / pc")
        else:
            self.set_card_text(self.fastest_card, "⚡ Fastest Option\nNo specific lead times found.")

        # 4. Extract Safest Option (Lowest risk, active validity date)
        safest_row = None
        for r in filtered_data:
            risk = (r.get("sourcing_risk") or "").lower()
            val = (r.get("validity_date") or "").lower()
            if "low" in risk and "expired" not in val:
                safest_row = r
                break
        
        if not safest_row:
            for r in filtered_data:
                risk = (r.get("sourcing_risk") or "").lower()
                if "low" in risk:
                    safest_row = r
                    break
        
        if not safest_row and filtered_data:
            safest_row = filtered_data[0]

        if safest_row:
            self.set_card_text(self.safest_card, f"🛡 SAFEST OPTION (LOWEST RISK):\nSupplier: {self.clean_supplier_name(safest_row['supplier'])}\nProduct: {safest_row.get('product')}\nRisk Level: {safest_row.get('sourcing_risk')}\nValidity: {safest_row.get('validity_date')}")
        else:
            self.set_card_text(self.safest_card, "🛡 Safest Option (Lowest Risk)\nNo risk evaluation details found.")

    def set_card_text(self, card_widget, text):
        card_widget.configure(state="normal")
        card_widget.delete("1.0", tk.END)
        card_widget.insert("1.0", text)
        card_widget.configure(state="disabled")

    def generate_negotiation_email(self):
        supplier = self.negotiate_supplier_cb.get()
        goal = self.negotiate_goal_cb.get()
        if not supplier or supplier == "No suppliers loaded":
            messagebox.showwarning("Select Supplier", "Please select a valid supplier to negotiate with.")
            return

        self.btn_gen_email.configure(state="disabled", text="Writing...")
        
        def run_email_gen():
            try:
                category_filter = self.insights_category_cb.get()
                supplier_quotes = [q for q in self.extracted_data if q.get("supplier") == supplier]
                other_quotes = [q for q in self.extracted_data if q.get("supplier") != supplier]
                
                # Apply product category filtering
                if category_filter and category_filter != "All":
                    supplier_quotes = [q for q in supplier_quotes if (q.get("product") or "").strip().lower() == category_filter.lower()]
                    other_quotes = [q for q in other_quotes if (q.get("product") or "").strip().lower() == category_filter.lower()]
                
                context_str = f"Supplier Quotation Details:\n{json.dumps(supplier_quotes, indent=2)}\n\nOther Competitor Quotes:\n{json.dumps(other_quotes, indent=2)}"
                
                ai_prompt = f"""
                You are an expert global procurement advisor. Write a professional, polite, and persuasive negotiation email to the supplier '{supplier}'.
                Goal of negotiation: {goal}.
                
                Here is the current quotation database context for references:
                {context_str}
                
                Make sure:
                - Use the actual competitor pricing or lead times from the context to negotiate a better deal (e.g., if a competitor offers Cap at $0.0059 and this supplier quoted $0.0071, ask them to match $0.0059 or meet half-way).
                - Sound highly professional, polite, yet firm.
                - Keep placeholders for sender name/company.
                - Output ONLY the ready-to-copy email body. No markdown formatting.
                """
                
                email_body = self.generate_with_fallback([], ai_prompt, json_response=False)
                
                self.after(0, lambda text=email_body: self.display_negotiation_email(text))
            except Exception as e:
                self.after(0, lambda err=e: self.display_negotiation_email(f"Error generating negotiation email: {err}"))
                
        threading.Thread(target=run_email_gen, daemon=True).start()

    def display_negotiation_email(self, text):
        self.btn_gen_email.configure(state="normal", text="✍ Generate")
        self.email_text_box.delete("1.0", tk.END)
        self.email_text_box.insert("1.0", text)

    def copy_email_to_clipboard(self):
        email_content = self.email_text_box.get("1.0", tk.END).strip()
        if not email_content:
            return
        self.clipboard_clear()
        self.clipboard_append(email_content)
        messagebox.showinfo("Copied", "Negotiation email copied to clipboard successfully!")

    def copy_to_clipboard(self, text):
        if not text:
            return
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("Copied", "Text copied to clipboard successfully!")

    def select_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Supplier Quotes",
            filetypes=[("Quote files", "*.pdf;*.png;*.jpg;*.jpeg;*.txt;*.xlsx;*.xls")]
        )
        if not file_paths:
            return
            
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT filename FROM processed_files")
        processed = {row[0] for row in c.fetchall()}
        conn.close()
        
        # If no folder was selected, default folder path to the parent directory of the first file
        if not self.selected_folder and file_paths:
            self.selected_folder = os.path.dirname(file_paths[0])
            self.folder_entry.delete(0, tk.END)
            self.folder_entry.insert(0, self.selected_folder)
            
        existing_names = {item[0] for item in self.files_list}
        
        for path in file_paths:
            name = os.path.basename(path)
            if name not in existing_names:
                self.files_list.append((name, path))
                if name in processed:
                    self.files_box_synced.insert(tk.END, f"✅ {name}")
                else:
                    self.files_box_unsynced.insert(tk.END, f"⏳ {name}")
                    
        has_unsynced = any(name not in processed for (name, path) in self.files_list)
        if has_unsynced:
            self.btn_start.configure(state="normal")

    def clean_filename_part(self, s):
        if not s:
            return ""
        import re
        # Remove anything that isn't alphanumeric or space
        s = re.sub(r'[^a-zA-Z0-9\s-]', '', s)
        # Convert spaces to CamelCase
        return "".join(word.capitalize() for word in s.split())

    def start_file_organizer_thread(self):
        if not self.selected_folder:
            messagebox.showerror("Folder Missing", "Please select a quotes folder or load files first.")
            return
            
        ans = messagebox.askyesno(
            "Organize Files",
            "This will create a new folder 'Organized_Quotes' inside your selected folder.\n\nIt will rename and copy your quotation files into subfolders sorted by Supplier Name (e.g. 'Organized_Quotes/XiantaoTopmed/XiantaoTopmed_Hairnet_20260801.pdf').\n\nWould you like to proceed?"
        )
        if not ans:
            return
            
        threading.Thread(target=self.run_file_organizer, daemon=True).start()

    def run_file_organizer(self):
        try:
            # Query all quotes
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT filename, supplier, product, validity_date FROM extracted_quotes")
            rows = c.fetchall()
            conn.close()
            
            if not rows:
                self.after(0, lambda: messagebox.showwarning("No Data", "No quotes found in database to organize."))
                return
                
            # Group by unique source filename to avoid duplicating the same source file
            from collections import defaultdict
            grouped = defaultdict(list)
            for filename, supplier, product, validity in rows:
                grouped[filename].append((supplier, product, validity))
                
            organized_dir = os.path.join(self.selected_folder, "Organized_Quotes")
            os.makedirs(organized_dir, exist_ok=True)
            
            copied_count = 0
            filename_to_path = {item[0]: item[1] for item in self.files_list}
            
            for filename, items in grouped.items():
                src_path = filename_to_path.get(filename)
                if not src_path:
                    src_path = os.path.join(self.selected_folder, filename)
                    
                if not os.path.exists(src_path):
                    continue
                    
                # Use details from the first item to name the file
                supplier, first_product, validity = items[0]
                
                # Get unique products in this file
                unique_products = list(dict.fromkeys([it[1] for it in items if it[1] and it[1] != "N/A"]))
                
                if len(unique_products) > 1:
                    clean_product = self.clean_filename_part(unique_products[0]) or "Product"
                    clean_product = f"{clean_product}_etc"
                elif len(unique_products) == 1:
                    clean_product = self.clean_filename_part(unique_products[0]) or "Product"
                else:
                    clean_product = "Product"
                    
                clean_supplier = self.clean_filename_part(supplier) or "UnknownSupplier"
                
                # Format date
                date_part = ""
                if validity and validity != "N/A":
                    date_part = "_" + validity.replace("-", "")
                
                ext = os.path.splitext(filename)[1]
                new_filename = f"{clean_supplier}_{clean_product}{date_part}{ext}"
                
                supplier_dir = os.path.join(organized_dir, clean_supplier)
                os.makedirs(supplier_dir, exist_ok=True)
                
                dest_path = os.path.join(supplier_dir, new_filename)
                
                import shutil
                shutil.copy2(src_path, dest_path)
                copied_count += 1
                
            self.after(0, lambda: messagebox.showinfo("Success", f"Quotes organized successfully!\nCopied and renamed {copied_count} files into:\n{organized_dir}"))
        except Exception as e:
            self.after(0, lambda err=e: messagebox.showerror("Error", f"Failed to organize files: {err}"))

    def setup_scorecard_tab(self):
        tab_scorecard = self.scorecard_tabview.tab("🏆 Supplier Scorecard")
        tab_scorecard.grid_columnconfigure(0, weight=3)
        tab_scorecard.grid_columnconfigure(1, weight=1)
        tab_scorecard.grid_rowconfigure(1, weight=1)
        tab_scorecard.grid_rowconfigure(2, weight=0)

        # Header Frame (Spans both columns)
        header_row = ctk.CTkFrame(tab_scorecard, fg_color="transparent")
        header_row.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 5), sticky="ew")

        title_lbl = ctk.CTkLabel(header_row, text="Supplier Scorecard Matrix", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.pack(side="left")

        # Category Filter Combobox
        self.scorecard_category_cb = ctk.CTkComboBox(header_row, values=["All"], command=lambda choice: self.update_scorecard_tab(), width=150)
        self.scorecard_category_cb.pack(side="right", padx=10)
        self.scorecard_category_cb.set("All")

        scorecard_lbl = ctk.CTkLabel(header_row, text="Filter Product Category:")
        scorecard_lbl.pack(side="right", padx=5)

        # --- LEFT COLUMN: Treeview Scorecard Table ---
        table_frame = ctk.CTkFrame(tab_scorecard, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        table_frame.grid(row=1, column=0, padx=(20, 10), pady=10, sticky="nsew")
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        # Set up custom styles for Scorecard treeview
        self.style_workspace_table("Scorecard.Treeview")
        style = ttk.Style()
        style.configure("Scorecard.Treeview", rowheight=31, font=("Segoe UI", 10))
        style.configure("Scorecard.Treeview.Heading", font=("Segoe UI", 10, "bold"))

        cols = ("Rank", "Supplier", "Product", "Price", "MOQ", "Lead Time", "Risk", "Score", "Rating")
        self.scorecard_tree = ttk.Treeview(table_frame, columns=cols, show="headings", style="Scorecard.Treeview")
        self.scorecard_tree.bind("<<TreeviewSelect>>", self.on_scorecard_select)
        
        col_widths = {
            "Rank": 45,
            "Supplier": 110,
            "Product": 95,
            "Price": 70,
            "MOQ": 70,
            "Lead Time": 80,
            "Risk": 90,
            "Score": 60,
            "Rating": 80
        }
        
        for col in cols:
            self.scorecard_tree.heading(col, text=col)
            self.scorecard_tree.column(col, width=col_widths.get(col, 90), anchor="center" if col in ["Rank", "Price", "MOQ", "Score", "Rating"] else "w")

        # Scrollbar
        sb = ttk.Scrollbar(table_frame, orient="vertical", command=self.scorecard_tree.yview)
        self.scorecard_tree.configure(yscrollcommand=sb.set)
        
        self.scorecard_tree.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")

        # Recommendation Text Card at bottom
        self.rec_card = ctk.CTkTextbox(tab_scorecard, height=75, wrap="word", font=ctk.CTkFont(size=12))
        self.rec_card.grid(row=2, column=0, padx=(20, 10), pady=(5, 15), sticky="ew")

        # --- RIGHT COLUMN: Sourcing Weight Simulator ---
        self.scorecard_sim_frame = ctk.CTkFrame(
            tab_scorecard,
            fg_color=self.THEME["surface"],
            border_color=self.THEME["border"],
            border_width=1,
            corner_radius=8,
        )
        self.scorecard_sim_frame.grid(row=1, column=1, rowspan=2, padx=(10, 20), pady=10, sticky="nsew")
        self.scorecard_sim_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self.scorecard_sim_frame, text="🎛 Sourcing Priorities", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=(15, 10), sticky="w")

        # Initial default weights
        self.weight_price = 40
        self.weight_lead = 20
        self.weight_moq = 20
        self.weight_risk = 20

        # Sliders
        self.lbl_w_price = ctk.CTkLabel(self.scorecard_sim_frame, text="Price Weight: 40%", font=ctk.CTkFont(size=11))
        self.lbl_w_price.grid(row=1, column=0, padx=15, pady=(10, 2), sticky="w")
        self.slider_price = ctk.CTkSlider(self.scorecard_sim_frame, from_=0, to=100, number_of_steps=100, command=lambda v: self.on_weight_changed("price", v))
        self.slider_price.grid(row=2, column=0, padx=15, pady=2, sticky="ew")
        self.slider_price.set(40)

        self.lbl_w_lead = ctk.CTkLabel(self.scorecard_sim_frame, text="Lead Time Weight: 20%", font=ctk.CTkFont(size=11))
        self.lbl_w_lead.grid(row=3, column=0, padx=15, pady=(10, 2), sticky="w")
        self.slider_lead = ctk.CTkSlider(self.scorecard_sim_frame, from_=0, to=100, number_of_steps=100, command=lambda v: self.on_weight_changed("lead", v))
        self.slider_lead.grid(row=4, column=0, padx=15, pady=2, sticky="ew")
        self.slider_lead.set(20)

        self.lbl_w_moq = ctk.CTkLabel(self.scorecard_sim_frame, text="MOQ Weight: 20%", font=ctk.CTkFont(size=11))
        self.lbl_w_moq.grid(row=5, column=0, padx=15, pady=(10, 2), sticky="w")
        self.slider_moq = ctk.CTkSlider(self.scorecard_sim_frame, from_=0, to=100, number_of_steps=100, command=lambda v: self.on_weight_changed("moq", v))
        self.slider_moq.grid(row=6, column=0, padx=15, pady=2, sticky="ew")
        self.slider_moq.set(20)

        self.lbl_w_risk = ctk.CTkLabel(self.scorecard_sim_frame, text="Risk Weight: 20%", font=ctk.CTkFont(size=11))
        self.lbl_w_risk.grid(row=7, column=0, padx=15, pady=(10, 2), sticky="w")
        self.slider_risk = ctk.CTkSlider(self.scorecard_sim_frame, from_=0, to=100, number_of_steps=100, command=lambda v: self.on_weight_changed("risk", v))
        self.slider_risk.grid(row=8, column=0, padx=15, pady=2, sticky="ew")
        self.slider_risk.set(20)

        # Reset button
        btn_reset_weights = self.make_button(self.scorecard_sim_frame, text="Reset Defaults", command=self.reset_weights, variant="secondary")
        btn_reset_weights.grid(row=9, column=0, padx=15, pady=25, sticky="ew")

    def update_scorecard_tab(self):
        # Clear existing rows
        self.scorecard_tree.delete(*self.scorecard_tree.get_children())
        
        # Calculate scorecard
        matrix = self.calculate_supplier_scorecard()
        if not matrix:
            self.set_card_text(self.rec_card, "🏆 AI Recommendation Scorecard\nNo quotes data available in database to rank.")
            return

        for idx, row in enumerate(matrix):
            price_display = f"${float(row['price']):.5f}" if row['price'] is not None else "N/A"
            self.scorecard_tree.insert("", "end", values=(
                f"#{idx+1}",
                self.clean_supplier_name(row["supplier"]),
                row["product"],
                price_display,
                row["moq"] or "N/A",
                row["lead_time"] or "N/A",
                row["risk"] or "N/A",
                f"{row['score']}/100",
                row["stars"]
            ))

        # Highlight best option in recommendation text
        best = matrix[0]
        best_supplier = self.clean_supplier_name(best["supplier"])
        self.set_card_text(self.rec_card, f"🏆 TOP SOURCING RECOMMENDATION:\nBased on our AI Scorecard algorithm, {best_supplier} is the highly recommended choice for {best['product']} with an overall score of {best['score']}/100. They offer a highly competitive price of ${float(best['price']):.5f}/pc, MOQ of {best['moq'] or 'N/A'}, and lead time of {best['lead_time'] or 'N/A'} with a risk level of: {best['risk'] or 'N/A'}.")

        # Select first row by default to display radar chart
        children = self.scorecard_tree.get_children()
        if children:
            self.scorecard_tree.selection_set(children[0])
            self.on_scorecard_select(None)

    def calculate_supplier_scorecard(self):
        if not hasattr(self, 'scorecard_category_cb'):
            return []
            
        category = self.scorecard_category_cb.get()
        
        # Filter quotes by selected category
        filtered_data = self.get_decision_quotes()
        if category and category != "All":
            filtered_data = [r for r in filtered_data if (r.get("product") or "").strip().lower() == category.lower()]
            
        if not filtered_data:
            return []
            
        # Parse numeric prices, lead times, and MOQs for scaling
        prices = []
        lead_times = []
        moqs = []
        
        for r in filtered_data:
            try:
                if r["price"] is not None:
                    prices.append(float(r["price"]))
            except (ValueError, TypeError):
                pass
            import re
            lt_match = re.search(r'\d+', r.get("lead_time") or "")
            if lt_match:
                lead_times.append(int(lt_match.group(0)))
            moq_match = re.search(r'\d+', (r.get("moq") or "").replace(",", ""))
            if moq_match:
                moqs.append(int(moq_match.group(0)))
                
        min_price = min(prices) if prices else 0.001
        max_price = max(prices) if prices else 1.0
        
        min_lead = min(lead_times) if lead_times else 5
        max_lead = max(lead_times) if lead_times else 30
        
        min_moq = min(moqs) if moqs else 5000
        max_moq = max(moqs) if moqs else 100000
        
        scorecard = []
        for r in filtered_data:
            supplier = r.get("supplier") or "Unknown"
            product = r.get("product") or "Unknown"
            
            # --- Price Score (100 pts max) ---
            try:
                p_val = float(r["price"])
                if max_price == min_price:
                    price_score = 100.0
                else:
                    price_score = 100.0 * (1.0 - (p_val - min_price) / (max_price - min_price))
            except Exception:
                price_score = 40.0
                
            # --- Lead Time Score (100 pts max) ---
            import re
            lt_match = re.search(r'\d+', r.get("lead_time") or "")
            if lt_match:
                lt_val = int(lt_match.group(0))
                if max_lead == min_lead:
                    lt_score = 100.0
                else:
                    lt_score = 100.0 * (1.0 - (lt_val - min_lead) / (max_lead - min_lead))
            else:
                lt_score = 50.0
                
            # --- MOQ Score (100 pts max) ---
            moq_match = re.search(r'\d+', (r.get("moq") or "").replace(",", ""))
            if moq_match:
                moq_val = int(moq_match.group(0))
                if max_moq == min_moq:
                    moq_score = 100.0
                else:
                    moq_score = 100.0 * (1.0 - (moq_val - min_moq) / (max_moq - min_moq))
            else:
                moq_score = 50.0
                
            # --- Risk Score (100 pts max) ---
            risk = (r.get("sourcing_risk") or "").lower()
            if "high" in risk:
                risk_score = 20.0
            elif "medium" in risk:
                risk_score = 60.0
            elif "low" in risk:
                risk_score = 100.0
            else:
                risk_score = 50.0

            # Query incidents database for this supplier and deduct penalties
            try:
                conn = sqlite3.connect(DB_FILE)
                c = conn.cursor()
                c.execute("SELECT supplier, severity FROM supplier_incidents")
                incidents = c.fetchall()
                conn.close()
                
                penalty = 0
                for inc_row in incidents:
                    inc_sup = self.clean_supplier_name(inc_row[0])
                    if inc_sup == self.clean_supplier_name(supplier):
                        sev = str(inc_row[1]).lower()
                        if "high" in sev:
                            penalty += 15
                        elif "medium" in sev:
                            penalty += 5
                
                risk_score = max(0.0, risk_score - penalty)
            except Exception as e:
                print(f"Failed to query incidents penalty for scorecard: {e}")
                
            # Weighted average
            total_w = self.weight_price + self.weight_lead + self.weight_moq + self.weight_risk
            if total_w == 0:
                total_w = 1.0
            total_score = (price_score * self.weight_price + lt_score * self.weight_lead + moq_score * self.weight_moq + risk_score * self.weight_risk) / total_w
            
            star_rating = "⭐" * int(round(total_score / 20.0))
            if not star_rating:
                star_rating = "⭐"
                
            scorecard.append({
                "supplier": supplier,
                "product": product,
                "price": r.get("price"),
                "moq": r.get("moq"),
                "lead_time": r.get("lead_time"),
                "risk": r.get("sourcing_risk"),
                "score": int(round(total_score)),
                "stars": star_rating,
                "price_score": int(round(price_score)),
                "lead_score": int(round(lt_score)),
                "moq_score": int(round(moq_score)),
                "risk_score": int(round(risk_score))
            })
            
        scorecard.sort(key=lambda x: x["score"], reverse=True)
        return scorecard

    def on_weight_changed(self, weight_type, value):
        val = int(value)
        if weight_type == "price":
            self.weight_price = val
            self.lbl_w_price.configure(text=f"Price Weight: {val}%")
        elif weight_type == "lead":
            self.weight_lead = val
            self.lbl_w_lead.configure(text=f"Lead Time Weight: {val}%")
        elif weight_type == "moq":
            self.weight_moq = val
            self.lbl_w_moq.configure(text=f"MOQ Weight: {val}%")
        elif weight_type == "risk":
            self.weight_risk = val
            self.lbl_w_risk.configure(text=f"Risk Weight: {val}%")
            
        self.update_scorecard_tab()

    def reset_weights(self):
        self.weight_price = 40
        self.weight_lead = 20
        self.weight_moq = 20
        self.weight_risk = 20
        
        self.slider_price.set(40)
        self.slider_lead.set(20)
        self.slider_moq.set(20)
        self.slider_risk.set(20)
        
        self.lbl_w_price.configure(text="Price Weight: 40%")
        self.lbl_w_lead.configure(text="Lead Time Weight: 20%")
        self.lbl_w_moq.configure(text="MOQ Weight: 20%")
        self.lbl_w_risk.configure(text="Risk Weight: 20%")
        
        self.update_scorecard_tab()

    def launch_email_client(self):
        supplier = self.negotiate_supplier_cb.get()
        body = self.email_text_box.get("1.0", tk.END).strip()
        if not body:
            return
            
        email = ""
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT contact_info FROM supplier_contacts WHERE supplier = ?", (supplier,))
        row = c.fetchone()
        conn.close()
        
        if row and row[0]:
            contact_str = row[0]
            import re
            emails = re.findall(r'[\w\.-]+@[\w\.-]+', contact_str)
            if emails:
                email = emails[0]
                
        subject = f"Price Negotiation - Quotation Review"
        import urllib.parse
        import webbrowser
        mailto_url = f"mailto:{email}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        try:
            webbrowser.open(mailto_url)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open email client:\n{e}")

    def update_preview_metrics_overlay(self, vals):
        for widget in self.preview_metrics_frame.winfo_children():
            widget.destroy()

        if not vals:
            ctk.CTkLabel(self.preview_metrics_frame, text="Select a quote to view key details", font=("Segoe UI", 10, "italic"), text_color=self.THEME["muted"]).pack(pady=10)
            return

        self.preview_metrics_frame.grid_columnconfigure(0, weight=1)
        self.preview_metrics_frame.grid_columnconfigure(1, weight=2)

        # Title Label
        ctk.CTkLabel(self.preview_metrics_frame, text="🔍 Extracted AI Sourcing Metrics", font=ctk.CTkFont(size=11, weight="bold"), text_color="#a6e3e9").grid(row=0, column=0, columnspan=2, pady=(8, 4))

        metrics = [
            ("Status:", vals[1], "#92400E"),
            ("Supplier:", vals[3], "#1f538d"),
            ("Product:", vals[4], "#15592e"),
            ("Price:", f"${vals[8]} / {vals[9]}", "#6e4513"),
            ("MOQ:", vals[10], "#5a1f1f"),
            ("Lead Time:", vals[13], "#1f538d"),
            ("Validity:", vals[14], "#6e4513"),
            ("Risk Alert:", vals[15], "#5a1f1f")
        ]

        for idx, (label, val, color) in enumerate(metrics):
            lbl = ctk.CTkLabel(self.preview_metrics_frame, text=label, font=ctk.CTkFont(size=10, weight="bold"), anchor="w", text_color=self.THEME["text"])
            lbl.grid(row=idx+1, column=0, sticky="w", padx=10, pady=2)
            
            # Format status values nicely
            val_str = str(val)
            val_lbl = ctk.CTkLabel(self.preview_metrics_frame, text=val_str, font=ctk.CTkFont(size=10), anchor="w", wraplength=180, justify="left", text_color=self.THEME["muted"])
            val_lbl.grid(row=idx+1, column=1, sticky="w", padx=10, pady=2)

    def setup_timeline_tab(self):
        tab_timeline = self.logistics_tabview.tab("📅 Gantt Timeline")
        tab_timeline.grid_columnconfigure(0, weight=1)
        tab_timeline.grid_columnconfigure(1, weight=1)
        tab_timeline.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_timeline, text="Sourcing Timeline & Expiration Planner", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Expiration Alerts ---
        left_frame = ctk.CTkFrame(tab_timeline)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="📅 Quote Expiration Alerts", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")
        
        self.expiration_scroll = ctk.CTkScrollableFrame(left_frame, fg_color=self.THEME["surface_soft"])
        self.expiration_scroll.grid(row=1, column=0, padx=15, pady=10, sticky="nsew")
        self.expiration_scroll.grid_columnconfigure(0, weight=1)

        # --- RIGHT PANEL: Production Planner ---
        right_frame = ctk.CTkFrame(tab_timeline)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(right_frame, text="⚡ Lead Time Production Planner", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        # Filters
        filter_row = ctk.CTkFrame(right_frame, fg_color="transparent")
        filter_row.grid(row=1, column=0, padx=15, pady=5, sticky="ew")

        ctk.CTkLabel(filter_row, text="Category:").pack(side="left", padx=5)
        self.timeline_category_cb = ctk.CTkComboBox(filter_row, values=["All"], command=lambda c: self.update_timeline_tab(), width=130)
        self.timeline_category_cb.pack(side="left", padx=5)
        self.timeline_category_cb.set("All")

        ctk.CTkLabel(filter_row, text="Order Date:").pack(side="left", padx=5)
        self.order_date_entry = ctk.CTkEntry(filter_row, width=110)
        self.order_date_entry.pack(side="left", padx=5)
        self.order_date_entry.insert(0, "2026-08-03")
        self.order_date_entry.bind("<Return>", lambda e: self.update_timeline_tab())
        self.order_date_entry.bind("<FocusOut>", lambda e: self.update_timeline_tab())

        # Timeline display scroll frame
        self.timeline_display_scroll = ctk.CTkScrollableFrame(right_frame, fg_color=self.THEME["surface_soft"])
        self.timeline_display_scroll.grid(row=2, column=0, padx=15, pady=10, sticky="nsew")
        self.timeline_display_scroll.grid_columnconfigure(0, weight=1)

    def update_timeline_tab(self):
        # 1. Update category values
        if hasattr(self, 'timeline_category_cb'):
            products = set()
            for r in self.extracted_data:
                prod = (r.get("product") or "").strip().title()
                if prod:
                    products.add(prod)
            sorted_prods = ["All"] + sorted(list(products))
            current = self.timeline_category_cb.get()
            self.timeline_category_cb.configure(values=sorted_prods)
            if current in sorted_prods:
                self.timeline_category_cb.set(current)
            else:
                self.timeline_category_cb.set("All")

        # 2. Populate Expiration list
        for widget in self.expiration_scroll.winfo_children():
            widget.destroy()

        import datetime
        today = datetime.date(2026, 8, 3)
        
        if not self.extracted_data:
            ctk.CTkLabel(self.expiration_scroll, text="No quotes data loaded in database.", text_color="grey").pack(pady=20)
        else:
            for idx, r in enumerate(self.extracted_data):
                supplier = self.clean_supplier_name(r.get("supplier"))
                product = r.get("product") or "Product"
                val_date_str = r.get("validity_date")
                
                status_text = "🟢 Active"
                text_color = "#a6ffa6"
                bg_color = "#1e4620"
                
                if val_date_str and val_date_str != "N/A":
                    try:
                        val_date = datetime.datetime.strptime(val_date_str, "%Y-%m-%d").date()
                        delta = (val_date - today).days
                        if delta < 0:
                            status_text = f"🔴 Expired (on {val_date_str})"
                            text_color = "#991B1B"
                            bg_color = "#FEF2F2"
                        elif delta <= 14:
                            status_text = f"🟡 Expiry Warning ({delta} days left)"
                            text_color = "#92400E"
                            bg_color = "#FEF3C7"
                        else:
                            status_text = f"🟢 Active ({delta} days left)"
                    except Exception:
                        pass
                else:
                    status_text = "⚪ Validity Unknown"
                    text_color = self.THEME["muted"]
                    bg_color = self.THEME["surface_alt"]

                row_fr = ctk.CTkFrame(self.expiration_scroll, fg_color=bg_color, height=45)
                row_fr.pack(fill="x", pady=3, padx=5)
                
                ctk.CTkLabel(row_fr, text=f"{supplier} - {product}", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
                
                # Add request button next to status label
                btn_req = ctk.CTkButton(row_fr, text="✉ Request Extension", width=120, height=24, fg_color="#1f538d", hover_color="#153e6b", font=ctk.CTkFont(size=11), command=lambda record=r: self.open_extension_request_popup(record))
                btn_req.pack(side="right", padx=10)
                
                ctk.CTkLabel(row_fr, text=status_text, text_color=text_color, font=ctk.CTkFont(weight="bold")).pack(side="right", padx=10)

        # 3. Populate Production Planner progress bars
        for widget in self.timeline_display_scroll.winfo_children():
            widget.destroy()

        category = self.timeline_category_cb.get()
        filtered = self.extracted_data
        if category and category != "All":
            filtered = [r for r in self.extracted_data if (r.get("product") or "").strip().lower() == category.lower()]

        # Parse Order Date
        order_date_str = self.order_date_entry.get().strip()
        try:
            order_date = datetime.datetime.strptime(order_date_str, "%Y-%m-%d").date()
        except Exception:
            order_date = today

        if not filtered:
            ctk.CTkLabel(self.timeline_display_scroll, text="No quotes matching selected category.", text_color="grey").pack(pady=20)
            return

        # Render visual Gantt Chart using Matplotlib
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        
        fig, ax = plt.subplots(figsize=(6, 3.5), dpi=100)
        fig.patch.set_facecolor(self.THEME["surface"])
        ax.set_facecolor(self.THEME["surface"])
        
        y_ticks = []
        y_labels = []
        
        for idx, r_item in enumerate(filtered):
            sup = self.clean_supplier_name(r_item.get("supplier"))
            prod = r_item.get("product") or "Product"
            lead_str = r_item.get("lead_time") or "30 days"
            
            # Parse lead time
            lead_days = 30
            import re
            lead_match = re.search(r'([0-9.]+)\s*(day|week|month)', lead_str.lower())
            if lead_match:
                try:
                    val = float(lead_match.group(1))
                    unit = lead_match.group(2)
                    if "week" in unit: lead_days = val * 7
                    elif "month" in unit: lead_days = val * 30
                    else: lead_days = val
                except: pass
            
            # Production Phase vs Shipping Phase (FOB to DDP delivery estimation)
            start_prod = 0
            end_prod = lead_days
            start_ship = lead_days
            end_ship = lead_days + 30 # default 30 days transit
            
            ax.barh(idx, end_prod - start_prod, left=start_prod, height=0.4, color=self.THEME["primary"], align='center')
            ax.barh(idx, end_ship - start_ship, left=start_ship, height=0.4, color=self.THEME["success"], align='center')
            
            y_ticks.append(idx)
            label_text = f"{sup[:15]} ({prod[:10]})"
            y_labels.append(label_text)
            
        ax.set_yticks(y_ticks)
        ax.set_yticklabels(y_labels, color=self.THEME["text"], fontsize=8)
        ax.set_xlabel("Timeline (Days from Order Date)", color=self.THEME["muted"], fontsize=9)
        ax.xaxis.label.set_color(self.THEME["muted"])
        ax.tick_params(colors=self.THEME["muted"], labelsize=8)
        ax.grid(axis="x", color=self.THEME["border"], linestyle="--", linewidth=0.8)
        
        # Remove borders
        for spine in ax.spines.values():
            spine.set_visible(False)
            
        # Add legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor=self.THEME["primary"], label='Production'),
            Patch(facecolor=self.THEME["success"], label='Transit/Sea Freight')
        ]
        leg = ax.legend(handles=legend_elements, loc='upper right', facecolor=self.THEME["surface"], edgecolor='none')
        for text in leg.get_texts():
            text.set_color(self.THEME["text"])
            text.set_fontsize(7)
            
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.timeline_display_scroll)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True, pady=5)
        plt.close(fig)

    def update_preview_gallery_bar(self, row_id, vals):
        for widget in self.preview_gallery_bar.winfo_children():
            widget.destroy()

        if not vals:
            return

        attached_media_str = ""
        for r in self.extracted_data:
            if str(r.get("id")) == str(row_id):
                attached_media_str = r.get("attached_media") or ""
                break

        options = [("📄 Quote", "quote", vals[2])]
        
        if attached_media_str:
            media_files = attached_media_str.split(";")
            img_idx = 1
            vid_idx = 1
            for m in media_files:
                if m.strip():
                    path = m.strip()
                    ext = os.path.splitext(path)[1].lower()
                    if ext in [".mp4", ".avi", ".mov", ".mkv", ".mpg", ".mpeg", ".wmv"]:
                        options.append((f"🎥 Video {vid_idx}", "video", path))
                        vid_idx += 1
                    elif ext in [".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"]:
                        options.append((f"🖼 Image {img_idx}", "image", path))
                        img_idx += 1
                    else:
                        options.append((f"📎 File {img_idx}", "other", path))
                        img_idx += 1

        self.gallery_buttons = {}
        for idx, (label, media_type, path) in enumerate(options):
            btn = ctk.CTkButton(
                self.preview_gallery_bar,
                text=label,
                width=80,
                height=26,
                fg_color="#1f538d" if idx == 0 else "#3c3c3c",
                command=lambda p=path, t=media_type, idx=idx: self.select_gallery_item(p, t, idx)
            )
            btn.pack(side="left", padx=3, pady=2)
            self.gallery_buttons[idx] = btn

    def select_gallery_item(self, path, media_type, btn_idx):
        for idx, btn in self.gallery_buttons.items():
            if idx == btn_idx:
                btn.configure(fg_color="#1f538d")
            else:
                btn.configure(fg_color="#3c3c3c")

        self.preview_text_box.pack_forget()
        self.preview_image_lbl.pack_forget()
        for widget in self.preview_display_frame.winfo_children():
            if widget not in [self.preview_image_lbl, self.preview_text_box]:
                widget.destroy()

        if media_type == "quote":
            self.show_preview(path)
            return

        if not path or not os.path.exists(path):
            self.show_preview_message(f"Attached file not found:\n{os.path.basename(path)}")
            self.btn_open_external.configure(state="disabled")
            return

        self.current_preview_path = path
        self.btn_open_external.configure(state="normal")

        if media_type == "video":
            self.show_video_play_overlay(path)
        elif media_type == "image":
            self.render_image_preview(path)
        else:
            self.show_preview_message(f"Attached file:\n{os.path.basename(path)}\n(Click button below to open.)")

    def show_video_play_overlay(self, path):
        for widget in self.preview_display_frame.winfo_children():
            if widget not in [self.preview_image_lbl, self.preview_text_box]:
                widget.destroy()
                
        self.preview_image_lbl.pack_forget()
        self.preview_text_box.pack_forget()
        
        play_btn = ctk.CTkButton(self.preview_display_frame, text="▶ Play Video", font=ctk.CTkFont(size=16, weight="bold"), fg_color="#8c1c1c", hover_color="#6e1313", height=50, command=lambda: os.startfile(path))
        play_btn.pack(expand=True, pady=100, padx=20)

    def attach_media_to_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Warning", "Please select a quotation row to attach media to.")
            return

        item = sel[0]
        vals = self.tree.item(item, "values")
        row_id = vals[0]
        supplier = vals[3]
        product = vals[4]

        files = filedialog.askopenfilenames(
            title="Attach Images/Videos to Quotation",
            filetypes=[("Media Files", "*.png;*.jpg;*.jpeg;*.gif;*.webp;*.bmp;*.mp4;*.avi;*.mov;*.mkv;*.wmv")]
        )
        if not files:
            return

        # Setup top-level options window
        options_win = ctk.CTkToplevel(self)
        options_win.title("Attachment Scope")
        options_win.geometry("450x260")
        options_win.resizable(False, False)
        options_win.attributes("-topmost", True)
        
        # Center the window
        options_win.update_idletasks()
        width = options_win.winfo_width()
        height = options_win.winfo_height()
        x = (options_win.winfo_screenwidth() // 2) - (width // 2)
        y = (options_win.winfo_screenheight() // 2) - (height // 2)
        options_win.geometry(f'+{x}+{y}')

        ctk.CTkLabel(options_win, text="Select Attachment Scope:", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 10), padx=20, anchor="w")
        
        scope_var = tk.IntVar(value=1)
        
        r1 = ctk.CTkRadioButton(options_win, text="Selected quote row only", variable=scope_var, value=1)
        r1.pack(pady=5, padx=30, anchor="w")

        r2 = ctk.CTkRadioButton(options_win, text=f"All quotes from {supplier} for '{product}'", variable=scope_var, value=2)
        r2.pack(pady=5, padx=30, anchor="w")

        r3 = ctk.CTkRadioButton(options_win, text=f"All quotes from {supplier} (All products)", variable=scope_var, value=3)
        r3.pack(pady=5, padx=30, anchor="w")

        btn_row = ctk.CTkFrame(options_win, fg_color="transparent")
        btn_row.pack(fill="x", side="bottom", pady=15, padx=20)

        def proceed():
            choice = scope_var.get()
            options_win.destroy()
            self.execute_media_attachment(row_id, supplier, product, files, choice)

        def cancel():
            options_win.destroy()

        btn_cancel = ctk.CTkButton(btn_row, text="Cancel", width=100, fg_color="#3c3c3c", command=cancel)
        btn_cancel.pack(side="right", padx=5)

        btn_confirm = ctk.CTkButton(btn_row, text="Attach Media", width=120, fg_color="#1f538d", command=proceed)
        btn_confirm.pack(side="right", padx=5)

    def execute_media_attachment(self, row_id, supplier, product, files, choice):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        # Determine target query based on choice
        if choice == 1:
            c.execute("SELECT id, attached_media FROM extracted_quotes WHERE id = ?", (row_id,))
        elif choice == 2:
            c.execute("SELECT id, attached_media FROM extracted_quotes WHERE supplier = ? AND product = ?", (supplier, product))
        else:
            c.execute("SELECT id, attached_media FROM extracted_quotes WHERE supplier = ?", (supplier,))
            
        rows = c.fetchall()
        
        updated_count = 0
        for r_id, existing_media_str in rows:
            media_list = []
            if existing_media_str:
                media_list = [f.strip() for f in existing_media_str.split(";") if f.strip()]

            for f in files:
                norm_f = os.path.abspath(f).replace("\\", "/")
                if norm_f not in media_list:
                    media_list.append(norm_f)

            new_val = ";".join(media_list)
            c.execute("UPDATE extracted_quotes SET attached_media = ? WHERE id = ?", (new_val, r_id))
            updated_count += 1

        conn.commit()
        conn.close()

        self.load_all_quotes_from_db()
        
        scope_lbls = {
            1: "selected quote only",
            2: f"all {supplier} quotes for '{product}'",
            3: f"all quotes from {supplier}"
        }
        messagebox.showinfo("Success", f"Attached {len(files)} media files to {updated_count} quotation(s) ({scope_lbls[choice]}) successfully!")

    def move_file_to_synced(self, name):
        for idx in range(self.files_box_unsynced.size()):
            item_text = self.files_box_unsynced.get(idx)
            if name in item_text:
                self.files_box_unsynced.delete(idx)
                break
        self.files_box_synced.insert(tk.END, f"✅ {name}")

    def open_extension_request_popup(self, r):
        # Get Details
        supplier = r.get("supplier") or "Supplier"
        product = r.get("product") or "Product"
        val_date = r.get("validity_date") or "N/A"
        price = r.get("price") or "N/A"
        currency = r.get("currency") or "USD"
        
        # Query contact email
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT contact_info FROM supplier_contacts WHERE supplier = ?", (supplier,))
        contact_row = c.fetchone()
        conn.close()
        
        email = ""
        if contact_row and contact_row[0]:
            import re
            emails = re.findall(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', contact_row[0])
            if emails:
                email = emails[0]
                
        # Setup top-level window
        popup = ctk.CTkToplevel(self)
        popup.title("Request Quote Validity Extension")
        popup.geometry("600x520")
        popup.resizable(False, False)
        popup.attributes("-topmost", True)
        
        # Center the window
        popup.update_idletasks()
        width = popup.winfo_width()
        height = popup.winfo_height()
        x = (popup.winfo_screenwidth() // 2) - (width // 2)
        y = (popup.winfo_screenheight() // 2) - (height // 2)
        popup.geometry(f'+{x}+{y}')
        
        # Title Label
        ctk.CTkLabel(popup, text="✉ Quotation Extension Email Draft", font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(15, 5), padx=20, anchor="w")
        ctk.CTkLabel(popup, text=f"Generate validity extension request for {supplier} ({product})", text_color="grey").pack(pady=(0, 10), padx=20, anchor="w")
        
        # Recipient email row
        email_frame = ctk.CTkFrame(popup, fg_color="transparent")
        email_frame.pack(fill="x", padx=20, pady=5)
        
        ctk.CTkLabel(email_frame, text="To: ", width=50, anchor="w").pack(side="left")
        email_entry = ctk.CTkEntry(email_frame, width=350)
        email_entry.pack(side="left", fill="x", expand=True)
        if email:
            email_entry.insert(0, email)
            
        # Subject row
        subject_frame = ctk.CTkFrame(popup, fg_color="transparent")
        subject_frame.pack(fill="x", padx=20, pady=5)
        
        ctk.CTkLabel(subject_frame, text="Subject: ", width=50, anchor="w").pack(side="left")
        subject_entry = ctk.CTkEntry(subject_frame, width=350)
        subject_entry.pack(side="left", fill="x", expand=True)
        subject_entry.insert(0, f"Quote Validity Extension Request: {product} - {supplier}")
        
        # Text box for email body
        text_box = ctk.CTkTextbox(popup, height=280, font=("Consolas", 11))
        text_box.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Button container row
        btn_row = ctk.CTkFrame(popup, fg_color="transparent")
        btn_row.pack(fill="x", side="bottom", pady=15, padx=20)
        
        def run_generation():
            btn_regen.configure(state="disabled", text="Generating...")
            popup.update()
            
            ai_prompt = f"""
            Write a professional and polite counter-extension email to request a quotation validity extension.
            
            Details:
            - Supplier Name: {supplier}
            - Product Item: {product}
            - Current Expiry Date: {val_date}
            - Quoted Price: {price} {currency}
            
            Instructions:
            - Be concise, polite, and formal.
            - State that we are finalizing our sourcing evaluation and require a 30-day extension of the quote validity.
            - Address the supplier professionally.
            - Output ONLY the email body. Do not include subject line or placeholders like [Your Name]. Use 'Procurement Team' as the sender signature.
            """
            
            try:
                email_body = self.generate_with_fallback([], ai_prompt, json_response=False)
                self.after(0, lambda text=email_body: update_text_box(text))
            except Exception as e:
                self.after(0, lambda err=e: update_text_box(f"Error generating email: {err}"))
                
        def update_text_box(text):
            text_box.delete("1.0", tk.END)
            text_box.insert("1.0", text)
            btn_regen.configure(state="normal", text="✍ Regenerate")
            
        def copy_email():
            content = text_box.get("1.0", tk.END).strip()
            if content:
                self.clipboard_clear()
                self.clipboard_append(content)
                messagebox.showinfo("Copied", "Email body copied to clipboard successfully!", parent=popup)
                
        def launch_client():
            to_email = email_entry.get().strip()
            subj = subject_entry.get().strip()
            body = text_box.get("1.0", tk.END).strip()
            
            import urllib.parse
            import webbrowser
            mailto_url = f"mailto:{to_email}?subject={urllib.parse.quote(subj)}&body={urllib.parse.quote(body)}"
            webbrowser.open(mailto_url)
            
        btn_close = ctk.CTkButton(btn_row, text="Close", width=80, fg_color="#3c3c3c", command=popup.destroy)
        btn_close.pack(side="left", padx=5)
        
        btn_regen = ctk.CTkButton(btn_row, text="✍ Generate", width=120, command=lambda: threading.Thread(target=run_generation, daemon=True).start())
        btn_regen.pack(side="right", padx=5)
        
        btn_copy = ctk.CTkButton(btn_row, text="📋 Copy", width=100, fg_color="#6e4513", hover_color="#52320b", command=copy_email)
        btn_copy.pack(side="right", padx=5)
        
        btn_launch = ctk.CTkButton(btn_row, text="🚀 Launch Mail", width=120, fg_color="#1f538d", hover_color="#153e6b", command=launch_client)
        btn_launch.pack(side="right", padx=5)
        
        # Start first generation automatically
        threading.Thread(target=run_generation, daemon=True).start()

    def setup_landed_cost_tab(self):
        tab_landed = self.logistics_tabview.tab("📦 Landed Cost Simulator")
        tab_landed.grid_columnconfigure(0, weight=1)
        tab_landed.grid_columnconfigure(1, weight=3)
        tab_landed.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_landed, text="Landed Cost & Logistics Freight Simulator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Input Parameters ---
        left_frame = ctk.CTkFrame(tab_landed)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(left_frame, text="⚙ Simulator Parameters", font=ctk.CTkFont(size=15, weight="bold")).pack(pady=10, padx=15, anchor="w")

        # Quantity Input
        qty_lbl = ctk.CTkLabel(left_frame, text="Order Size (pcs):", font=ctk.CTkFont(size=12))
        qty_lbl.pack(pady=(5, 0), padx=15, anchor="w")
        self.sim_qty_entry = ctk.CTkEntry(left_frame, placeholder_text="100,000")
        self.sim_qty_entry.pack(pady=2, padx=15, fill="x")
        self.sim_qty_entry.insert(0, "100000")

        # Freight Mode Segmented Button
        mode_lbl = ctk.CTkLabel(left_frame, text="Freight Shipment Mode:", font=ctk.CTkFont(size=12))
        mode_lbl.pack(pady=(10, 0), padx=15, anchor="w")
        self.sim_freight_mode = ctk.CTkSegmentedButton(left_frame, values=["LCL (per CBM)", "FCL 20GP", "FCL 40HQ"], command=lambda m: self.on_freight_mode_changed(m))
        self.sim_freight_mode.pack(pady=2, padx=15, fill="x")
        self.sim_freight_mode.set("LCL (per CBM)")

        # Freight Rate Input
        self.rate_lbl_text = tk.StringVar(value="LCL Rate per CBM ($):")
        rate_lbl = ctk.CTkLabel(left_frame, textvariable=self.rate_lbl_text, font=ctk.CTkFont(size=12))
        rate_lbl.pack(pady=(10, 0), padx=15, anchor="w")
        self.sim_rate_entry = ctk.CTkEntry(left_frame)
        self.sim_rate_entry.pack(pady=2, padx=15, fill="x")
        self.sim_rate_entry.insert(0, "120")

        # Duty Rate Percent Slider/Input
        duty_lbl = ctk.CTkLabel(left_frame, text="Customs Duty Rate (%):", font=ctk.CTkFont(size=12))
        duty_lbl.pack(pady=(10, 0), padx=15, anchor="w")
        
        duty_row = ctk.CTkFrame(left_frame, fg_color="transparent")
        duty_row.pack(fill="x", padx=15, pady=2)
        
        self.sim_duty_slider = ctk.CTkSlider(duty_row, from_=0, to=30, number_of_steps=60, command=lambda v: self.on_duty_slider_changed(v))
        self.sim_duty_slider.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.sim_duty_slider.set(6.5)
        
        self.sim_duty_val_lbl = ctk.CTkLabel(duty_row, text="6.5%", width=45)
        self.sim_duty_val_lbl.pack(side="right")

        # Local Customs / Clearance Flat Fee
        local_lbl = ctk.CTkLabel(left_frame, text="Local Fees & Clearance (Flat $):", font=ctk.CTkFont(size=12))
        local_lbl.pack(pady=(10, 0), padx=15, anchor="w")
        self.sim_local_entry = ctk.CTkEntry(left_frame)
        self.sim_local_entry.pack(pady=2, padx=15, fill="x")
        self.sim_local_entry.insert(0, "350")

        # Recalculate Button
        self.btn_calc_sim = ctk.CTkButton(left_frame, text="📊 Run Simulation", fg_color="#1f538d", hover_color="#153e6b", command=self.update_landed_cost_tab)
        self.btn_calc_sim.pack(pady=25, padx=15, fill="x")

        # --- RIGHT PANEL: Comparison Grid ---
        right_frame = ctk.CTkFrame(tab_landed)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📊 Landed Cost Comparison (FOB to DDP)", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        # Treeview Comparison table
        cols = ("supplier", "product", "unit_fob", "est_cbm", "freight_pc", "duty_pc", "local_pc", "landed_pc", "total_cost")
        
        scroll_y = ttk.Scrollbar(right_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(right_frame, orient="horizontal")
        
        self.sim_tree = ttk.Treeview(
            right_frame,
            columns=cols,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )
        
        scroll_y.config(command=self.sim_tree.yview)
        scroll_x.config(command=self.sim_tree.xview)
        
        scroll_y.grid(row=1, column=1, sticky="ns")
        scroll_x.grid(row=2, column=0, sticky="ew")
        self.sim_tree.grid(row=1, column=0, sticky="nsew", padx=(15, 0))

        # Setup Headings
        self.sim_tree.heading("supplier", text="Supplier")
        self.sim_tree.heading("product", text="Product")
        self.sim_tree.heading("unit_fob", text="FOB ($/pc)")
        self.sim_tree.heading("est_cbm", text="Est CBM")
        self.sim_tree.heading("freight_pc", text="Freight/pc")
        self.sim_tree.heading("duty_pc", text="Duty/pc")
        self.sim_tree.heading("local_pc", text="Local/pc")
        self.sim_tree.heading("landed_pc", text="Landed ($/pc)")
        self.sim_tree.heading("total_cost", text="Total Landed")

        # Setup Widths
        self.sim_tree.column("supplier", width=140, anchor="w")
        self.sim_tree.column("product", width=120, anchor="w")
        self.sim_tree.column("unit_fob", width=80, anchor="center")
        self.sim_tree.column("est_cbm", width=70, anchor="center")
        self.sim_tree.column("freight_pc", width=80, anchor="center")
        self.sim_tree.column("duty_pc", width=80, anchor="center")
        self.sim_tree.column("local_pc", width=80, anchor="center")
        self.sim_tree.column("landed_pc", width=95, anchor="center")
        self.sim_tree.column("total_cost", width=105, anchor="center")

        self.sim_tree.tag_configure("winner", background="#DCFCE7", foreground="#14532D")

    def on_freight_mode_changed(self, mode):
        if mode == "LCL (per CBM)":
            self.rate_lbl_text.set("LCL Rate per CBM ($):")
            self.sim_rate_entry.delete(0, tk.END)
            self.sim_rate_entry.insert(0, "120")
        elif mode == "FCL 20GP":
            self.rate_lbl_text.set("20GP Container Cost ($):")
            self.sim_rate_entry.delete(0, tk.END)
            self.sim_rate_entry.insert(0, "3000")
        else:
            self.rate_lbl_text.set("40HQ Container Cost ($):")
            self.sim_rate_entry.delete(0, tk.END)
            self.sim_rate_entry.insert(0, "4500")
        self.update_landed_cost_tab()

    def on_duty_slider_changed(self, val):
        self.sim_duty_val_lbl.configure(text=f"{val:.1f}%")
        self.update_landed_cost_tab()

    def parse_packing_metrics(self, packing_str):
        pcs_per_ctn = 1000
        unit_cbm = 0.05
        
        if not packing_str or packing_str == "N/A":
            return pcs_per_ctn, unit_cbm
            
        import re
        packing_lower = packing_str.lower()
        
        # 1. Try to find CBM
        cbm_match = re.search(r'([0-9.]+)\s*(cbm|m3|cubic|meas)', packing_lower)
        if cbm_match:
            try:
                unit_cbm = float(cbm_match.group(1))
            except Exception:
                pass
        else:
            dim_match = re.findall(r'([0-9.]+)\s*[\*xX]\s*([0-9.]+)\s*[\*xX]\s*([0-9.]+)', packing_lower)
            if dim_match:
                try:
                    w, h, d = map(float, dim_match[0])
                    unit_cbm = (w * h * d) / 1000000.0
                except Exception:
                    pass
                    
        # 2. Try to find pieces per carton
        ctn_match = re.search(r'([0-9,]+)\s*(pcs|pieces)?\s*(/|per)\s*(ctn|carton|box|case)', packing_lower)
        if ctn_match:
            try:
                num_str = ctn_match.group(1).replace(",", "")
                pcs_per_ctn = int(num_str)
            except Exception:
                pass
        else:
            pcs_match = re.findall(r'([0-9,]+)\s*(pcs|pieces|bags)?/ctn', packing_lower)
            if pcs_match:
                try:
                    pcs_per_ctn = int(pcs_match[0][0].replace(",", ""))
                except Exception:
                    pass
                    
        return pcs_per_ctn, unit_cbm

    def update_landed_cost_tab(self):
        for item in self.sim_tree.get_children():
            self.sim_tree.delete(item)

        if not self.extracted_data:
            return

        import math
        try:
            order_qty = int(self.sim_qty_entry.get().replace(",", "").strip())
        except Exception:
            order_qty = 100000
            
        try:
            freight_rate = float(self.sim_rate_entry.get().replace(",", "").strip())
        except Exception:
            freight_rate = 120.0

        duty_percent = self.sim_duty_slider.get()

        try:
            local_flat = float(self.sim_local_entry.get().replace(",", "").strip())
        except Exception:
            local_flat = 350.0

        freight_mode = self.sim_freight_mode.get()

        sim_rows = []
        for r in self.extracted_data:
            fob = r.get("price")
            if fob is None or fob == "N/A":
                continue
                
            supplier = self.clean_supplier_name(r.get("supplier"))
            product = r.get("product") or "Product"
            packing = r.get("packing") or "N/A"
            
            pcs_per_ctn, unit_cbm = self.parse_packing_metrics(packing)
            
            total_ctns = math.ceil(order_qty / pcs_per_ctn)
            total_cbm = total_ctns * unit_cbm
            
            total_fob = order_qty * fob
            
            if freight_mode == "LCL (per CBM)":
                total_freight = total_cbm * freight_rate
            elif freight_mode == "FCL 20GP":
                containers = math.ceil(total_cbm / 28.0)
                total_freight = containers * freight_rate
            else:
                containers = math.ceil(total_cbm / 68.0)
                total_freight = containers * freight_rate
                
            total_duty = total_fob * (duty_percent / 100.0)
            total_local = local_flat
            
            total_landed = total_fob + total_freight + total_duty + total_local
            landed_pc = total_landed / order_qty
            
            freight_pc = total_freight / order_qty
            duty_pc = total_duty / order_qty
            local_pc = total_local / order_qty
            
            sim_rows.append({
                "supplier": supplier,
                "product": product,
                "unit_fob": fob,
                "est_cbm": total_cbm,
                "freight_pc": freight_pc,
                "duty_pc": duty_pc,
                "local_pc": local_pc,
                "landed_pc": landed_pc,
                "total_cost": total_landed
            })

        if not sim_rows:
            return

        sim_rows.sort(key=lambda x: x["landed_pc"])
        best_landed = sim_rows[0]["landed_pc"]

        for r in sim_rows:
            tag = ""
            if abs(r["landed_pc"] - best_landed) < 1e-7:
                tag = "winner"
                
            self.sim_tree.insert(
                "",
                tk.END,
                values=(
                    r["supplier"],
                    r["product"],
                    f"${r['unit_fob']:.5f}" if r["unit_fob"] < 0.1 else f"${r['unit_fob']:.2f}",
                    f"{r['est_cbm']:.2f} m³",
                    f"${r['freight_pc']:.4f}",
                    f"${r['duty_pc']:.4f}",
                    f"${r['local_pc']:.4f}",
                    f"${r['landed_pc']:.4f}",
                    f"${r['total_cost']:.2f}"
                ),
                tags=(tag,)
            )

    def setup_purchase_optimizer_tab(self):
        tab_opt = self.logistics_tabview.tab("🎯 Purchase Optimizer")
        tab_opt.grid_columnconfigure(0, weight=1)
        tab_opt.grid_columnconfigure(1, weight=1)
        tab_opt.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_opt, text="Multi-Product Consolidated Purchase Optimizer", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Products & Quantities Selection ---
        left_frame = ctk.CTkFrame(tab_opt)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="📦 Select Products & Quantities", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        self.opt_scroll = ctk.CTkScrollableFrame(left_frame, fg_color=self.THEME["surface_soft"])
        self.opt_scroll.grid(row=1, column=0, padx=15, pady=5, sticky="nsew")
        self.opt_scroll.grid_columnconfigure(1, weight=1)

        # Bottom logistics controls for optimizer
        opt_ctrls = ctk.CTkFrame(left_frame, fg_color="transparent")
        opt_ctrls.grid(row=2, column=0, padx=15, pady=10, sticky="ew")

        ctk.CTkLabel(opt_ctrls, text="Freight per CBM ($):", font=ctk.CTkFont(size=11)).pack(side="left", padx=5)
        self.opt_freight_rate_entry = ctk.CTkEntry(opt_ctrls, width=70)
        self.opt_freight_rate_entry.pack(side="left", padx=5)
        self.opt_freight_rate_entry.insert(0, "120")

        ctk.CTkLabel(opt_ctrls, text="Duty (%):", font=ctk.CTkFont(size=11)).pack(side="left", padx=5)
        self.opt_duty_entry = ctk.CTkEntry(opt_ctrls, width=60)
        self.opt_duty_entry.pack(side="left", padx=5)
        self.opt_duty_entry.insert(0, "6.5")

        self.btn_run_opt = ctk.CTkButton(left_frame, text="⚡ Run Optimization", fg_color="#1f538d", hover_color="#153e6b", command=self.run_purchase_optimization)
        self.btn_run_opt.grid(row=3, column=0, padx=15, pady=(5, 15), sticky="ew")

        # --- RIGHT PANEL: Optimization Report ---
        right_frame = ctk.CTkFrame(tab_opt)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="⚡ Optimization Results & Strategy", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        self.opt_results_scroll = ctk.CTkScrollableFrame(right_frame, fg_color=self.THEME["surface_soft"])
        self.opt_results_scroll.grid(row=1, column=0, padx=15, pady=10, sticky="nsew")
        self.opt_results_scroll.grid_columnconfigure(0, weight=1)

        self.make_empty_state(
            self.opt_results_scroll,
            "Ready for optimization",
            "Select products, confirm quantities, then run the optimizer to compare split sourcing versus supplier consolidation.",
        )

    def update_purchase_optimizer_tab(self):
        current_selections = {}
        if hasattr(self, 'opt_checkboxes'):
            for p, cb in self.opt_checkboxes.items():
                current_selections[p] = (cb.get(), self.opt_qty_entries[p].get())

        for w in self.opt_scroll.winfo_children():
            w.destroy()

        self.opt_checkboxes = {}
        self.opt_qty_entries = {}

        unique_prods = set()
        for r in self.extracted_data:
            p = (r.get("product") or "").strip().title()
            if p:
                unique_prods.add(p)

        if not unique_prods:
            self.make_empty_state(self.opt_scroll, "No products found", "Extract or sync quotes first so the optimizer can build a product list.")
            return

        for idx, p in enumerate(sorted(list(unique_prods))):
            row_fr = ctk.CTkFrame(self.opt_scroll, fg_color="transparent")
            row_fr.pack(fill="x", pady=2, padx=5)

            cb_var = tk.IntVar(value=1)
            qty_val = "50000"

            if p in current_selections:
                cb_var.set(current_selections[p][0])
                qty_val = current_selections[p][1]

            cb = ctk.CTkCheckBox(row_fr, text=p, variable=cb_var, font=ctk.CTkFont(size=12))
            cb.pack(side="left", padx=5)

            qty_ent = ctk.CTkEntry(row_fr, width=90, font=ctk.CTkFont(size=11))
            qty_ent.pack(side="right", padx=5)
            qty_ent.insert(0, qty_val)

            self.opt_checkboxes[p] = cb_var
            self.opt_qty_entries[p] = qty_ent

    def run_purchase_optimization(self):
        for w in self.opt_results_scroll.winfo_children():
            w.destroy()

        selected_items = {}
        for p, cb in self.opt_checkboxes.items():
            if cb.get() == 1:
                try:
                    qty = int(self.opt_qty_entries[p].get().replace(",", "").strip())
                    if qty > 0:
                        selected_items[p.lower()] = qty
                except Exception:
                    pass

        if not selected_items:
            ctk.CTkLabel(self.opt_results_scroll, text="❌ Please select at least one product with a valid quantity.", text_color="#ffa6a6").pack(pady=20)
            return

        try:
            freight_rate = float(self.opt_freight_rate_entry.get().strip())
        except Exception:
            freight_rate = 120.0
        try:
            duty_percent = float(self.opt_duty_entry.get().strip())
        except Exception:
            duty_percent = 6.5

        quotes_by_product = {}
        for r in self.extracted_data:
            p = (r.get("product") or "").strip().lower()
            if p in selected_items:
                if p not in quotes_by_product:
                    quotes_by_product[p] = []
                quotes_by_product[p].append(r)

        split_items = []
        total_split_fob = 0.0
        total_split_cbm = 0.0
        suppliers_used = set()

        for p_name, qty in selected_items.items():
            options = quotes_by_product.get(p_name, [])
            if not options:
                continue
            valid_opts = [o for o in options if o.get("price") is not None and o.get("price") != "N/A"]
            if not valid_opts:
                continue
            valid_opts.sort(key=lambda x: x.get("price"))
            best_opt = valid_opts[0]
            
            fob = best_opt.get("price")
            supplier = self.clean_supplier_name(best_opt.get("supplier"))
            packing = best_opt.get("packing") or "N/A"
            
            pcs_per_ctn, unit_cbm = self.parse_packing_metrics(packing)
            import math
            ctns = math.ceil(qty / pcs_per_ctn)
            cbm = ctns * unit_cbm
            
            item_cost = qty * fob
            total_split_fob += item_cost
            total_split_cbm += cbm
            suppliers_used.add(supplier)
            
            split_items.append({
                "product": p_name.title(),
                "supplier": supplier,
                "qty": qty,
                "fob": fob,
                "cost": item_cost,
                "cbm": cbm
            })

        if not split_items:
            ctk.CTkLabel(self.opt_results_scroll, text="❌ No valid quotes available for selected items.", text_color="#ffa6a6").pack(pady=20)
            return

        split_freight = total_split_cbm * freight_rate
        split_duty = total_split_fob * (duty_percent / 100.0)
        split_local = len(suppliers_used) * 350.0
        total_split_cost = total_split_fob + split_freight + split_duty + split_local

        all_suppliers = set(self.clean_supplier_name(r.get("supplier")) for r in self.extracted_data if r.get("supplier"))
        
        consolidation_results = []
        for s in all_suppliers:
            s_items = []
            s_fob_total = 0.0
            s_cbm_total = 0.0
            can_quote_all = True
            
            for p_name, qty in selected_items.items():
                options = quotes_by_product.get(p_name, [])
                s_options = [o for o in options if self.clean_supplier_name(o.get("supplier")) == s and o.get("price") is not None and o.get("price") != "N/A"]
                if not s_options:
                    can_quote_all = False
                    break
                s_options.sort(key=lambda x: x.get("price"))
                opt = s_options[0]
                
                fob = opt.get("price")
                packing = opt.get("packing") or "N/A"
                pcs_per_ctn, unit_cbm = self.parse_packing_metrics(packing)
                import math
                ctns = math.ceil(qty / pcs_per_ctn)
                cbm = ctns * unit_cbm
                
                item_cost = qty * fob
                s_fob_total += item_cost
                s_cbm_total += cbm
                
                s_items.append({
                    "product": p_name.title(),
                    "qty": qty,
                    "fob": fob,
                    "cost": item_cost,
                    "cbm": cbm
                })
                
            if can_quote_all:
                s_freight = s_cbm_total * freight_rate
                s_duty = s_fob_total * (duty_percent / 100.0)
                s_local = 350.0
                s_total = s_fob_total + s_freight + s_duty + s_local
                
                consolidation_results.append({
                    "supplier": s,
                    "items": s_items,
                    "fob_total": s_fob_total,
                    "cbm_total": s_cbm_total,
                    "freight": s_freight,
                    "duty": s_duty,
                    "local": s_local,
                    "total_cost": s_total
                })

        rec_fr = ctk.CTkFrame(self.opt_results_scroll, fg_color=self.THEME["info_soft"], border_color=self.THEME["info_border"], border_width=1, corner_radius=8, height=70)
        rec_fr.pack(fill="x", pady=(0, 15), padx=5)
        
        best_consolidated = None
        if consolidation_results:
            consolidation_results.sort(key=lambda x: x["total_cost"])
            best_consolidated = consolidation_results[0]

        if best_consolidated and best_consolidated["total_cost"] < total_split_cost:
            savings = total_split_cost - best_consolidated["total_cost"]
            rec_text = f"💡 RECOMMENDATION: Consolidate with {best_consolidated['supplier']}!\nConsolidating saves you ${savings:,.2f} in split freight & handling fees."
            rec_color = self.THEME["success"]
        else:
            if best_consolidated:
                savings = best_consolidated["total_cost"] - total_split_cost
                rec_text = f"💡 RECOMMENDATION: Split purchases across suppliers!\nSplitting is ${savings:,.2f} cheaper than consolidating with {best_consolidated['supplier']}."
            else:
                rec_text = f"💡 RECOMMENDATION: Split purchases across suppliers!\nNo single supplier can fulfill all selected products."
            rec_color = self.THEME["warning"]
            
        ctk.CTkLabel(rec_fr, text=rec_text, font=ctk.CTkFont(size=13, weight="bold"), text_color=rec_color, justify="left").pack(padx=15, pady=12, fill="both")

        split_fr = ctk.CTkFrame(self.opt_results_scroll, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        split_fr.pack(fill="x", pady=5, padx=5)
        
        ctk.CTkLabel(split_fr, text="📊 Strategy A: Split Sourcing (Cheapest FOB per Item)", font=ctk.CTkFont(weight="bold", size=14)).pack(pady=8, padx=15, anchor="w")
        
        for item in split_items:
            row_lbl = f"• Buy {item['qty']:,}x {item['product']} from {item['supplier']} @ ${item['fob']:.4f}/pc — Cost: ${item['cost']:,.2f} ({item['cbm']:.2f} m³)"
            ctk.CTkLabel(split_fr, text=row_lbl, font=ctk.CTkFont(size=11), text_color="grey").pack(pady=1, padx=25, anchor="w")
            
        summary_split = f"Items Cost: ${total_split_fob:,.2f} | Freight: ${split_freight:,.2f} | Duties: ${split_duty:,.2f} | Handling: ${split_local:,.2f}"
        ctk.CTkLabel(split_fr, text=summary_split, font=ctk.CTkFont(size=11, weight="bold"), text_color=self.THEME["text"]).pack(pady=(5, 2), padx=15, anchor="w")
        ctk.CTkLabel(split_fr, text=f"TOTAL DDP SPLIT COST: ${total_split_cost:,.2f}", font=ctk.CTkFont(size=13, weight="bold"), text_color=self.THEME["primary"]).pack(pady=(2, 8), padx=15, anchor="w")

        if consolidation_results:
            con_fr = ctk.CTkFrame(self.opt_results_scroll, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
            con_fr.pack(fill="x", pady=10, padx=5)
            
            ctk.CTkLabel(con_fr, text="🏢 Strategy B: Consolidated Sourcing Options", font=ctk.CTkFont(weight="bold", size=14)).pack(pady=8, padx=15, anchor="w")
            
            for s_res in consolidation_results:
                s_name = s_res["supplier"]
                s_total = s_res["total_cost"]
                
                mark = ""
                if best_consolidated and s_name == best_consolidated["supplier"]:
                    mark = " ★ Best Consolidated"
                
                ctk.CTkLabel(con_fr, text=f"• {s_name}{mark}", font=ctk.CTkFont(weight="bold", size=12)).pack(pady=2, padx=15, anchor="w")
                
                for item in s_res["items"]:
                    row_lbl = f"  - {item['product']}: {item['qty']:,} pcs @ ${item['fob']:.4f}/pc — Cost: ${item['cost']:,.2f}"
                    ctk.CTkLabel(con_fr, text=row_lbl, font=ctk.CTkFont(size=11), text_color="grey").pack(pady=1, padx=25, anchor="w")
                    
                summary_con = f"  Items: ${s_res['fob_total']:,.2f} | Freight: ${s_res['freight']:,.2f} | Duties: ${s_res['duty']:,.2f} | Handling: ${s_res['local']:,.2f}"
                ctk.CTkLabel(con_fr, text=summary_con, font=ctk.CTkFont(size=11), text_color="grey").pack(pady=1, padx=25, anchor="w")
                total_color = self.THEME["success"] if best_consolidated and s_name == best_consolidated["supplier"] else self.THEME["primary"]
                ctk.CTkLabel(con_fr, text=f"  TOTAL CONSOLIDATED COST: ${s_total:,.2f}", font=ctk.CTkFont(size=12, weight="bold"), text_color=total_color).pack(pady=(2, 6), padx=25, anchor="w")
        else:
            self.make_empty_state(self.opt_results_scroll, "No full consolidation option", "No single supplier in the database quotes every selected product.")

    def setup_rfq_generator_tab(self):
        tab_rfq = self.rfqs_tabview.tab("📝 RFQ Generator")
        tab_rfq.grid_columnconfigure(0, weight=1)
        tab_rfq.grid_columnconfigure(1, weight=1)
        tab_rfq.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_rfq, text="AI Request For Quotation (RFQ) Generator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: RFQ Fields ---
        left_frame = ctk.CTkFrame(tab_rfq)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="📝 RFQ Sourcing Requirements", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=10, sticky="w")

        # Product Dropdown & Select
        ctk.CTkLabel(left_frame, text="Select Product category:").grid(row=1, column=0, padx=15, pady=2, sticky="w")
        self.rfq_product_cb = ctk.CTkComboBox(left_frame, values=["Custom"], command=self.on_rfq_product_changed)
        self.rfq_product_cb.grid(row=1, column=1, padx=15, pady=2, sticky="ew")

        # Custom Product Input
        ctk.CTkLabel(left_frame, text="Or Custom Product Name:").grid(row=2, column=0, padx=15, pady=2, sticky="w")
        self.rfq_name_entry = ctk.CTkEntry(left_frame)
        self.rfq_name_entry.grid(row=2, column=1, padx=15, pady=2, sticky="ew")
        self.rfq_name_entry.bind("<KeyRelease>", lambda e: self.update_rfq_readiness_panel())

        # Industry Template Dropdown
        ctk.CTkLabel(left_frame, text="Industry Template:").grid(row=3, column=0, padx=15, pady=2, sticky="w")
        self.rfq_template_cb = ctk.CTkComboBox(left_frame, values=["Custom / None", "Medical PPE", "Consumer Electronics", "Apparel & Fabrics", "Industrial Packaging"], command=self.on_rfq_template_changed)
        self.rfq_template_cb.grid(row=3, column=1, padx=15, pady=2, sticky="ew")
        self.rfq_template_cb.set("Custom / None")

        # Target Qty
        ctk.CTkLabel(left_frame, text="Target Quantity (pcs):").grid(row=4, column=0, padx=15, pady=2, sticky="w")
        self.rfq_qty_entry = ctk.CTkEntry(left_frame)
        self.rfq_qty_entry.grid(row=4, column=1, padx=15, pady=2, sticky="ew")
        self.rfq_qty_entry.insert(0, "100000")
        self.rfq_qty_entry.bind("<KeyRelease>", lambda e: self.update_rfq_readiness_panel())

        # Target Price Term
        ctk.CTkLabel(left_frame, text="Price Terms (FOB/EXW):").grid(row=5, column=0, padx=15, pady=2, sticky="w")
        self.rfq_term_cb = ctk.CTkComboBox(left_frame, values=["FOB Wuhan", "FOB Shanghai", "FOB Ningbo", "EXW", "CIF", "DDP"], command=lambda choice: self.update_rfq_readiness_panel())
        self.rfq_term_cb.grid(row=5, column=1, padx=15, pady=2, sticky="ew")
        self.rfq_term_cb.set("FOB Shanghai")

        # Lead Time
        ctk.CTkLabel(left_frame, text="Target Lead Time:").grid(row=6, column=0, padx=15, pady=2, sticky="w")
        self.rfq_lead_entry = ctk.CTkEntry(left_frame)
        self.rfq_lead_entry.grid(row=6, column=1, padx=15, pady=2, sticky="ew")
        self.rfq_lead_entry.insert(0, "30 days")
        self.rfq_lead_entry.bind("<KeyRelease>", lambda e: self.update_rfq_readiness_panel())

        # Payment Term
        ctk.CTkLabel(left_frame, text="Payment Terms:").grid(row=7, column=0, padx=15, pady=2, sticky="w")
        self.rfq_payment_entry = ctk.CTkEntry(left_frame)
        self.rfq_payment_entry.grid(row=7, column=1, padx=15, pady=2, sticky="ew")
        self.rfq_payment_entry.insert(0, "30% Deposit, 70% Balance against B/L")
        self.rfq_payment_entry.bind("<KeyRelease>", lambda e: self.update_rfq_readiness_panel())

        # Specs Label
        ctk.CTkLabel(left_frame, text="Product Specifications:").grid(row=8, column=0, padx=15, pady=5, sticky="w")
        
        btn_spec_helper = ctk.CTkButton(left_frame, text="✍ Refine Specs with AI", fg_color="#6e4513", hover_color="#52320b", font=ctk.CTkFont(size=11), command=self.refine_rfq_specs_with_ai)
        btn_spec_helper.grid(row=8, column=1, padx=15, pady=5, sticky="e")

        # Text specs field
        self.rfq_specs_text = ctk.CTkTextbox(left_frame, height=130)
        self.rfq_specs_text.grid(row=9, column=0, columnspan=2, padx=15, pady=(2, 15), sticky="nsew")
        self.rfq_specs_text.bind("<KeyRelease>", lambda e: self.update_rfq_readiness_panel())

        # --- RIGHT PANEL: Actions & Preview ---
        right_frame = ctk.CTkFrame(tab_rfq)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📋 RFQ PDF Action Controls", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        self.rfq_preview_box = ctk.CTkTextbox(right_frame, height=280, font=("Consolas", 10))
        self.rfq_preview_box.grid(row=1, column=0, padx=15, pady=5, sticky="nsew")
        self.rfq_preview_box.insert("1.0", "Fill out the fields on the left and click 'Generate PDF RFQ' to compile your sourcing document.")
        self.rfq_preview_box.configure(state="disabled")

        self.rfq_readiness_box = ctk.CTkTextbox(right_frame, height=110, wrap="word")
        self.rfq_readiness_box.grid(row=2, column=0, padx=15, pady=(0, 8), sticky="ew")
        self.style_text_output(self.rfq_readiness_box)

        action_fr = ctk.CTkFrame(right_frame, fg_color="transparent")
        action_fr.grid(row=3, column=0, padx=15, pady=15, sticky="ew")

        self.btn_gen_rfq_pdf = ctk.CTkButton(action_fr, text="📝 Generate PDF RFQ", fg_color="#1f538d", hover_color="#153e6b", command=self.generate_rfq_pdf)
        self.btn_gen_rfq_pdf.pack(side="right", padx=5)

        self.btn_broadcast_rfq = ctk.CTkButton(action_fr, text="📡 Broadcast RFQ", fg_color="#6e4513", hover_color="#52320b", command=self.open_rfq_broadcast_popup)
        self.btn_broadcast_rfq.pack(side="left", padx=5)

        self.btn_save_rfq_draft = self.make_button(action_fr, "Save Draft", command=lambda: self.save_rfq_record("Draft"), variant="secondary", width=95)
        self.btn_save_rfq_draft.pack(side="left", padx=5)
        self.update_rfq_readiness_panel()

    def setup_rfq_register_tab(self):
        tab = self.rfqs_tabview.tab("RFQ Register")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_columnconfigure(1, weight=0, minsize=360)
        tab.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(tab, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 8))
        header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header, text="RFQ Register", font=ctk.CTkFont(size=20, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(header, text="Saved sourcing requests with status, target quantity, terms, and generated document path.", font=ctk.CTkFont(size=12), text_color=self.THEME["muted"]).grid(row=1, column=0, sticky="w", pady=(3, 0))
        self.make_button(header, "Refresh", command=self.load_rfq_register, variant="secondary", width=90).grid(row=0, column=1, rowspan=2, padx=5)
        self.make_button(header, "Mark Sent", command=lambda: self.update_selected_rfq_status("Sent"), variant="primary", width=90).grid(row=0, column=2, rowspan=2, padx=5)
        self.make_button(header, "Close", command=lambda: self.update_selected_rfq_status("Closed"), variant="success", width=80).grid(row=0, column=3, rowspan=2, padx=5)
        self.make_button(header, "Cancel", command=lambda: self.update_selected_rfq_status("Cancelled"), variant="danger", width=80).grid(row=0, column=4, rowspan=2, padx=5)

        frame = ctk.CTkFrame(tab, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        frame.grid(row=1, column=0, sticky="nsew", padx=(16, 8), pady=(0, 16))
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        cols = ("id", "rfq", "product", "qty", "terms", "deadline", "status", "pdf")
        self.rfq_register_tree = ttk.Treeview(frame, columns=cols, show="headings", style="Treeview")
        for col, label, width in [
            ("id", "ID", 50),
            ("rfq", "RFQ No.", 150),
            ("product", "Product", 160),
            ("qty", "Qty", 90),
            ("terms", "Terms", 90),
            ("deadline", "Lead Time", 90),
            ("status", "Status", 90),
            ("pdf", "PDF Path", 260),
        ]:
            self.rfq_register_tree.heading(col, text=label)
            self.rfq_register_tree.column(col, width=width, anchor="w")
        self.rfq_register_tree.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        self.rfq_register_tree.bind("<<TreeviewSelect>>", lambda event: self.show_rfq_register_detail())

        detail = ctk.CTkFrame(tab, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        detail.grid(row=1, column=1, sticky="nsew", padx=(8, 16), pady=(0, 16))
        detail.grid_columnconfigure(0, weight=1)
        detail.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(detail, text="RFQ Detail", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 8))
        self.rfq_detail_box = ctk.CTkTextbox(detail, wrap="word", fg_color=self.THEME["surface_alt"], text_color=self.THEME["text"], border_width=0)
        self.rfq_detail_box.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 10))
        self.rfq_detail_box.insert("1.0", "Select an RFQ row to view details, document path, and workflow audit trail.")
        self.rfq_detail_box.configure(state="disabled")

        rfq_actions = ctk.CTkFrame(detail, fg_color="transparent")
        rfq_actions.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 14))
        rfq_actions.grid_columnconfigure(0, weight=1)
        rfq_actions.grid_columnconfigure(1, weight=1)
        self.make_button(rfq_actions, "Open PDF", command=lambda: self.open_selected_register_document("RFQ"), variant="primary").grid(row=0, column=0, sticky="ew", padx=(0, 4), pady=4)
        self.make_button(rfq_actions, "Open Folder", command=lambda: self.open_selected_register_folder("RFQ"), variant="secondary").grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=4)
        self.make_button(rfq_actions, "Copy Path", command=lambda: self.copy_selected_register_path("RFQ"), variant="secondary").grid(row=1, column=0, sticky="ew", padx=(0, 4), pady=4)
        self.make_button(rfq_actions, "Load Generator", command=self.load_selected_rfq_into_generator, variant="success").grid(row=1, column=1, sticky="ew", padx=(4, 0), pady=4)
        self.load_rfq_register()

    def load_rfq_register(self):
        if not hasattr(self, "rfq_register_tree"):
            return
        self.rfq_register_tree.delete(*self.rfq_register_tree.get_children())
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT id, rfq_number, product_name, target_quantity, price_terms, lead_time, status, COALESCE(pdf_path, '')
            FROM rfq_register
            ORDER BY updated_at DESC, id DESC
        """)
        for row in c.fetchall():
            self.rfq_register_tree.insert("", tk.END, values=row)
        conn.close()
        if hasattr(self, "rfq_detail_box"):
            self.set_detail_text(self.rfq_detail_box, "Select an RFQ row to view details, document path, and workflow audit trail.")

    def set_detail_text(self, textbox, text):
        textbox.configure(state="normal")
        textbox.delete("1.0", tk.END)
        textbox.insert("1.0", text)
        textbox.configure(state="disabled")

    def get_selected_register_id_and_path(self, register_type):
        tree = self.rfq_register_tree if register_type == "RFQ" else self.po_register_tree
        sel = tree.selection()
        if not sel:
            return None, ""
        vals = tree.item(sel[0], "values")
        if not vals:
            return None, ""
        try:
            record_id = int(vals[0])
        except Exception:
            record_id = None
        return record_id, vals[-1] if vals[-1] else ""

    def get_workflow_audit_lines(self, workflow_type, workflow_id):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT created_at, action, status, COALESCE(note, '')
            FROM workflow_audit_log
            WHERE workflow_type=? AND workflow_id=?
            ORDER BY datetime(created_at) DESC, id DESC
            LIMIT 12
        """, (workflow_type, workflow_id))
        rows = c.fetchall()
        conn.close()
        if not rows:
            return ["No audit events recorded yet."]
        return [f"{created_at} | {action} -> {status}\n{note}" for created_at, action, status, note in rows]

    def show_rfq_register_detail(self):
        rfq_id, _ = self.get_selected_register_id_and_path("RFQ")
        if not rfq_id:
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT rfq_number, product_name, target_quantity, price_terms, lead_time,
                   payment_terms, COALESCE(selected_suppliers, ''), status,
                   COALESCE(pdf_path, ''), COALESCE(specs, ''), created_at, updated_at
            FROM rfq_register
            WHERE id=?
        """, (rfq_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            return
        rfq_number, product, qty, terms, lead, payment, suppliers, status, pdf_path, specs, created, updated = row
        audit = "\n\n".join(self.get_workflow_audit_lines("RFQ", rfq_id))
        text = (
            f"RFQ: {rfq_number}\n"
            f"Status: {status}\n"
            f"Product: {product}\n"
            f"Target Quantity: {qty}\n"
            f"Price Terms: {terms}\n"
            f"Lead Time: {lead}\n"
            f"Payment Terms: {payment}\n"
            f"Selected Suppliers: {suppliers or 'Not selected'}\n"
            f"PDF Path: {pdf_path or 'Not generated'}\n"
            f"Created: {created}\n"
            f"Updated: {updated}\n\n"
            f"Specifications:\n{specs or 'No specifications saved.'}\n\n"
            f"Workflow Audit:\n{audit}"
        )
        self.set_detail_text(self.rfq_detail_box, text)

    def show_po_register_detail(self):
        po_id, _ = self.get_selected_register_id_and_path("PO")
        if not po_id:
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT po_number, supplier_name, product_name, quantity, unit_cost, total_value,
                   payment_terms, delivery_address, status, COALESCE(pdf_path, ''),
                   quote_id, created_at, updated_at
            FROM po_register
            WHERE id=?
        """, (po_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            return
        po_number, supplier, product, qty, unit_cost, total_value, payment, address, status, pdf_path, quote_id, created, updated = row
        audit = "\n\n".join(self.get_workflow_audit_lines("PO", po_id))
        unit_text = f"${float(unit_cost):.5f}" if unit_cost is not None else "N/A"
        total_text = f"${float(total_value):,.2f}" if total_value is not None else "N/A"
        text = (
            f"PO: {po_number}\n"
            f"Status: {status}\n"
            f"Supplier: {supplier}\n"
            f"Product: {product}\n"
            f"Quantity: {qty:,} pcs\n"
            f"Unit Cost: {unit_text}\n"
            f"Total Value: {total_text}\n"
            f"Payment Terms: {payment}\n"
            f"Delivery Address: {address}\n"
            f"Related Quote ID: {quote_id or 'N/A'}\n"
            f"PDF Path: {pdf_path or 'Not generated'}\n"
            f"Created: {created}\n"
            f"Updated: {updated}\n\n"
            f"Workflow Audit:\n{audit}"
        )
        self.set_detail_text(self.po_detail_box, text)

    def open_selected_register_document(self, register_type):
        _, path = self.get_selected_register_id_and_path(register_type)
        if not path:
            messagebox.showwarning("No Document", f"No {register_type} PDF path is saved for the selected record.")
            return
        if not os.path.exists(path):
            messagebox.showerror("File Missing", f"The saved document path does not exist:\n{path}")
            return
        os.startfile(path)

    def open_selected_register_folder(self, register_type):
        _, path = self.get_selected_register_id_and_path(register_type)
        if not path:
            messagebox.showwarning("No Document", f"No {register_type} PDF path is saved for the selected record.")
            return
        folder = os.path.dirname(path)
        if not folder or not os.path.isdir(folder):
            messagebox.showerror("Folder Missing", f"The saved document folder does not exist:\n{folder}")
            return
        os.startfile(folder)

    def copy_selected_register_path(self, register_type):
        _, path = self.get_selected_register_id_and_path(register_type)
        if not path:
            messagebox.showwarning("No Document", f"No {register_type} PDF path is saved for the selected record.")
            return
        self.copy_to_clipboard(path)

    def set_tab_by_contains(self, tabview, text):
        try:
            names = list(getattr(tabview, "_name_list", []))
            for name in names:
                if text.lower() in str(name).lower():
                    tabview.set(name)
                    return
        except Exception:
            pass

    def load_selected_rfq_into_generator(self):
        rfq_id, _ = self.get_selected_register_id_and_path("RFQ")
        if not rfq_id:
            messagebox.showwarning("Select RFQ", "Please select an RFQ record first.")
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT product_name, target_quantity, price_terms, lead_time, payment_terms, specs
            FROM rfq_register
            WHERE id=?
        """, (rfq_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            return
        product, qty, terms, lead, payment, specs = row
        self.show_page("RFQs Outreach")
        self.set_tab_by_contains(self.rfqs_tabview, "RFQ Generator")
        self.rfq_product_cb.set("Custom")
        self.rfq_name_entry.delete(0, tk.END)
        self.rfq_name_entry.insert(0, product or "")
        self.rfq_qty_entry.delete(0, tk.END)
        self.rfq_qty_entry.insert(0, qty or "")
        self.rfq_term_cb.set(terms or "FOB Shanghai")
        self.rfq_lead_entry.delete(0, tk.END)
        self.rfq_lead_entry.insert(0, lead or "")
        self.rfq_payment_entry.delete(0, tk.END)
        self.rfq_payment_entry.insert(0, payment or "")
        self.rfq_specs_text.delete("1.0", tk.END)
        self.rfq_specs_text.insert("1.0", specs or "")
        self.update_rfq_readiness_panel()

    def load_selected_po_into_generator(self):
        po_id, _ = self.get_selected_register_id_and_path("PO")
        if not po_id:
            messagebox.showwarning("Select PO", "Please select a PO record first.")
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT po_number, supplier_name, product_name, quantity, payment_terms,
                   delivery_address, quote_id, unit_cost
            FROM po_register
            WHERE id=?
        """, (po_id,))
        row = c.fetchone()
        conn.close()
        if not row:
            return
        po_number, supplier, product, qty, payment, address, quote_id, unit_cost = row
        self.show_page("Logistics Costing")
        self.set_tab_by_contains(self.logistics_tabview, "PO Generator")
        self.po_supplier_cb.set(supplier or "Select Supplier")
        self.po_product_cb.set(product or "Select Product")
        self.po_number_entry.delete(0, tk.END)
        self.po_number_entry.insert(0, po_number or "")
        self.po_qty_entry.delete(0, tk.END)
        self.po_qty_entry.insert(0, str(qty or ""))
        self.po_payment_entry.delete(0, tk.END)
        self.po_payment_entry.insert(0, payment or "")
        self.po_address_entry.delete(0, tk.END)
        self.po_address_entry.insert(0, address or "")
        self.po_active_price = float(unit_cost or 0)
        self.po_active_quote = next((r for r in self.extracted_data if r.get("id") == quote_id), {"id": quote_id, "review_status": "Approved"})
        self.update_po_preview()

    def update_selected_rfq_status(self, status):
        sel = self.rfq_register_tree.selection()
        if not sel:
            messagebox.showwarning("Select RFQ", "Please select an RFQ record first.")
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        for item in sel:
            rfq_id = int(self.rfq_register_tree.item(item, "values")[0])
            c.execute("UPDATE rfq_register SET status=?, updated_at=datetime('now') WHERE id=?", (status, rfq_id))
            c.execute("""
                INSERT INTO workflow_audit_log (workflow_type, workflow_id, action, status, note)
                VALUES ('RFQ', ?, 'Status Changed', ?, ?)
            """, (rfq_id, status, f"RFQ marked {status}."))
        conn.commit()
        conn.close()
        self.load_rfq_register()
        if hasattr(self, "dashboard_cards"):
            self.update_dashboard_page()

    def on_rfq_product_changed(self, choice):
        if choice == "Custom":
            self.rfq_name_entry.delete(0, tk.END)
            self.rfq_specs_text.delete("1.0", tk.END)
        else:
            self.rfq_name_entry.delete(0, tk.END)
            self.rfq_name_entry.insert(0, choice)
            
            best_specs = ""
            for r in self.extracted_data:
                if (r.get("product") or "").strip().lower() == choice.lower():
                    spec = r.get("spec") or ""
                    color = r.get("color") or ""
                    elastic = r.get("elastic") or ""
                    
                    details = []
                    if spec and spec != "N/A": details.append(f"Specifications: {spec}")
                    if color and color != "N/A": details.append(f"Color: {color}")
                    if elastic and elastic != "N/A": details.append(f"Elastic Style: {elastic}")
                    
                    if details:
                        best_specs = "\n".join(details)
                        break
            
            self.rfq_specs_text.delete("1.0", tk.END)
            if best_specs:
                self.rfq_specs_text.insert("1.0", best_specs)
            else:
                self.rfq_specs_text.insert("1.0", f"High-quality {choice} matching standard industry specifications.")
        self.update_rfq_readiness_panel()

    def update_rfq_generator_tab(self):
        products = set()
        for r in self.extracted_data:
            prod = (r.get("product") or "").strip().title()
            if prod:
                products.add(prod)
        sorted_prods = ["Custom"] + sorted(list(products))
        
        current = self.rfq_product_cb.get()
        self.rfq_product_cb.configure(values=sorted_prods)
        if current in sorted_prods:
            self.rfq_product_cb.set(current)
        else:
            self.rfq_product_cb.set("Custom")
        self.update_rfq_readiness_panel()

    def refine_rfq_specs_with_ai(self):
        prod_name = self.rfq_name_entry.get().strip()
        current_specs = self.rfq_specs_text.get("1.0", tk.END).strip()
        
        if not prod_name:
            messagebox.showwarning("Warning", "Please enter a product name first!")
            return
            
        self.rfq_specs_text.delete("1.0", tk.END)
        self.rfq_specs_text.insert("1.0", "AI is writing premium technical specs...")
        self.update()
        
        def run_ai_specs():
            prompt = f"""
            You are an expert global sourcing manager. Write a detailed, professional technical specification list for a Request for Quotation (RFQ).
            Product: {prod_name}
            Current specifications input: {current_specs}
            
            Please list out:
            1. Raw Materials & Grade standards (e.g. non-woven, PE film, weight in gsm if applicable).
            2. Structural details & Dimensions (dimensions, elastic types, seam style).
            3. Quality control standards (FDA/CE conformity if standard).
            4. Packing instructions (bag/carton counts).
            
            Keep the specifications clear, professional, and formatted in bullet points.
            Output ONLY the specification bullet points. Do not write introductory words.
            """
            try:
                refined = self.generate_with_fallback([], prompt, json_response=False)
                self.after(0, lambda text=refined: update_textbox(text))
            except Exception as e:
                self.after(0, lambda err=e: update_textbox(f"Error refining specs: {err}"))
                
        def update_textbox(text):
            self.rfq_specs_text.delete("1.0", tk.END)
            self.rfq_specs_text.insert("1.0", text)
            self.update_rfq_readiness_panel()
            
        threading.Thread(target=run_ai_specs, daemon=True).start()

    def generate_rfq_pdf(self):
        issues, warnings = self.validate_rfq_ready()
        self.update_rfq_readiness_panel()
        if issues:
            note = "; ".join(issues)
            self.log_workflow_blocked_attempt("RFQ", "Generate PDF Blocked", note)
            messagebox.showerror(
                "RFQ Not Ready",
                "Fix these issues before generating the RFQ:\n\n- " + "\n- ".join(issues)
            )
            return
        if warnings and not messagebox.askyesno(
            "RFQ Readiness Warnings",
            "These items should be improved for an enterprise RFQ:\n\n- "
            + "\n- ".join(warnings)
            + "\n\nContinue anyway?"
        ):
            self.log_workflow_blocked_attempt("RFQ", "Generate PDF Blocked", "; ".join(warnings))
            return

        prod_name = self.rfq_name_entry.get().strip()
        qty = self.rfq_qty_entry.get().strip()
        terms = self.rfq_term_cb.get().strip()
        lead_time = self.rfq_lead_entry.get().strip()
        payments = self.rfq_payment_entry.get().strip()
        specs = self.rfq_specs_text.get("1.0", tk.END).strip()
        
        if not prod_name:
            messagebox.showwarning("Warning", "Please specify a product name before generating the RFQ PDF!")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Documents", "*.pdf")],
            initialfile=f"RFQ_Request_{prod_name.replace(' ', '_')}.pdf",
            title="Save RFQ PDF Document"
        )
        
        if not file_path:
            return
            
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            import datetime
            
            doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            
            styles = getSampleStyleSheet()
            
            primary_color = colors.HexColor("#1f538d")
            
            title_style = ParagraphStyle(
                'RFQTitle',
                parent=styles['Heading1'],
                fontSize=22,
                textColor=primary_color,
                spaceAfter=15
            )
            
            body_style = ParagraphStyle(
                'RFQBody',
                parent=styles['Normal'],
                fontSize=10,
                leading=14,
                textColor=colors.HexColor("#333333")
            )
            
            sub_title_style = ParagraphStyle(
                'RFQSubTitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=primary_color,
                spaceBefore=12,
                spaceAfter=8
            )
            
            story.append(Paragraph("REQUEST FOR QUOTATION (RFQ)", title_style))
            story.append(Paragraph(f"<b>Document Ref:</b> RFQ-{datetime.date.today().strftime('%Y%m%d')}-{prod_name[:4].upper()}", body_style))
            story.append(Paragraph(f"<b>Date Generated:</b> {datetime.date.today().strftime('%B %d, %Y')}", body_style))
            story.append(Spacer(1, 15))
            
            terms_data = [
                [Paragraph("<b>Sourcing Category</b>", body_style), Paragraph(prod_name, body_style)],
                [Paragraph("<b>Target Order Volume</b>", body_style), Paragraph(f"{qty} pieces", body_style)],
                [Paragraph("<b>Required Price Terms</b>", body_style), Paragraph(terms, body_style)],
                [Paragraph("<b>Target Delivery Time</b>", body_style), Paragraph(lead_time, body_style)],
                [Paragraph("<b>Required Payment Terms</b>", body_style), Paragraph(payments, body_style)]
            ]
            
            terms_table = Table(terms_data, colWidths=[180, 350])
            terms_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,-1), colors.HexColor("#f2f5f9")),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor("#333333")),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dcdcdc")),
                ('BOX', (0,0), (-1,-1), 1, primary_color),
            ]))
            story.append(terms_table)
            story.append(Spacer(1, 20))
            
            story.append(Paragraph("Technical Product Specifications", sub_title_style))
            specs_formatted = specs.replace("\n", "<br/>")
            story.append(Paragraph(specs_formatted, body_style))
            story.append(Spacer(1, 20))
            
            story.append(Paragraph("Submission Instructions", sub_title_style))
            instructions = (
                "Please submit your formal quotation matching or improving upon the above terms. "
                "Quote submissions must include: unit price, carton packing configurations, carton dimensions (MEAS/CBM), "
                "production lead times, and country of origin certificates."
            )
            story.append(Paragraph(instructions, body_style))
            
            doc.build(story)
            
            self.rfq_preview_box.configure(state="normal")
            self.rfq_preview_box.delete("1.0", tk.END)
            self.rfq_preview_box.insert("1.0", f"✅ RFQ PDF generated successfully!\nSaved to: {file_path}\n\n=== Sourcing Target Summary ===\nProduct: {prod_name}\nQuantity: {qty}\nPrice Terms: {terms}\nLead Time: {lead_time}\nPayment Terms: {payments}\n\n=== Specifications Draft ===\n{specs}")
            self.rfq_preview_box.configure(state="disabled")
            self.save_rfq_record("Draft", file_path)
            
            messagebox.showinfo("Success", f"RFQ PDF saved successfully at:\n{file_path}!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate RFQ PDF: {e}")

    def setup_profit_simulator_tab(self):
        tab_prof = self.logistics_tabview.tab("💰 Profit Simulator")
        tab_prof.grid_columnconfigure(0, weight=1)
        tab_prof.grid_columnconfigure(1, weight=3)
        tab_prof.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_prof, text="Profit Margin, ROI & Retail Pricing Simulator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Revenue & Operating Cost Inputs ---
        left_frame = ctk.CTkFrame(tab_prof)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left_frame, text="⚙ Price & Cost Inputs", font=ctk.CTkFont(size=15, weight="bold")).pack(pady=10, padx=15, anchor="w")

        # Retail Price per piece
        ctk.CTkLabel(left_frame, text="Target Retail Price ($/pc):", font=ctk.CTkFont(size=12)).pack(pady=(5, 0), padx=15, anchor="w")
        self.prof_retail_entry = ctk.CTkEntry(left_frame)
        self.prof_retail_entry.pack(pady=2, padx=15, fill="x")
        self.prof_retail_entry.insert(0, "1.50")

        # Fulfillment/Warehouse Cost per piece
        ctk.CTkLabel(left_frame, text="Warehouse/Fulfillment Fee ($/pc):", font=ctk.CTkFont(size=12)).pack(pady=(10, 0), padx=15, anchor="w")
        self.prof_fba_entry = ctk.CTkEntry(left_frame)
        self.prof_fba_entry.pack(pady=2, padx=15, fill="x")
        self.prof_fba_entry.insert(0, "0.30")

        # Marketing Cost per piece
        ctk.CTkLabel(left_frame, text="Marketing/CAC Cost ($/pc):", font=ctk.CTkFont(size=12)).pack(pady=(10, 0), padx=15, anchor="w")
        self.prof_mkt_entry = ctk.CTkEntry(left_frame)
        self.prof_mkt_entry.pack(pady=2, padx=15, fill="x")
        self.prof_mkt_entry.insert(0, "0.15")

        # Other OPEX / Overhead per piece
        ctk.CTkLabel(left_frame, text="Other Overhead/OPEX ($/pc):", font=ctk.CTkFont(size=12)).pack(pady=(10, 0), padx=15, anchor="w")
        self.prof_opex_entry = ctk.CTkEntry(left_frame)
        self.prof_opex_entry.pack(pady=2, padx=15, fill="x")
        self.prof_opex_entry.insert(0, "0.05")

        # Calculate Button
        self.btn_calc_prof = ctk.CTkButton(left_frame, text="💰 Run Profit Analysis", fg_color="#1f538d", hover_color="#153e6b", command=self.update_profit_simulator_tab)
        self.btn_calc_prof.pack(pady=25, padx=15, fill="x")

        # --- RIGHT PANEL: Profitability Comparison Treeview ---
        right_frame = ctk.CTkFrame(tab_prof)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📊 Profitability Comparison Matrix", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        cols = ("supplier", "product", "landed", "operating", "total_cost", "net_profit", "gross_margin", "net_margin", "roi")
        
        scroll_y = ttk.Scrollbar(right_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(right_frame, orient="horizontal")
        
        self.profit_tree = ttk.Treeview(
            right_frame,
            columns=cols,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )
        
        scroll_y.config(command=self.profit_tree.yview)
        scroll_x.config(command=self.profit_tree.xview)
        
        scroll_y.grid(row=1, column=1, sticky="ns")
        scroll_x.grid(row=2, column=0, sticky="ew")
        self.profit_tree.grid(row=1, column=0, sticky="nsew", padx=(15, 0))

        # Headings
        self.profit_tree.heading("supplier", text="Supplier")
        self.profit_tree.heading("product", text="Product")
        self.profit_tree.heading("landed", text="Landed ($)")
        self.profit_tree.heading("operating", text="Opex ($)")
        self.profit_tree.heading("total_cost", text="Total Cost ($)")
        self.profit_tree.heading("net_profit", text="Net Profit ($)")
        self.profit_tree.heading("gross_margin", text="Gross Margin")
        self.profit_tree.heading("net_margin", text="Net Margin")
        self.profit_tree.heading("roi", text="ROI (%)")

        # Widths
        self.profit_tree.column("supplier", width=130, anchor="w")
        self.profit_tree.column("product", width=110, anchor="w")
        self.profit_tree.column("landed", width=75, anchor="center")
        self.profit_tree.column("operating", width=75, anchor="center")
        self.profit_tree.column("total_cost", width=85, anchor="center")
        self.profit_tree.column("net_profit", width=95, anchor="center")
        self.profit_tree.column("gross_margin", width=90, anchor="center")
        self.profit_tree.column("net_margin", width=90, anchor="center")
        self.profit_tree.column("roi", width=80, anchor="center")

        self.profit_tree.tag_configure("winner", background="#DCFCE7", foreground="#14532D")

    def update_profit_simulator_tab(self):
        for item in self.profit_tree.get_children():
            self.profit_tree.delete(item)

        if not self.extracted_data:
            return

        try:
            retail_price = float(self.prof_retail_entry.get().replace(",", "").strip())
        except Exception:
            retail_price = 1.50
        try:
            warehouse_fee = float(self.prof_fba_entry.get().replace(",", "").strip())
        except Exception:
            warehouse_fee = 0.30
        try:
            marketing_fee = float(self.prof_mkt_entry.get().replace(",", "").strip())
        except Exception:
            marketing_fee = 0.15
        try:
            opex_fee = float(self.prof_opex_entry.get().replace(",", "").strip())
        except Exception:
            opex_fee = 0.05

        order_qty = 100000
        freight_rate = 120.0
        duty_percent = 6.5
        local_flat = 350.0
        freight_mode = "LCL (per CBM)"
        
        if hasattr(self, 'sim_qty_entry'):
            try:
                order_qty = int(self.sim_qty_entry.get().replace(",", "").strip())
            except Exception: pass
        if hasattr(self, 'sim_rate_entry'):
            try:
                freight_rate = float(self.sim_rate_entry.get().replace(",", "").strip())
            except Exception: pass
        if hasattr(self, 'sim_duty_slider'):
            try:
                duty_percent = self.sim_duty_slider.get()
            except Exception: pass
        if hasattr(self, 'sim_local_entry'):
            try:
                local_flat = float(self.sim_local_entry.get().replace(",", "").strip())
            except Exception: pass
        if hasattr(self, 'sim_freight_mode'):
            try:
                freight_mode = self.sim_freight_mode.get()
            except Exception: pass

        rows = []
        for r in self.extracted_data:
            fob = r.get("price")
            if fob is None or fob == "N/A":
                continue
                
            supplier = self.clean_supplier_name(r.get("supplier"))
            product = r.get("product") or "Product"
            packing = r.get("packing") or "N/A"
            
            pcs_per_ctn, unit_cbm = self.parse_packing_metrics(packing)
            
            import math
            total_ctns = math.ceil(order_qty / pcs_per_ctn)
            total_cbm = total_ctns * unit_cbm
            total_fob = order_qty * fob
            
            if freight_mode == "LCL (per CBM)":
                total_freight = total_cbm * freight_rate
            elif freight_mode == "FCL 20GP":
                containers = math.ceil(total_cbm / 28.0)
                total_freight = containers * freight_rate
            else:
                containers = math.ceil(total_cbm / 68.0)
                total_freight = containers * freight_rate
                
            total_duty = total_fob * (duty_percent / 100.0)
            total_local = local_flat
            
            total_landed = total_fob + total_freight + total_duty + total_local
            landed_pc = total_landed / order_qty
            
            opex_total_pc = warehouse_fee + marketing_fee + opex_fee
            unit_total_cost = landed_pc + opex_total_pc
            net_profit = retail_price - unit_total_cost
            
            gross_margin = ((retail_price - landed_pc) / retail_price) * 100.0 if retail_price > 0 else 0.0
            net_margin = (net_profit / retail_price) * 100.0 if retail_price > 0 else 0.0
            roi = (net_profit / landed_pc) * 100.0 if landed_pc > 0 else 0.0
            
            rows.append({
                "supplier": supplier,
                "product": product,
                "landed": landed_pc,
                "opex": opex_total_pc,
                "total_cost": unit_total_cost,
                "net_profit": net_profit,
                "gross_margin": gross_margin,
                "net_margin": net_margin,
                "roi": roi
            })

        if not rows:
            return

        rows.sort(key=lambda x: x["net_profit"], reverse=True)
        best_profit = rows[0]["net_profit"]

        for r in rows:
            tag = ""
            if abs(r["net_profit"] - best_profit) < 1e-7:
                tag = "winner"
                
            self.profit_tree.insert(
                "",
                tk.END,
                values=(
                    r["supplier"],
                    r["product"],
                    f"${r['landed']:.4f}",
                    f"${r['opex']:.2f}",
                    f"${r['total_cost']:.4f}",
                    f"${r['net_profit']:.4f}",
                    f"{r['gross_margin']:.1f}%",
                    f"{r['net_margin']:.1f}%",
                    f"{r['roi']:.1f}%"
                ),
                tags=(tag,)
            )

    def setup_factory_qc_tab(self):
        tab_qc = self.scorecard_tabview.tab("🏢 Factory Audit & QC")
        tab_qc.grid_columnconfigure(0, weight=1)
        tab_qc.grid_columnconfigure(1, weight=1)
        tab_qc.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_qc, text="Factory Compliance Standards & QC Inspection", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Compliance Entry ---
        left_frame = ctk.CTkFrame(tab_qc)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="🏢 Compliance & Audit Log", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=10, sticky="w")

        ctk.CTkLabel(left_frame, text="Select Supplier:").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        self.qc_supplier_cb = ctk.CTkComboBox(left_frame, values=[], command=self.load_compliance_record)
        self.qc_supplier_cb.grid(row=1, column=1, padx=15, pady=5, sticky="ew")

        # Certification Checkboxes
        ctk.CTkLabel(left_frame, text="Compliance Standards:").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        
        cb_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        cb_frame.grid(row=2, column=1, padx=15, pady=5, sticky="w")
        
        self.qc_ce_var = tk.IntVar()
        self.qc_fda_var = tk.IntVar()
        self.qc_iso_var = tk.IntVar()
        self.qc_bsci_var = tk.IntVar()
        self.qc_sgs_var = tk.IntVar()
        
        ctk.CTkCheckBox(cb_frame, text="CE Certificate", variable=self.qc_ce_var, font=ctk.CTkFont(size=11)).pack(anchor="w", pady=1)
        ctk.CTkCheckBox(cb_frame, text="FDA Registration", variable=self.qc_fda_var, font=ctk.CTkFont(size=11)).pack(anchor="w", pady=1)
        ctk.CTkCheckBox(cb_frame, text="ISO 9001 (Quality)", variable=self.qc_iso_var, font=ctk.CTkFont(size=11)).pack(anchor="w", pady=1)
        ctk.CTkCheckBox(cb_frame, text="BSCI (Social)", variable=self.qc_bsci_var, font=ctk.CTkFont(size=11)).pack(anchor="w", pady=1)
        ctk.CTkCheckBox(cb_frame, text="SGS Third Party", variable=self.qc_sgs_var, font=ctk.CTkFont(size=11)).pack(anchor="w", pady=1)

        # Audit Score
        ctk.CTkLabel(left_frame, text="Audit Score (0-100):").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        self.qc_audit_entry = ctk.CTkEntry(left_frame)
        self.qc_audit_entry.grid(row=3, column=1, padx=15, pady=5, sticky="ew")
        self.qc_audit_entry.insert(0, "85")

        # Defect Rate
        ctk.CTkLabel(left_frame, text="Est Defect Rate (%):").grid(row=4, column=0, padx=15, pady=5, sticky="w")
        self.qc_defect_entry = ctk.CTkEntry(left_frame)
        self.qc_defect_entry.grid(row=4, column=1, padx=15, pady=5, sticky="ew")
        self.qc_defect_entry.insert(0, "1.2")

        self.btn_save_qc = ctk.CTkButton(left_frame, text="💾 Save Compliance Record", fg_color="#1f538d", hover_color="#153e6b", command=self.save_compliance_record)
        self.btn_save_qc.grid(row=5, column=0, columnspan=2, padx=15, pady=20, sticky="ew")

        # --- RIGHT PANEL: QC Checks & PDF ---
        right_frame = ctk.CTkFrame(tab_qc)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(right_frame, text="📋 QC Inspection Checklists", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        info_box = ctk.CTkTextbox(right_frame, height=220, font=("Consolas", 10))
        info_box.grid(row=1, column=0, padx=15, pady=5, sticky="nsew")
        info_box.insert("1.0", "Factory Quality Control checks standard procedures:\n\n1. Carton Drop Test (1.2m drop test)\n2. Barcode & Carton Labeling matches shipment details\n3. Functional Stress Check (Tensile elasticity checks)\n4. Visual Checks (Dirt, stains, seam completeness)\n5. Dimension and weight conformities")
        info_box.configure(state="disabled")

        self.btn_export_qc_pdf = ctk.CTkButton(right_frame, text="📋 Export QC Checklist PDF", fg_color="#6e4513", hover_color="#52320b", command=self.generate_qc_checklist_pdf)
        self.btn_export_qc_pdf.grid(row=2, column=0, padx=15, pady=20, sticky="ew")

        # Incident Log section
        ctk.CTkLabel(right_frame, text="🚨 Log Operational Incident / Defect", font=ctk.CTkFont(size=14, weight="bold")).grid(row=3, column=0, padx=15, pady=(15, 5), sticky="w")
        
        inc_input_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        inc_input_frame.grid(row=4, column=0, padx=15, pady=5, sticky="ew")
        inc_input_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(inc_input_frame, text="Type:").grid(row=0, column=0, padx=(0, 10), pady=2, sticky="w")
        self.qc_inc_type_cb = ctk.CTkComboBox(inc_input_frame, values=["Delay", "Defect", "Communication", "Payment"], width=120)
        self.qc_inc_type_cb.grid(row=0, column=1, padx=5, pady=2, sticky="w")
        
        ctk.CTkLabel(inc_input_frame, text="Severity:").grid(row=0, column=2, padx=10, pady=2, sticky="w")
        self.qc_inc_sev_cb = ctk.CTkComboBox(inc_input_frame, values=["Low", "Medium", "High"], width=100)
        self.qc_inc_sev_cb.grid(row=0, column=3, padx=5, pady=2, sticky="w")
        
        ctk.CTkLabel(right_frame, text="Description:").grid(row=5, column=0, padx=15, pady=(5, 2), sticky="w")
        self.qc_inc_desc_entry = ctk.CTkEntry(right_frame, placeholder_text="e.g. 5 days delay on batch #4, 2% seam tear defect rate")
        self.qc_inc_desc_entry.grid(row=6, column=0, padx=15, pady=2, sticky="ew")
        
        self.btn_log_incident = ctk.CTkButton(right_frame, text="🚨 Log Performance Incident", fg_color="#bf3b3b", hover_color="#9e2d2d", command=self.log_supplier_incident)
        self.btn_log_incident.grid(row=7, column=0, padx=15, pady=(10, 15), sticky="ew")

    def update_factory_qc_tab(self):
        suppliers = set()
        for r in self.extracted_data:
            s = self.clean_supplier_name(r.get("supplier"))
            if s and s != "Unknown":
                suppliers.add(s)
        sorted_sups = sorted(list(suppliers))
        self.qc_supplier_cb.configure(values=sorted_sups)
        if sorted_sups:
            self.qc_supplier_cb.set(sorted_sups[0])
            self.load_compliance_record(sorted_sups[0])

    def save_compliance_record(self):
        supplier = self.qc_supplier_cb.get()
        if not supplier or supplier == "Select Supplier":
            messagebox.showwarning("Warning", "Please select a supplier first!")
            return
            
        try:
            score = float(self.qc_audit_entry.get().strip())
            defect = float(self.qc_defect_entry.get().strip())
        except Exception:
            messagebox.showerror("Error", "Audit score and defect rate must be valid numbers!")
            return

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            INSERT OR REPLACE INTO supplier_compliance (supplier, has_ce, has_fda, has_iso, has_bsci, has_sgs, audit_score, defect_rate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            supplier,
            self.qc_ce_var.get(),
            self.qc_fda_var.get(),
            self.qc_iso_var.get(),
            self.qc_bsci_var.get(),
            self.qc_sgs_var.get(),
            score,
            defect
        ))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", f"Compliance record saved for {supplier} successfully!")

    def load_compliance_record(self, supplier):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT has_ce, has_fda, has_iso, has_bsci, has_sgs, audit_score, defect_rate FROM supplier_compliance WHERE supplier = ?", (supplier,))
        row = c.fetchone()
        conn.close()

        if row:
            self.qc_ce_var.set(row[0])
            self.qc_fda_var.set(row[1])
            self.qc_iso_var.set(row[2])
            self.qc_bsci_var.set(row[3])
            self.qc_sgs_var.set(row[4])
            
            self.qc_audit_entry.delete(0, tk.END)
            self.qc_audit_entry.insert(0, str(row[5]))
            
            self.qc_defect_entry.delete(0, tk.END)
            self.qc_defect_entry.insert(0, str(row[6]))
        else:
            self.qc_ce_var.set(0)
            self.qc_fda_var.set(0)
            self.qc_iso_var.set(0)
            self.qc_bsci_var.set(0)
            self.qc_sgs_var.set(0)
            
            self.qc_audit_entry.delete(0, tk.END)
            self.qc_audit_entry.insert(0, "85")
            
            self.qc_defect_entry.delete(0, tk.END)
            self.qc_defect_entry.insert(0, "1.2")

    def generate_qc_checklist_pdf(self):
        supplier = self.qc_supplier_cb.get()
        if not supplier or supplier == "Select Supplier":
            messagebox.showwarning("Warning", "Please select a supplier first!")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Documents", "*.pdf")],
            initialfile=f"QC_Inspection_Checklist_{supplier.replace(' ', '_')}.pdf",
            title="Save QC Checklist PDF"
        )
        if not file_path:
            return
            
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            import datetime
            
            doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            styles = getSampleStyleSheet()
            
            primary_color = colors.HexColor("#1f538d")
            
            title_style = ParagraphStyle(
                'QCTitle', parent=styles['Heading1'], fontSize=20, textColor=primary_color, spaceAfter=15
            )
            body_style = ParagraphStyle(
                'QCBody', parent=styles['Normal'], fontSize=9, leading=13, textColor=colors.HexColor("#333333")
            )
            sub_title_style = ParagraphStyle(
                'QCSubTitle', parent=styles['Heading2'], fontSize=12, textColor=primary_color, spaceBefore=10, spaceAfter=6
            )
            
            story.append(Paragraph("PRE-SHIPMENT QUALITY CONTROL CHECKLIST", title_style))
            story.append(Paragraph(f"<b>Supplier Factory:</b> {supplier}", body_style))
            story.append(Paragraph(f"<b>Inspection Date:</b> {datetime.date.today().strftime('%B %d, %Y')}", body_style))
            story.append(Spacer(1, 10))
            
            story.append(Paragraph("Inspection Standards & AQL Thresholds", sub_title_style))
            aql_text = (
                "<b>Inspection Level:</b> General Inspection Level II<br/>"
                "<b>Acceptable Quality Limit (AQL):</b> Critical defects: 0% | Major defects: 2.5% | Minor defects: 4.0%"
            )
            story.append(Paragraph(aql_text, body_style))
            story.append(Spacer(1, 15))
            
            qc_items = [
                ["[  ]", "Carton Drop Test", "Perform drop test from 1.2m height on 1 corner, 3 edges, and 6 faces. Assess for damage."],
                ["[  ]", "Carton Labeling Check", "Verify shipper labels, barcode format, lot numbers, SKU details, and gross weight matches packing list."],
                ["[  ]", "Visual Defects Inspection", "Inspect unit samples for dirt, stains, tears, discoloration, or manufacturing defects."],
                ["[  ]", "Dimension Measurement Check", "Measure width, length, elastic band stretch capacity, and weight against product specification sheets."],
                ["[  ]", "Functional Testing Check", "Verify performance under tension. Perform tear tests on non-woven fabrics and check seal strength."]
            ]
            
            table_data = [[
                Paragraph("<b>Status</b>", ParagraphStyle('H1', parent=body_style, textColor=colors.white)),
                Paragraph("<b>Inspection Point</b>", ParagraphStyle('H2', parent=body_style, textColor=colors.white)),
                Paragraph("<b>Inspection Task & Standard Procedure</b>", ParagraphStyle('H3', parent=body_style, textColor=colors.white))
            ]]
            
            for item in qc_items:
                table_data.append([
                    Paragraph(item[0], body_style),
                    Paragraph(f"<b>{item[1]}</b>", body_style),
                    Paragraph(item[2], body_style)
                ])
                
            qc_table = Table(table_data, colWidths=[50, 150, 340])
            qc_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), primary_color),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#dcdcdc")),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('TOPPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(qc_table)
            
            doc.build(story)
            messagebox.showinfo("Success", f"QC Checklist PDF generated successfully at:\n{file_path}!")
        except Exception as e:
            messagebox.showerror("Error", f"Could not generate QC PDF: {e}")

    def fetch_live_exchange_rates(self):
        import urllib.request
        import json
        try:
            url = "https://open.er-api.com/v6/latest/USD"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode())
                if data.get("result") == "success":
                    rates = data.get("rates", {})
                    cny_rate = rates.get("CNY", 7.25)
                    eur_rate = rates.get("EUR", 0.92)
                    self.exchange_rates["CNY"] = cny_rate
                    self.exchange_rates["EUR"] = eur_rate
                    print(f"Successfully fetched live exchange rates: CNY={cny_rate:.4f}, EUR={eur_rate:.4f}")
                    if hasattr(self, 'currency_status_lbl'):
                        self.after(0, lambda: self.currency_status_lbl.configure(text=f"Live Rates: 1 USD = {cny_rate:.2f} CNY | {eur_rate:.2f} EUR", text_color="green"))
        except Exception as e:
            print(f"Failed to fetch live exchange rates: {e}")
        except Exception as e:
            print(f"Failed to fetch live exchange rates: {e}")
            if hasattr(self, 'currency_status_lbl'):
                self.after(0, lambda: self.currency_status_lbl.configure(text="Offline Rates: 1 USD = 7.25 CNY | 0.92 EUR", text_color="grey"))

    def setup_sourcing_matrix_tab(self):
        self.matrix_tab = self.sourcing_tabview.tab("🧮 Sourcing Matrix")
        self.matrix_tab.grid_columnconfigure(0, weight=1)
        self.matrix_tab.grid_rowconfigure(1, weight=1)

        # Header Control Frame
        self.matrix_ctrl_frame = ctk.CTkFrame(self.matrix_tab, fg_color="transparent")
        self.matrix_ctrl_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(15, 10))

        title_lbl = ctk.CTkLabel(self.matrix_ctrl_frame, text="Supplier vs Product Side-by-Side Sourcing Matrix", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.pack(side="left")

        # Dynamic Category combobox
        self.matrix_category_cb = ctk.CTkComboBox(self.matrix_ctrl_frame, values=["All"], command=lambda choice: self.update_sourcing_matrix_tab(), width=180)
        self.matrix_category_cb.pack(side="right", padx=10)
        self.matrix_category_cb.set("All")

        self.matrix_cat_lbl = ctk.CTkLabel(self.matrix_ctrl_frame, text="Select Product Category:")
        self.matrix_cat_lbl.pack(side="right", padx=5)

        # Scrollable Frame to hold Treeview
        self.matrix_frame = ctk.CTkFrame(self.matrix_tab)
        self.matrix_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        self.matrix_frame.grid_columnconfigure(0, weight=1)
        self.matrix_frame.grid_rowconfigure(0, weight=1)

        self.matrix_tree = None

    def update_sourcing_matrix_tab(self):
        for w in self.matrix_frame.winfo_children():
            w.destroy()

        if not self.extracted_data:
            self.make_empty_state(self.matrix_frame, "No sourcing data loaded", "Extract quote files first to build the side-by-side supplier matrix.")
            return

        # Dynamically update category dropdown options
        all_cats = set()
        for r in self.extracted_data:
            p = (r.get("product") or "").strip().title()
            if p:
                all_cats.add(p)
        sorted_cats = ["All"] + sorted(list(all_cats))
        
        # Only configure values if they differ to avoid reset loops
        if hasattr(self, 'matrix_category_cb'):
            current_vals = self.matrix_category_cb.cget("values")
            if list(current_vals) != sorted_cats:
                self.matrix_category_cb.configure(values=sorted_cats)

        suppliers = set()
        products = set()
        for r in self.extracted_data:
            s = self.clean_supplier_name(r.get("supplier"))
            p = (r.get("product") or "").strip().title()
            if s and s != "Unknown":
                suppliers.add(s)
            if p:
                products.add(p)

        sorted_suppliers = sorted(list(suppliers))
        sorted_products = sorted(list(products))

        # Filter products and active suppliers based on category dropdown choice
        if hasattr(self, 'matrix_category_cb'):
            choice = self.matrix_category_cb.get()
            if choice and choice != "All":
                # Find suppliers who quote this product
                active_sups = set()
                for r in self.extracted_data:
                    r_prod = (r.get("product") or "").strip().lower()
                    r_sup = self.clean_supplier_name(r.get("supplier"))
                    if r_prod == choice.lower() and r_sup and r_sup != "Unknown":
                        active_sups.add(r_sup)
                sorted_suppliers = sorted(list(active_sups))
                sorted_products = [p for p in sorted_products if p.lower() == choice.lower()]

        if not sorted_suppliers or not sorted_products:
            ctk.CTkLabel(self.matrix_frame, text="No suppliers or products found to build comparison matrix for selected category.", text_color="grey").pack(pady=40)
            return

        cols = ("product",) + tuple(sorted_suppliers)
        
        scroll_y = ttk.Scrollbar(self.matrix_frame, orient="vertical")
        scroll_x = ttk.Scrollbar(self.matrix_frame, orient="horizontal")
        
        self.matrix_tree = ttk.Treeview(
            self.matrix_frame,
            columns=cols,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            style="Treeview"
        )
        
        scroll_y.config(command=self.matrix_tree.yview)
        scroll_x.config(command=self.matrix_tree.xview)
        
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        self.matrix_tree.pack(side="left", fill="both", expand=True)

        self.matrix_tree.heading("product", text="Product Category")
        self.matrix_tree.column("product", width=160, anchor="w")
        for s in sorted_suppliers:
            self.matrix_tree.heading(s, text=s)
            self.matrix_tree.column(s, width=130, anchor="center")

        currency_choice = self.currency_cb.get()
        factor = 1.0
        symbol = "$"
        if "CNY" in currency_choice:
            factor = self.exchange_rates["CNY"]
            symbol = "¥"
        elif "EUR" in currency_choice:
            factor = self.exchange_rates["EUR"]
            symbol = "€"

        for p in sorted_products:
            row_vals = [p]
            
            prices = {}
            for s in sorted_suppliers:
                for r in self.extracted_data:
                    r_sup = self.clean_supplier_name(r.get("supplier"))
                    r_prod = (r.get("product") or "").strip().lower()
                    if r_sup == s and r_prod == p.lower():
                        price = r.get("price")
                        try:
                            prices[s] = float(price)
                        except (ValueError, TypeError):
                            pass
            
            best_sup = None
            if prices:
                best_sup = min(prices, key=prices.get)

            for s in sorted_suppliers:
                if s in prices:
                    price_val = prices[s] * factor
                    formatted = f"{symbol}{price_val:.5f}" if price_val < 0.1 else f"{symbol}{price_val:.2f}"
                    if s == best_sup and len(prices) > 1:
                        formatted += " (Best)"
                    row_vals.append(formatted)
                else:
                    row_vals.append("N/A")

            self.matrix_tree.insert("", tk.END, values=row_vals)

    def open_rfq_broadcast_popup(self):
        issues, warnings = self.validate_rfq_ready()
        self.update_rfq_readiness_panel()
        if issues:
            note = "; ".join(issues)
            self.log_workflow_blocked_attempt("RFQ", "Broadcast Blocked", note)
            messagebox.showerror(
                "RFQ Not Ready",
                "Fix these issues before broadcasting the RFQ:\n\n- " + "\n- ".join(issues)
            )
            return
        suppliers = set()
        for r in self.extracted_data:
            s = self.clean_supplier_name(r.get("supplier"))
            if s and s != "Unknown":
                suppliers.add(s)
        sorted_sups = sorted(list(suppliers))

        if not sorted_sups:
            messagebox.showwarning("Warning", "No suppliers found in the database to broadcast to.")
            return

        popup = ctk.CTkToplevel(self)
        popup.title("Multi-Supplier RFQ Broadcaster")
        popup.geometry("500x450")
        popup.resizable(False, False)
        popup.attributes("-topmost", True)

        ctk.CTkLabel(popup, text="Select Suppliers for RFQ Outreach", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)

        checklist_scroll = ctk.CTkScrollableFrame(popup, height=250)
        checklist_scroll.pack(fill="both", expand=True, padx=20, pady=10)

        checkbox_vars = {}
        supplier_emails = {}
        for s in sorted_sups:
            var = tk.IntVar(value=1)
            checkbox_vars[s] = var
            
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT contact_info FROM supplier_contacts WHERE supplier = ?", (s,))
            row = c.fetchone()
            conn.close()
            email_info = ""
            if row:
                import re
                email_match = re.search(r'[\w\.-]+@[\w\.-]+', row[0])
                if email_match:
                    email_info = f" ({email_match.group(0)})"
                    supplier_emails[s] = email_match.group(0)
            
            ctk.CTkCheckBox(checklist_scroll, text=f"{s}{email_info}", variable=var).pack(anchor="w", pady=4, padx=10)

        def proceed_outreach():
            selected_sups = [s for s, var in checkbox_vars.items() if var.get() == 1]
            if not selected_sups:
                self.log_workflow_blocked_attempt("RFQ", "Broadcast Blocked", "No suppliers selected for RFQ outreach.")
                messagebox.showwarning("Warning", "Please select at least one supplier!")
                return
            missing_emails = [s for s in selected_sups if not supplier_emails.get(s)]
            combined_warnings = list(warnings)
            if missing_emails:
                combined_warnings.append(
                    "Missing supplier email for: " + ", ".join(missing_emails[:8])
                    + ("..." if len(missing_emails) > 8 else "")
                )
            if combined_warnings and not messagebox.askyesno(
                "RFQ Broadcast Warnings",
                "Review these outreach warnings:\n\n- "
                + "\n- ".join(combined_warnings)
                + "\n\nContinue generating drafts?"
            ):
                self.log_workflow_blocked_attempt("RFQ", "Broadcast Blocked", "; ".join(combined_warnings))
                return
            popup.destroy()
            self.generate_personalized_rfq_outreach(selected_sups)

        btn_proceed = ctk.CTkButton(popup, text="⚡ Generate Personalized Outreach", fg_color="#1f538d", hover_color="#153e6b", command=proceed_outreach)
        btn_proceed.pack(pady=15, padx=20, fill="x")

    def generate_personalized_rfq_outreach(self, selected_suppliers):
        outreach_win = ctk.CTkToplevel(self)
        outreach_win.title("Broadcasting RFQ Outreach Drafts")
        outreach_win.geometry("700x550")
        outreach_win.attributes("-topmost", True)

        ctk.CTkLabel(outreach_win, text="AI Personalized RFQ Outreach Emails", font=ctk.CTkFont(size=18, weight="bold")).pack(pady=10)

        tab_control = ctk.CTkTabview(outreach_win)
        tab_control.pack(fill="both", expand=True, padx=15, pady=10)

        prod_name = self.rfq_name_entry.get().strip() or "Product"
        target_qty = self.rfq_qty_entry.get().strip() or "100000"
        price_term = self.rfq_term_cb.get() or "FOB"
        lead_time = self.rfq_lead_entry.get().strip() or "30 days"
        payment_term = self.rfq_payment_entry.get().strip() or "30/70"
        specs = self.rfq_specs_text.get("1.0", "end-1c").strip()
        rfq_id = self.save_rfq_record("Sent")
        if rfq_id:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute(
                "UPDATE rfq_register SET selected_suppliers=?, updated_at=datetime('now') WHERE id=?",
                (", ".join(selected_suppliers), rfq_id)
            )
            c.execute("""
                INSERT INTO workflow_audit_log (workflow_type, workflow_id, action, status, note)
                VALUES ('RFQ', ?, 'Outreach Generated', 'Sent', ?)
            """, (rfq_id, f"RFQ outreach generated for {len(selected_suppliers)} supplier(s)."))
            conn.commit()
            conn.close()
            self.load_rfq_register()

        for s in selected_suppliers:
            tab_control.add(s)
            s_tab = tab_control.tab(s)
            s_tab.grid_columnconfigure(0, weight=1)
            s_tab.grid_rowconfigure(1, weight=1)

            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT contact_info FROM supplier_contacts WHERE supplier = ?", (s,))
            row = c.fetchone()
            conn.close()
            
            sup_email = "sales@supplier.com"
            if row:
                import re
                email_match = re.search(r'[\w\.-]+@[\w\.-]+', row[0])
                if email_match:
                    sup_email = email_match.group(0)

            prev_quote_text = ""
            for r in self.extracted_data:
                r_sup = self.clean_supplier_name(r.get("supplier"))
                if r_sup == s:
                    prev_quote_text = f"In our database, we see this supplier previously quoted {r.get('product')} at ${r.get('price')} FOB."
                    break

            ai_prompt = f"""
            You are a professional purchasing agent. Write a personalized RFQ email to the supplier '{s}'.
            Email Address: {sup_email}
            We want to request a quotation for:
            - Product: {prod_name}
            - Quantity: {target_qty} pcs
            - Required Price Terms: {price_term}
            - Lead Time: {lead_time}
            - Payment Terms: {payment_term}
            - Technical Specs: {specs}
            
            Personalization Context: {prev_quote_text} (If this supplier previously gave a quote, politely reference it and ask if they can improve the pricing/MOQ for this larger RFQ).
            
            Keep the email highly professional, clear, and ready to send.
            Do not use markdown formatting.
            """

            try:
                outreach_body = self.generate_with_fallback([], ai_prompt, json_response=False)
            except Exception:
                outreach_body = f"Dear {s} Sales Team,\n\nWe would like to request a quotation for {prod_name}. Details are as follows:\n- Quantity: {target_qty}\n- Specs: {specs}\n\nPlease let us know your best pricing and lead time.\n\nBest regards,\nProcurement Team"

            txt = ctk.CTkTextbox(s_tab, height=300)
            txt.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
            txt.insert("1.0", outreach_body)

            ctrl_fr = ctk.CTkFrame(s_tab, fg_color="transparent")
            ctrl_fr.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

            def make_launch_cmd(email=sup_email, body_txt_box=txt, s_name=s):
                return lambda: self.launch_rfq_email_client(email, prod_name, body_txt_box.get("1.0", "end-1c").strip())

            btn_launch = ctk.CTkButton(ctrl_fr, text="✉ Open in Mail Client", fg_color="#1f538d", hover_color="#153e6b", command=make_launch_cmd(sup_email, txt, s))
            btn_launch.pack(side="right", padx=5)

            btn_copy = ctk.CTkButton(ctrl_fr, text="📋 Copy Draft", command=lambda b=txt: self.copy_to_clipboard(b.get("1.0", "end-1c").strip()))
            btn_copy.pack(side="right", padx=5)

    def launch_rfq_email_client(self, email, product, body):
        import urllib.parse
        import webbrowser
        subject = f"Request for Quote: {product}"
        mail_url = f"mailto:{email}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        webbrowser.open(mail_url)

    def setup_uae_customs_tab(self):
        tab_uae = self.search_tabview.tab("🇦🇪 UAE Customs & HS Code")
        tab_uae.grid_columnconfigure(0, weight=1)
        tab_uae.grid_columnconfigure(1, weight=1)
        tab_uae.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_uae, text="UAE Customs Tariff & HS Code Classifier", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Sourcing Inputs ---
        left_frame = ctk.CTkFrame(tab_uae)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="🇦🇪 Import Configuration", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=10, sticky="w")

        ctk.CTkLabel(left_frame, text="Product Category:").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        self.uae_category_cb = ctk.CTkComboBox(left_frame, values=["Custom"])
        self.uae_category_cb.grid(row=1, column=1, padx=15, pady=5, sticky="ew")

        ctk.CTkLabel(left_frame, text="Or Custom Product Name:").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        self.uae_custom_name = ctk.CTkEntry(left_frame, placeholder_text="e.g. Paper Cup / Mask")
        self.uae_custom_name.grid(row=2, column=1, padx=15, pady=5, sticky="ew")

        ctk.CTkLabel(left_frame, text="CIF Value (USD):").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        self.uae_cif_entry = ctk.CTkEntry(left_frame)
        self.uae_cif_entry.grid(row=3, column=1, padx=15, pady=5, sticky="ew")
        self.uae_cif_entry.insert(0, "15000")

        # Country of Origin Dropdown for FTA CEPA Exemption Checker
        ctk.CTkLabel(left_frame, text="Country of Origin:").grid(row=4, column=0, padx=15, pady=5, sticky="w")
        self.uae_origin_cb = ctk.CTkComboBox(left_frame, values=["China", "India", "GCC (Saudi/Oman/etc)", "USA", "Europe"])
        self.uae_origin_cb.grid(row=4, column=1, padx=15, pady=5, sticky="ew")
        self.uae_origin_cb.set("India")

        self.btn_classify_uae = ctk.CTkButton(left_frame, text="🔍 Classify HS Code & Duties", fg_color="#1f538d", hover_color="#153e6b", command=self.run_uae_customs_evaluation)
        self.btn_classify_uae.grid(row=5, column=0, columnspan=2, padx=15, pady=25, sticky="ew")

        # --- RIGHT PANEL: AI Customs Analysis Summary ---
        right_frame = ctk.CTkFrame(tab_uae)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📋 UAE Customs Duty Projections", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        self.uae_customs_output = ctk.CTkTextbox(right_frame, font=("Consolas", 11))
        self.uae_customs_output.grid(row=1, column=0, padx=15, pady=10, sticky="nsew")
        self.uae_customs_output.insert("1.0", "Select a product and CIF value to calculate UAE duties and look up HS Tariff codes.")
        self.uae_customs_output.configure(state="disabled")

    def update_uae_customs_tab(self):
        products = set()
        for r in self.extracted_data:
            prod = (r.get("product") or "").strip().title()
            if prod:
                products.add(prod)
        sorted_prods = ["Custom"] + sorted(list(products))
        self.uae_category_cb.configure(values=sorted_prods)
        if sorted_prods:
            self.uae_category_cb.set(sorted_prods[0])

    def run_uae_customs_evaluation(self):
        choice = self.uae_category_cb.get()
        custom = self.uae_custom_name.get().strip()
        
        prod_name = custom if choice == "Custom" else choice
        if not prod_name:
            messagebox.showwarning("Warning", "Please select a category or enter a custom product name!")
            return
            
        try:
            cif_val = float(self.uae_cif_entry.get().strip())
        except ValueError:
            messagebox.showerror("Error", "CIF Value must be a valid number!")
            return

        self.btn_classify_uae.configure(state="disabled", text="Querying UAE Tariff Database...")
        self.update()

        def run_ai_classification():
            ai_prompt = f"""
            Identify the Harmonized System (HS) Code and estimate the import customs tariff details for:
            - Product: {prod_name}
            - Import Country: United Arab Emirates (UAE)
            - CIF Value: ${cif_val:.2f} USD
            
            Determine:
            1. Recommended 6-digit Harmonized System (HS) Code.
            2. Standard UAE Customs Duty Rate (usually 5% for standard goods under the GCC Common Customs Law).
            3. Excise Tax (if applicable, e.g. sugary/energy drinks, tobacco).
            4. Import VAT Rate (5% in UAE).
            5. Sourcing & Documentation compliance recommendations for clearing Dubai/UAE Customs (e.g. Certificate of Origin legalized by UAE Embassy, Commercial Invoice, Packing List, Industrial/Commercial Trade License).
            
            Write the output in a clean, highly structured dashboard format.
            Do not use markdown syntax. Estimate the duty in AED (assume 1 USD = 3.6725 AED).
            """
            
            try:
                result_text = self.generate_with_fallback([], ai_prompt, json_response=False)
                
                origin = self.uae_origin_cb.get()
                cif_aed = cif_val * 3.6725
                duty_rate = 0.05
                is_fta = False
                
                if origin in ["India", "GCC (Saudi/Oman/etc)"]:
                    duty_rate = 0.0
                    is_fta = True
                
                excise_rate = 0.0
                if any(x in prod_name.lower() for x in ["drink", "tobacco", "cigarette", "sugar"]):
                    excise_rate = 0.50
                
                est_duty_aed = cif_aed * duty_rate
                est_excise_aed = cif_aed * excise_rate
                est_vat_aed = (cif_aed + est_duty_aed + est_excise_aed) * 0.05
                total_cleared_aed = cif_aed + est_duty_aed + est_excise_aed + est_vat_aed

                fta_banner = ""
                if is_fta:
                    fta_banner = f"★ FTA CEPA BENEFIT ACTIVE: 0.0% Preferential Duty (Origin: {origin})\n--------------------------------------------------\n"

                summary_header = f"""==================================================
🇦🇪 UAE CUSTOMS IMPORT CLEARANCE PROJECTION
==================================================
Target Product : {prod_name}
Country of Origin: {origin}
CIF Import Value: ${cif_val:,.2f} USD ({cif_aed:,.2f} AED)
--------------------------------------------------
{fta_banner}ESTIMATED CLEARANCE COST BREAKDOWN:
- GCC Customs Duty ({int(duty_rate*100)}%)    : {est_duty_aed:,.2f} AED
- Estimated Excise Tax           : {est_excise_aed:,.2f} AED
- Estimated UAE Import VAT (5%)  : {est_vat_aed:,.2f} AED
- Total Landed Cost (In UAE)     : {total_cleared_aed:,.2f} AED
--------------------------------------------------
AI TARIFF & COMPLIANCE DETAIL:
{result_text}
"""

                self.after(0, lambda: display_result(summary_header))
            except Exception as e:
                self.after(0, lambda err=e: display_result(f"Failed to run classification:\n{err}"))
            finally:
                self.after(0, lambda: self.btn_classify_uae.configure(state="normal", text="🔍 Classify HS Code & Duties"))

        def display_result(text):
            self.uae_customs_output.configure(state="normal")
            self.uae_customs_output.delete("1.0", tk.END)
            self.uae_customs_output.insert("1.0", text)
            self.uae_customs_output.configure(state="disabled")

        threading.Thread(target=run_ai_classification, daemon=True).start()

    def on_scorecard_select(self, event):
        selected = self.scorecard_tree.selection()
        if not selected:
            return
        item = self.scorecard_tree.item(selected[0])
        vals = item.get("values")
        if not vals:
            return
            
        supplier = vals[1]
        prod = vals[2]
        
        matrix = self.calculate_supplier_scorecard()
        for row in matrix:
            if self.clean_supplier_name(row["supplier"]) == supplier and row["product"] == prod:
                self.draw_scorecard_radar_chart(
                    supplier,
                    row.get("price_score", 80),
                    row.get("lead_score", 80),
                    row.get("moq_score", 80),
                    row.get("risk_score", 80)
                )
                break

    def draw_scorecard_radar_chart(self, supplier_name, price_score, lead_score, moq_score, risk_score):
        if hasattr(self, 'radar_canvas_frame'):
            self.radar_canvas_frame.destroy()

        self.radar_canvas_frame = ctk.CTkFrame(self.scorecard_sim_frame, fg_color="transparent")
        self.radar_canvas_frame.grid(row=10, column=0, padx=15, pady=10, sticky="nsew")

        import matplotlib.pyplot as plt
        import numpy as np
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        labels = np.array(['Price', 'Speed', 'MOQ', 'Compliance'])
        stats = np.array([price_score, lead_score, moq_score, risk_score])

        angles = np.linspace(0, 2*np.pi, len(labels), endpoint=False).tolist()
        stats = np.concatenate((stats,[stats[0]]))
        angles = np.concatenate((angles,[angles[0]]))

        fig, ax = plt.subplots(figsize=(2.2, 2.2), subplot_kw=dict(polar=True), dpi=100)
        fig.patch.set_facecolor(self.THEME["surface"])
        ax.set_facecolor(self.THEME["surface"])

        ax.plot(angles, stats, color='#1f538d', linewidth=2)
        ax.fill(angles, stats, color='#1f538d', alpha=0.25)

        ax.set_thetagrids(np.degrees(angles[:-1]), labels, color=self.THEME["text"], fontsize=8)
        ax.set_ylim(0, 100)
        ax.set_rgrids([25, 50, 75, 100], angle=0)
        ax.tick_params(colors=self.THEME["muted"])
        
        ax.spines['polar'].set_visible(False)
        ax.grid(color=self.THEME["border"], linestyle='--')

        ax.set_title(f"Strengths: {supplier_name[:12]}", color=self.THEME["text"], fontsize=9, pad=10)
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.radar_canvas_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    def setup_container_packing_tab(self):
        tab_packing = self.logistics_tabview.tab("🚢 Container Packing")
        tab_packing.grid_columnconfigure(0, weight=1)
        tab_packing.grid_columnconfigure(1, weight=1)
        tab_packing.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_packing, text="Logistics Container Capacity Packing Simulator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Packing Configuration ---
        left_frame = ctk.CTkFrame(tab_packing)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="📦 Packing Settings", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=10, sticky="w")

        ctk.CTkLabel(left_frame, text="Select Product:").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        self.packing_product_cb = ctk.CTkComboBox(left_frame, values=[], command=self.on_packing_product_changed)
        self.packing_product_cb.grid(row=1, column=1, padx=15, pady=5, sticky="ew")

        ctk.CTkLabel(left_frame, text="Carton Pack Info:").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        self.packing_info_lbl = ctk.CTkLabel(left_frame, text="1000 pcs/ctn, 0.045 CBM", font=ctk.CTkFont(weight="bold"))
        self.packing_info_lbl.grid(row=2, column=1, padx=15, pady=5, sticky="w")

        ctk.CTkLabel(left_frame, text="Order Quantity (pcs):").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        self.packing_qty_entry = ctk.CTkEntry(left_frame)
        self.packing_qty_entry.grid(row=3, column=1, padx=15, pady=5, sticky="ew")
        self.packing_qty_entry.insert(0, "100000")

        ctk.CTkLabel(left_frame, text="Container Type:").grid(row=4, column=0, padx=15, pady=5, sticky="w")
        self.packing_container_cb = ctk.CTkComboBox(left_frame, values=["20GP (28 CBM)", "40GP (58 CBM)", "40HQ (68 CBM)"])
        self.packing_container_cb.grid(row=4, column=1, padx=15, pady=5, sticky="ew")
        self.packing_container_cb.set("20GP (28 CBM)")

        self.btn_run_packing = ctk.CTkButton(left_frame, text="🚢 Simulate Container Packing", fg_color="#1f538d", hover_color="#153e6b", command=self.run_container_packing_simulation)
        self.btn_run_packing.grid(row=5, column=0, columnspan=2, padx=15, pady=25, sticky="ew")

        # --- RIGHT PANEL: Visual Map & Summary ---
        right_frame = ctk.CTkFrame(tab_packing)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📊 Utilization Summary", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        self.packing_summary_box = ctk.CTkTextbox(right_frame, height=120)
        self.packing_summary_box.grid(row=1, column=0, padx=15, pady=5, sticky="ew")
        self.packing_summary_box.insert("1.0", "Configure parameters on the left and click Simulate.")
        self.packing_summary_box.configure(state="disabled")

        self.packing_canvas_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        self.packing_canvas_frame.grid(row=2, column=0, padx=15, pady=10, sticky="nsew")

    def update_container_packing_tab(self):
        products = set()
        for r in self.extracted_data:
            prod = (r.get("product") or "").strip().title()
            if prod:
                products.add(prod)
        sorted_prods = sorted(list(products))
        self.packing_product_cb.configure(values=sorted_prods)
        if sorted_prods:
            self.packing_product_cb.set(sorted_prods[0])
            self.on_packing_product_changed(sorted_prods[0])

    def on_packing_product_changed(self, choice):
        for r in self.extracted_data:
            if (r.get("product") or "").strip().lower() == choice.lower():
                pack_details = r.get("packing") or "1000 pcs/ctn, 0.045 CBM"
                self.packing_info_lbl.configure(text=pack_details)
                break

    def run_container_packing_simulation(self):
        choice = self.packing_product_cb.get()
        if not choice:
            messagebox.showwarning("Warning", "Please select a product first!")
            return
            
        try:
            qty = float(self.packing_qty_entry.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Quantity must be a valid number!")
            return

        pack_text = self.packing_info_lbl.cget("text")
        pcs_per_ctn, unit_cbm = self.parse_packing_metrics(pack_text)

        import math
        total_ctns = math.ceil(qty / pcs_per_ctn)
        total_cbm = total_ctns * unit_cbm

        container_choice = self.packing_container_cb.get()
        max_cbm = 28.0
        container_name = "20GP"
        if "40GP" in container_choice:
            max_cbm = 58.0
            container_name = "40GP"
        elif "40HQ" in container_choice:
            max_cbm = 68.0
            container_name = "40HQ"

        fill_percent = (total_cbm / max_cbm) * 100.0

        if fill_percent > 100.0:
            rec = f"⚠️ OVERFILLED! Your shipment requires {total_cbm:.2f} CBM which exceeds the {container_name} capacity of {max_cbm} CBM. Upgrade your container size or split the load."
        elif fill_percent >= 80.0:
            rec = f"🟢 OPTIMAL! Your shipment fills {fill_percent:.1f}% of the {container_name} container, representing highly efficient container utilization."
        else:
            rec = f"🟡 UNDERFILLED! You are only utilizing {fill_percent:.1f}% of the {container_name} container. Consider ordering more products to maximize your logistics freight rates."

        summary = f"""Total Cartons: {total_ctns} ctns
Total CBM: {total_cbm:.3f} CBM
Container Capacity: {max_cbm} CBM ({container_name})
Space Utilized: {fill_percent:.1f}%

Sourcing Action Guidance:
{rec}"""

        self.packing_summary_box.configure(state="normal")
        self.packing_summary_box.delete("1.0", tk.END)
        self.packing_summary_box.insert("1.0", summary)
        self.packing_summary_box.configure(state="disabled")

        for w in self.packing_canvas_frame.winfo_children():
            w.destroy()

        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        fig, ax = plt.subplots(figsize=(5.5, 2.0), dpi=100)
        fig.patch.set_facecolor(self.THEME["surface"])
        ax.set_facecolor(self.THEME["surface"])

        fill_width = 10.0 * (min(fill_percent, 100.0) / 100.0)
        bar_color = "#1f538d" if fill_percent <= 100 else "#bf3b3b"

        if fill_width > 0:
            ax.barh(1, fill_width, height=1.0, color=bar_color, edgecolor=self.THEME["surface"], label='Filled')
        if fill_width < 10.0:
            ax.barh(1, 10.0 - fill_width, left=fill_width, height=1.0, color=self.THEME["surface_alt"], edgecolor=self.THEME["surface"], label='Empty')

        ax.set_xlim(0, 10)
        ax.set_ylim(0, 2)
        ax.axis('off')

        ax.text(5, 1, f"{fill_percent:.1f}% Filled", color=self.THEME["text"], ha='center', va='center', weight='bold', fontsize=11)
        ax.set_title(f"{container_name} Container Space Utilization Map", color=self.THEME["text"], fontsize=9, pad=5)

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.packing_canvas_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    def setup_visual_search_tab(self):
        tab_search = self.search_tabview.tab("🔍 AI Visual Search")
        tab_search.grid_columnconfigure(0, weight=1)
        tab_search.grid_columnconfigure(1, weight=1)
        tab_search.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_search, text="AI Sourcing Search by Product Image Similarity", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Upload & Search Run ---
        left_frame = ctk.CTkFrame(tab_search)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left_frame, text="📷 Upload Target Product Image", font=ctk.CTkFont(size=15, weight="bold")).pack(pady=10, padx=15, anchor="w")

        self.search_img_preview = ctk.CTkLabel(left_frame, text="No Image Selected", width=220, height=220, fg_color=self.THEME["surface_soft"], text_color=self.THEME["muted"], corner_radius=8)
        self.search_img_preview.pack(pady=15)

        btn_upload = ctk.CTkButton(left_frame, text="📁 Select Product Photo", command=self.select_search_product_photo)
        btn_upload.pack(pady=5, padx=20, fill="x")

        self.btn_run_search = ctk.CTkButton(left_frame, text="⚡ Run Visual Search", fg_color="#1f538d", hover_color="#153e6b", command=self.run_visual_similarity_search)
        self.btn_run_search.pack(pady=20, padx=20, fill="x")
        self.uploaded_search_img_path = ""

        # --- RIGHT PANEL: Results list ---
        right_frame = ctk.CTkFrame(tab_search)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="🎯 Matching Supplier Quotes", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=15, pady=10, sticky="w")

        self.search_results_scroll = ctk.CTkScrollableFrame(right_frame, fg_color=self.THEME["surface_soft"])
        self.search_results_scroll.grid(row=1, column=0, padx=15, pady=10, sticky="nsew")
        self.search_results_scroll.grid_columnconfigure(0, weight=1)

        self.make_empty_state(self.search_results_scroll, "Ready for visual search", "Upload a product photo to match it against extracted supplier quote records.")

    def select_search_product_photo(self):
        path = filedialog.askopenfilename(
            filetypes=[("Image files", "*.png;*.jpg;*.jpeg")]
        )
        if not path:
            return
        self.uploaded_search_img_path = path
        try:
            from PIL import Image
            img = Image.open(path)
            img.thumbnail((220, 220))
            ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
            self.search_img_preview.configure(image=ctk_img, text="")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load image preview: {e}")

    def run_visual_similarity_search(self):
        if not self.uploaded_search_img_path:
            messagebox.showwarning("Warning", "Please select a product photo to search for first!")
            return

        for w in self.search_results_scroll.winfo_children():
            w.destroy()

        candidates = []
        for r in self.extracted_data:
            media = r.get("attached_media") or ""
            if media:
                media_files = [m.strip() for m in media.split(";") if m.strip()]
                for m_file in media_files:
                    ext = os.path.splitext(m_file)[1].lower()
                    if ext in {".png", ".jpg", ".jpeg"}:
                        abs_p = os.path.abspath(m_file)
                        if os.path.exists(abs_p):
                            candidates.append((r, abs_p))
                            break

        if not candidates:
            ctk.CTkLabel(self.search_results_scroll, text="⚠️ No quote records in database contain attached images to compare against.", text_color="grey").pack(pady=40)
            return

        self.btn_run_search.configure(state="disabled", text="Running AI Visual Scan...")
        self.update()

        def run_ai_scan():
            try:
                from google.genai import types
                
                with open(self.uploaded_search_img_path, "rb") as f:
                    target_bytes = f.read()

                parts = [
                    types.Part.from_bytes(data=target_bytes, mime_type="image/jpeg"),
                    "Above is the Target Product Image we are trying to find matches for in our database.\n"
                ]

                part_idx_to_candidate = {}
                for idx, (r, cand_path) in enumerate(candidates[:5]):
                    with open(cand_path, "rb") as f:
                        cand_bytes = f.read()
                    
                    p_num = idx + 2
                    parts.append(types.Part.from_bytes(data=cand_bytes, mime_type="image/jpeg"))
                    parts.append(f"Candidate Reference #{p_num} (Supplier: {self.clean_supplier_name(r.get('supplier'))}, Product: {r.get('product')})\n")
                    part_idx_to_candidate[p_num] = r

                ai_prompt = """
                Compare the Target Product Image (Image 1) against each of the Candidate Reference images.
                Rate the visual similarity (matching product design, application, material, color, shape) of each Candidate on a score from 0 to 100.
                
                Return the results strictly matching this JSON schema:
                {
                  "matches": [
                    {
                      "candidate_index": 2,
                      "score": 90,
                      "rationale": "Brief 1-sentence rationale of why it matches or differs"
                    }
                  ]
                }
                """
                parts.append(ai_prompt)

                response_text = self.generate_with_fallback(parts, "", json_response=True)
                result = json.loads(response_text)
                matches = result.get("matches", [])

                matches.sort(key=lambda x: x.get("score", 0), reverse=True)

                self.after(0, lambda: self.btn_run_search.configure(state="normal", text="⚡ Run Visual Search"))
                self.after(0, lambda: display_matches(matches, part_idx_to_candidate))

            except Exception as e:
                self.after(0, lambda: self.btn_run_search.configure(state="normal", text="⚡ Run Visual Search"))
                self.after(0, lambda err=e: display_error(err))

        def display_error(err):
            ctk.CTkLabel(self.search_results_scroll, text=f"❌ AI Visual Scan failed:\n{err}", text_color="#ffa6a6").pack(pady=20)

        def display_matches(matches, part_idx_to_candidate):
            if not matches:
                self.make_empty_state(self.search_results_scroll, "No similar products found", "Try another product image or extract more supplier quote data.")
                return

            for m in matches:
                cand_idx = m.get("candidate_index")
                score = m.get("score", 0)
                rationale = m.get("rationale", "")
                
                quote = part_idx_to_candidate.get(cand_idx)
                if not quote:
                    continue

                sup = self.clean_supplier_name(quote.get("supplier"))
                prod = quote.get("product") or "Product"
                price = quote.get("price") or 0.0
                
                card = ctk.CTkFrame(self.search_results_scroll, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
                card.pack(fill="x", pady=4, padx=5)

                header = ctk.CTkFrame(card, fg_color="transparent")
                header.pack(fill="x", padx=10, pady=5)

                badge_color = self.THEME["success_soft"] if score >= 80 else (self.THEME["warning_soft"] if score >= 50 else self.THEME["danger_soft"])
                badge_text_color = self.THEME["success"] if score >= 80 else (self.THEME["warning"] if score >= 50 else self.THEME["danger"])
                
                badge = ctk.CTkLabel(header, text=f" {score}% Match ", fg_color=badge_color, text_color=badge_text_color, font=ctk.CTkFont(weight="bold", size=11))
                badge.pack(side="left")

                ctk.CTkLabel(header, text=f" {sup} — {prod}", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10)
                ctk.CTkLabel(header, text=f"${price:.4f}/pc" if price < 0.1 else f"${price:.2f}/pc", text_color=self.THEME["primary"], font=ctk.CTkFont(weight="bold")).pack(side="right")

                ctk.CTkLabel(card, text=rationale, font=ctk.CTkFont(size=11), text_color="grey", justify="left").pack(padx=15, pady=(0, 10), fill="x")

        threading.Thread(target=run_ai_scan, daemon=True).start()

    def log_supplier_incident(self):
        supplier = self.qc_supplier_cb.get()
        if not supplier or supplier == "Select Supplier":
            messagebox.showwarning("Warning", "Please select a supplier first!")
            return
            
        inc_type = self.qc_inc_type_cb.get()
        severity = self.qc_inc_sev_cb.get()
        desc = self.qc_inc_desc_entry.get().strip()
        
        if not desc:
            messagebox.showwarning("Warning", "Please enter an incident description!")
            return
            
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                INSERT INTO supplier_incidents (supplier, incident_type, description, severity)
                VALUES (?, ?, ?, ?)
            """, (supplier, inc_type, desc, severity))
            conn.commit()
            conn.close()
            
            self.qc_inc_desc_entry.delete(0, tk.END)
            messagebox.showinfo("Success", f"Operational incident ({inc_type} - {severity}) logged for {supplier}!")
            self.load_all_quotes_from_db()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log incident: {e}")

    def setup_defect_log_tab(self):
        tab_incidents = self.scorecard_tabview.tab("⚠️ Defect Log")
        tab_incidents.grid_columnconfigure(0, weight=1)
        tab_incidents.grid_columnconfigure(1, weight=2)
        tab_incidents.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_incidents, text="Supplier Defect & Quality Incident Log", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Add Incident Form ---
        left_frame = ctk.CTkFrame(tab_incidents)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left_frame, text="🚨 Log New Incident", font=ctk.CTkFont(size=15, weight="bold")).pack(pady=10, padx=15, anchor="w")

        ctk.CTkLabel(left_frame, text="Select Supplier:").pack(padx=15, pady=(5, 2), anchor="w")
        self.inc_supplier_cb = ctk.CTkComboBox(left_frame, values=["Select Supplier"], width=230)
        self.inc_supplier_cb.pack(padx=15, pady=2, anchor="w")

        ctk.CTkLabel(left_frame, text="Incident Type:").pack(padx=15, pady=(5, 2), anchor="w")
        self.inc_type_cb = ctk.CTkComboBox(left_frame, values=["Defect Rate", "Delivery Delay", "Communication Delay", "Other Dispute"], width=230)
        self.inc_type_cb.pack(padx=15, pady=2, anchor="w")

        ctk.CTkLabel(left_frame, text="Severity Level:").pack(padx=15, pady=(5, 2), anchor="w")
        self.inc_sev_cb = ctk.CTkComboBox(left_frame, values=["Low", "Medium", "High"], width=230)
        self.inc_sev_cb.pack(padx=15, pady=2, anchor="w")

        ctk.CTkLabel(left_frame, text="Description:").pack(padx=15, pady=(5, 2), anchor="w")
        self.inc_desc_entry = ctk.CTkEntry(left_frame, placeholder_text="Details of quality issue...", width=230)
        self.inc_desc_entry.pack(padx=15, pady=2, anchor="w")

        btn_log = ctk.CTkButton(left_frame, text="🚨 Log Incident", fg_color="#bf3b3b", hover_color="#9e2d2d", command=self.add_incident_log)
        btn_log.pack(padx=15, pady=15, fill="x")

        # --- RIGHT PANEL: Incidents History Table ---
        right_frame = ctk.CTkFrame(tab_incidents)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(0, weight=1)

        scroll_y = ttk.Scrollbar(right_frame, orient="vertical")
        cols = ("id", "supplier", "type", "description", "severity", "date")
        self.inc_tree = ttk.Treeview(right_frame, columns=cols, show="headings", yscrollcommand=scroll_y.set, style="Treeview")
        scroll_y.config(command=self.inc_tree.yview)

        self.inc_tree.heading("id", text="ID")
        self.inc_tree.heading("supplier", text="Supplier")
        self.inc_tree.heading("type", text="Type")
        self.inc_tree.heading("description", text="Description")
        self.inc_tree.heading("severity", text="Severity")
        self.inc_tree.heading("date", text="Date")

        self.inc_tree.column("id", width=40, anchor="center")
        self.inc_tree.column("supplier", width=120, anchor="w")
        self.inc_tree.column("type", width=100, anchor="w")
        self.inc_tree.column("description", width=200, anchor="w")
        self.inc_tree.column("severity", width=80, anchor="center")
        self.inc_tree.column("date", width=120, anchor="center")

        self.inc_tree.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=10)
        scroll_y.grid(row=0, column=1, sticky="ns", pady=10, padx=(0, 10))

        btn_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        btn_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        
        btn_delete = ctk.CTkButton(btn_frame, text="🗑 Delete Selected Log", fg_color="#a83232", hover_color="#8c2626", command=self.delete_incident_log)
        btn_delete.pack(side="left", fill="x", expand=True, padx=(0, 5))

        btn_scar = ctk.CTkButton(btn_frame, text="📄 Generate SCAR PDF", fg_color="#1f538d", hover_color="#153e6b", command=self.export_scar_pdf)
        btn_scar.pack(side="right", fill="x", expand=True, padx=(5, 0))

        self.load_incident_logs()

    def load_incident_logs(self):
        # Update dropdown values first
        suppliers = set()
        for r in self.extracted_data:
            s = r.get("supplier")
            if s and s != "Unknown":
                suppliers.add(s)
        sorted_sups = sorted(list(suppliers))
        if hasattr(self, 'inc_supplier_cb'):
            self.inc_supplier_cb.configure(values=sorted_sups)
            if sorted_sups:
                self.inc_supplier_cb.set(sorted_sups[0])

        self.inc_tree.delete(*self.inc_tree.get_children())
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT id, supplier, incident_type, description, severity, logged_date FROM supplier_incidents ORDER BY logged_date DESC")
            for row in c.fetchall():
                self.inc_tree.insert("", "end", values=row)
            conn.close()
        except Exception as e:
            print(f"Failed to load incident logs: {e}")

    def add_incident_log(self):
        supplier = self.inc_supplier_cb.get()
        inc_type = self.inc_type_cb.get()
        severity = self.inc_sev_cb.get()
        desc = self.inc_desc_entry.get().strip()

        if not supplier or supplier == "Select Supplier":
            messagebox.showwarning("Warning", "Please select a supplier first!")
            return
        if not desc:
            messagebox.showwarning("Warning", "Please enter an incident description!")
            return

        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                INSERT INTO supplier_incidents (supplier, incident_type, description, severity)
                VALUES (?, ?, ?, ?)
            """, (supplier, inc_type, desc, severity))
            conn.commit()
            conn.close()

            self.inc_desc_entry.delete(0, tk.END)
            messagebox.showinfo("Success", f"Operational incident ({inc_type} - {severity}) logged for {supplier}!")
            self.load_incident_logs()
            self.update_scorecard_tab() # Refresh scorecard in real-time
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log incident: {e}")

    def delete_incident_log(self):
        selected = self.inc_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an incident log to delete!")
            return
        
        item = self.inc_tree.item(selected[0])
        log_id = item["values"][0]
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("DELETE FROM supplier_incidents WHERE id = ?", (log_id,))
            conn.commit()
            conn.close()
            messagebox.showinfo("Deleted", "Incident log deleted successfully!")
            self.load_incident_logs()
            self.update_scorecard_tab()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete incident log: {e}")

    def setup_currency_hedging_tab(self):
        tab_hedge = self.sourcing_tabview.tab("💵 Currency Hedging")
        tab_hedge.grid_columnconfigure(0, weight=1)
        tab_hedge.grid_columnconfigure(1, weight=2)
        tab_hedge.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_hedge, text="Exchange Rate Hedging & Landed Cost Fluctuation Simulator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Sliders & Parameter Controls ---
        left_frame = ctk.CTkFrame(tab_hedge)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left_frame, text="💵 Fluctuation Parameters", font=ctk.CTkFont(size=15, weight="bold")).pack(pady=10, padx=15, anchor="w")

        # CNY Slider
        ctk.CTkLabel(left_frame, text="CNY Exchange Rate Fluctuation (%):").pack(padx=15, pady=(5, 0), anchor="w")
        self.cny_fluc_lbl = ctk.CTkLabel(left_frame, text="0.0% (Baseline: 7.25 CNY)", text_color="grey")
        self.cny_fluc_lbl.pack(padx=15, pady=0, anchor="w")
        self.cny_slider = ctk.CTkSlider(left_frame, from_=-10.0, to=10.0, command=self.on_hedge_slider_change)
        self.cny_slider.pack(padx=15, pady=5, fill="x")
        self.cny_slider.set(0.0)

        # EUR Slider
        ctk.CTkLabel(left_frame, text="EUR Exchange Rate Fluctuation (%):").pack(padx=15, pady=(10, 0), anchor="w")
        self.eur_fluc_lbl = ctk.CTkLabel(left_frame, text="0.0% (Baseline: 0.92 EUR)", text_color="grey")
        self.eur_fluc_lbl.pack(padx=15, pady=0, anchor="w")
        self.eur_slider = ctk.CTkSlider(left_frame, from_=-10.0, to=10.0, command=self.on_hedge_slider_change)
        self.eur_slider.pack(padx=15, pady=5, fill="x")
        self.eur_slider.set(0.0)

        # Recommendation Card
        self.hedge_rec_card = ctk.CTkFrame(left_frame, fg_color=self.THEME["info_soft"], border_color=self.THEME["info_border"], border_width=1, corner_radius=8)
        self.hedge_rec_card.pack(pady=20, padx=15, fill="both", expand=True)
        self.hedge_rec_lbl = ctk.CTkLabel(self.hedge_rec_card, text="💵 Currency Risk Analysis\nAdjust exchange rate sliders to simulate landed cost exposures.", wraplength=220, justify="left")
        self.hedge_rec_lbl.pack(padx=15, pady=15, fill="both", expand=True)

        # --- RIGHT PANEL: Visual Chart Canvas ---
        self.hedge_display_frame = ctk.CTkFrame(tab_hedge, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        self.hedge_display_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        self.hedge_display_frame.grid_columnconfigure(0, weight=1)
        self.hedge_display_frame.grid_rowconfigure(0, weight=1)

        self.draw_hedge_chart()

    def on_hedge_slider_change(self, value):
        cny_val = self.cny_slider.get()
        eur_val = self.eur_slider.get()
        self.cny_fluc_lbl.configure(text=f"{cny_val:+.1f}% (Current: {7.25 * (1 + cny_val/100):.3f} CNY)")
        self.eur_fluc_lbl.configure(text=f"{eur_val:+.1f}% (Current: {0.92 * (1 + eur_val/100):.3f} EUR)")
        self.draw_hedge_chart()

    def draw_hedge_chart(self):
        for w in self.hedge_display_frame.winfo_children():
            w.destroy()

        if not self.extracted_data:
            self.make_empty_state(self.hedge_display_frame, "No quotes available", "Extract supplier quotes first to simulate currency exposure.")
            return

        cny_fluc = self.cny_slider.get() if hasattr(self, 'cny_slider') else 0.0
        eur_fluc = self.eur_slider.get() if hasattr(self, 'eur_slider') else 0.0

        # Select top 4 quotes with unit prices
        valid_quotes = []
        for q in self.extracted_data:
            try:
                if q["price"] is not None:
                    valid_quotes.append(q)
            except Exception:
                pass
        
        if not valid_quotes:
            self.make_empty_state(self.hedge_display_frame, "No numeric prices", "Add or correct unit prices before running the hedging simulator.")
            return

        # Sort and take top 4
        valid_quotes = sorted(valid_quotes, key=lambda x: float(x["price"] or 0))[:4]

        suppliers = []
        baseline_ddps = []
        simulated_ddps = []

        for q in valid_quotes:
            supplier_clean = self.clean_supplier_name(q.get("supplier"))
            product_clean = self.clean_product_name(q.get("product"))
            name = f"{supplier_clean}\n({product_clean})"
            
            unit_price_usd = float(q["price"])
            term = (q.get("term") or "").upper()
            
            # Simple DDP estimation
            baseline_ddp = unit_price_usd
            if "FOB" in term:
                baseline_ddp *= 1.15
            elif "EXW" in term:
                baseline_ddp *= 1.20
            else:
                baseline_ddp *= 1.05

            sim_ddp = baseline_ddp
            if "CNY" in term or any(tok in name.lower() for tok in ["anji", "xiantao", "juxian", "chen", "hefei", "guang"]):
                sim_ddp = baseline_ddp * (1 + cny_fluc / 100.0)
            elif "EUR" in term:
                sim_ddp = baseline_ddp * (1 + eur_fluc / 100.0)

            suppliers.append(name)
            baseline_ddps.append(baseline_ddp)
            simulated_ddps.append(sim_ddp)

        # Plot comparison bar chart
        fig, ax = plt.subplots(figsize=(6, 4), facecolor=self.THEME["surface"])
        ax.set_facecolor(self.THEME["surface"])

        import numpy as np
        x = np.arange(len(suppliers))
        width = 0.35

        ax.bar(x - width/2, baseline_ddps, width, label='Baseline DDP Landed Cost', color='#1f538d')
        ax.bar(x + width/2, simulated_ddps, width, label='Simulated Fluctuation', color='#a83232' if max(simulated_ddps) > max(baseline_ddps) else '#2da832')

        ax.set_ylabel('Cost per Piece (USD)', color=self.THEME["muted"])
        ax.set_title('Landed Cost Sensitivity to Currency Fluctuation', color=self.THEME["text"], pad=15)
        ax.set_xticks(x)
        ax.set_xticklabels(suppliers, color=self.THEME["muted"], rotation=15, fontsize=8)
        ax.tick_params(colors=self.THEME["muted"])
        ax.legend(facecolor=self.THEME["surface"], edgecolor='none', labelcolor=self.THEME["text"])
        ax.grid(color=self.THEME["border"], linestyle='--')

        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.hedge_display_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

        # Update recommendation message
        max_diff = max([sim - base for sim, base in zip(simulated_ddps, baseline_ddps)])
        if max_diff > 0.001:
            rec_text = f"🚨 CURRENCY EXPOSURE WARNING:\nStrengthening supplier currencies increase your landed costs by up to +{(max_diff/max(baseline_ddps))*100:.1f}%. Recommend purchasing forward contract hedges or setting a {(max_diff/max(baseline_ddps))*100 + 3:.1f}% price margin buffer."
        else:
            rec_text = "🟢 Favorable Exchange Trend:\nSimulated currency trends lower or stabilize your landed costs. Sourcing pricing is safe."
        self.hedge_rec_lbl.configure(text=rec_text)

    def setup_po_generator_tab(self):
        tab_po = self.logistics_tabview.tab("📄 PO Generator")
        tab_po.grid_columnconfigure(0, weight=1)
        tab_po.grid_columnconfigure(1, weight=1)
        tab_po.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_po, text="Enterprise Purchase Order (PO) Generator", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: PO Field Forms ---
        left_frame = ctk.CTkFrame(tab_po)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="📄 Purchase Order Parameters", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, pady=10, padx=15, sticky="w")

        # Supplier Choice
        ctk.CTkLabel(left_frame, text="Select Supplier:").grid(row=1, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_supplier_cb = ctk.CTkComboBox(left_frame, values=["Select Supplier"], command=self.on_po_supplier_selected)
        self.po_supplier_cb.grid(row=1, column=1, padx=15, pady=2, sticky="ew")

        # Product Choice
        ctk.CTkLabel(left_frame, text="Select Product:").grid(row=2, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_product_cb = ctk.CTkComboBox(left_frame, values=["Select Product"], command=self.on_po_quote_selected)
        self.po_product_cb.grid(row=2, column=1, padx=15, pady=2, sticky="ew")

        # PO Number
        ctk.CTkLabel(left_frame, text="PO Number:").grid(row=3, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_number_entry = ctk.CTkEntry(left_frame)
        self.po_number_entry.grid(row=3, column=1, padx=15, pady=2, sticky="ew")
        import random
        self.po_number_entry.insert(0, f"PO-2026-{random.randint(1000, 9999)}")
        self.po_number_entry.bind("<KeyRelease>", lambda e: self.update_po_preview())

        # Order Qty
        ctk.CTkLabel(left_frame, text="Order Quantity (pcs):").grid(row=4, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_qty_entry = ctk.CTkEntry(left_frame)
        self.po_qty_entry.grid(row=4, column=1, padx=15, pady=2, sticky="ew")
        self.po_qty_entry.insert(0, "10000")
        self.po_qty_entry.bind("<KeyRelease>", lambda e: self.update_po_preview())

        # Payment terms
        ctk.CTkLabel(left_frame, text="Payment Terms:").grid(row=5, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_payment_entry = ctk.CTkEntry(left_frame)
        self.po_payment_entry.grid(row=5, column=1, padx=15, pady=2, sticky="ew")
        self.po_payment_entry.insert(0, "30% Deposit, 70% before Shipment")
        self.po_payment_entry.bind("<KeyRelease>", lambda e: self.update_po_preview())

        # Shipping Address
        ctk.CTkLabel(left_frame, text="Delivery Address:").grid(row=6, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_address_entry = ctk.CTkEntry(left_frame)
        self.po_address_entry.grid(row=6, column=1, padx=15, pady=2, sticky="ew")
        self.po_address_entry.insert(0, "ProcureAI Hub warehouse, Dubai Airport Freezone, UAE")
        self.po_address_entry.bind("<KeyRelease>", lambda e: self.update_po_preview())

        # Unit Cost override display
        ctk.CTkLabel(left_frame, text="Unit Cost (USD):").grid(row=7, column=0, padx=15, pady=(5, 2), sticky="w")
        self.po_cost_lbl = ctk.CTkLabel(left_frame, text="$0.00", font=ctk.CTkFont(weight="bold"))
        self.po_cost_lbl.grid(row=7, column=1, padx=15, pady=2, sticky="w")

        # Generate & Export Buttons
        self.btn_gen_po = ctk.CTkButton(left_frame, text="📄 Generate PO PDF Document", command=self.export_po_pdf, fg_color="#1f7d44", hover_color="#15592e")
        self.btn_gen_po.grid(row=8, column=0, columnspan=2, padx=15, pady=20, sticky="ew")

        # --- RIGHT PANEL: Live PO Preview text ---
        right_frame = ctk.CTkFrame(tab_po)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📄 Live PO Document Draft", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, pady=10, padx=15, sticky="w")

        self.po_preview_box = ctk.CTkTextbox(right_frame, wrap="word")
        self.po_preview_box.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 8))

        self.po_readiness_box = ctk.CTkTextbox(right_frame, height=120, wrap="word")
        self.po_readiness_box.grid(row=2, column=0, sticky="ew", padx=15, pady=(0, 15))
        self.style_text_output(self.po_readiness_box)

        self.load_po_suppliers()

    def load_po_suppliers(self):
        suppliers = set()
        for r in self.extracted_data:
            if r.get("review_status") != "Approved":
                continue
            s = r.get("supplier")
            if s and s != "Unknown":
                suppliers.add(s)
        sorted_sups = sorted(list(suppliers))
        if hasattr(self, 'po_supplier_cb'):
            self.po_supplier_cb.configure(values=sorted_sups)
            if sorted_sups:
                self.po_supplier_cb.set(sorted_sups[0])
                self.on_po_supplier_selected(sorted_sups[0])
            else:
                self.po_supplier_cb.set("No approved quotes")
                self.po_product_cb.configure(values=["Approve quotes first"])
                self.po_product_cb.set("Approve quotes first")
                self.po_active_quote = None
                self.po_active_price = 0.0
                self.po_cost_lbl.configure(text="$0.00000")
                self.update_po_preview()

    def on_po_supplier_selected(self, choice):
        products = set()
        for r in self.extracted_data:
            if r.get("review_status") != "Approved":
                continue
            if r.get("supplier") == choice:
                p = (r.get("product") or "").strip().title()
                if p:
                    products.add(p)
        sorted_prods = sorted(list(products))
        if hasattr(self, 'po_product_cb'):
            self.po_product_cb.configure(values=sorted_prods)
            if sorted_prods:
                self.po_product_cb.set(sorted_prods[0])
                self.on_po_quote_selected(sorted_prods[0])

    def on_po_quote_selected(self, choice):
        supplier = self.po_supplier_cb.get()
        self.po_active_price = 0.0
        self.po_active_quote = None
        for r in self.extracted_data:
            if r.get("review_status") != "Approved":
                continue
            if r.get("supplier") == supplier and (r.get("product") or "").strip().title() == choice:
                self.po_active_quote = r
                try:
                    self.po_active_price = float(r["price"])
                except Exception:
                    self.po_active_price = 0.0
                break
        self.po_cost_lbl.configure(text=f"${self.po_active_price:.5f}")
        self.update_po_preview()

    def update_po_preview(self):
        supplier = self.po_supplier_cb.get()
        product = self.po_product_cb.get()
        po_num = self.po_number_entry.get().strip()
        address = self.po_address_entry.get().strip()
        payment = self.po_payment_entry.get().strip()
        
        try:
            qty = int(self.po_qty_entry.get().strip().replace(",", ""))
        except Exception:
            qty = 0
            
        unit_price = getattr(self, 'po_active_price', 0.0)
        total_cost = qty * unit_price

        preview_text = f"""==================================================
                 PURCHASE ORDER (DRAFT)
==================================================
PO NUMBER: {po_num}
DATE: 2026-08-04
BUYER: ProcureAI Hub Corp
DELIVERY ADDRESS: {address}

--------------------------------------------------
SUPPLIER: {supplier}
ITEM DESCRIPTION: {product}
ORDER QUANTITY: {qty:,} pcs
UNIT PRICE: ${unit_price:.5f}/pc
DELIVERY TERMS: {self.po_active_quote.get("term") if self.po_active_quote else "FOB"}
PAYMENT TERMS: {payment}

--------------------------------------------------
TOTAL ORDER VALUE: ${total_cost:,.2f}
==================================================
Authorized Signature: ___________________________
"""
        self.po_preview_box.configure(state="normal")
        self.po_preview_box.delete("1.0", tk.END)
        self.po_preview_box.insert("1.0", preview_text)
        self.po_preview_box.configure(state="disabled")
        self.update_po_readiness_panel()

    def export_po_pdf(self):
        supplier = self.po_supplier_cb.get()
        product = self.po_product_cb.get()
        po_num = self.po_number_entry.get().strip()
        address = self.po_address_entry.get().strip()
        payment = self.po_payment_entry.get().strip()
        issues, warnings = self.validate_po_ready()
        self.update_po_readiness_panel()
        if issues:
            note = "; ".join(issues)
            quote = getattr(self, "po_active_quote", None) or {}
            self.log_workflow_blocked_attempt("PO", "Issue PO Blocked", note, quote.get("id") or 0)
            messagebox.showerror(
                "PO Not Ready",
                "Fix these issues before issuing the PO:\n\n- " + "\n- ".join(issues)
            )
            return
        if warnings and not messagebox.askyesno(
            "PO Risk Warnings",
            "Review these commercial warnings before issuing:\n\n- "
            + "\n- ".join(warnings)
            + "\n\nContinue issuing this PO?"
        ):
            quote = getattr(self, "po_active_quote", None) or {}
            self.log_workflow_blocked_attempt("PO", "Issue PO Blocked", "; ".join(warnings), quote.get("id") or 0)
            return

        qty = self.parse_positive_int(self.po_qty_entry.get())
            
        unit_price = getattr(self, 'po_active_price', 0.0)
        total_cost = qty * unit_price

        file_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=f"PO_{po_num}.pdf")
        if not file_path:
            return

        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            styles = getSampleStyleSheet()

            primary_color = colors.HexColor("#1f538d")
            text_color = colors.HexColor("#333333")

            title_style = ParagraphStyle(
                'POTitle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=24,
                textColor=primary_color,
                spaceAfter=15
            )

            po_meta_style = ParagraphStyle(
                'POMeta',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=11,
                textColor=text_color,
                leading=14
            )

            table_header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=10,
                textColor=colors.white
            )

            table_cell_style = ParagraphStyle(
                'TableCell',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=10,
                textColor=text_color
            )

            story.append(Paragraph("PURCHASE ORDER", title_style))
            story.append(Spacer(1, 10))

            meta_data = [
                [
                    Paragraph(f"<b>Buyer:</b> ProcureAI Hub Corp<br/><b>Delivery Address:</b><br/>{address}", po_meta_style),
                    Paragraph(f"<b>PO Number:</b> {po_num}<br/><b>Date:</b> 2026-08-04<br/><b>Payment Terms:</b> {payment}", po_meta_style)
                ]
            ]
            meta_table = Table(meta_data, colWidths=[270, 270])
            meta_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('PADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 20))

            story.append(Paragraph(f"<b>Supplier:</b> {supplier}", po_meta_style))
            story.append(Spacer(1, 15))

            headers = [
                Paragraph("Item Description", table_header_style),
                Paragraph("Quantity", table_header_style),
                Paragraph("Unit Price", table_header_style),
                Paragraph("Total Cost", table_header_style)
            ]
            row1 = [
                Paragraph(product, table_cell_style),
                Paragraph(f"{qty:,} pcs", table_cell_style),
                Paragraph(f"${unit_price:.5f}", table_cell_style),
                Paragraph(f"${total_cost:,.2f}", table_cell_style)
            ]
            
            table_data = [headers, row1]
            items_table = Table(table_data, colWidths=[240, 100, 100, 100])
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), primary_color),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 40))

            sig_data = [
                [
                    Paragraph("<b>Prepared By:</b> ___________________________", po_meta_style),
                    Paragraph("<b>Authorized Signature:</b> ___________________________", po_meta_style)
                ]
            ]
            sig_table = Table(sig_data, colWidths=[270, 270])
            sig_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('PADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(sig_table)

            doc.build(story)
            self.save_po_record("Issued", file_path)
            messagebox.showinfo("Success", f"Purchase Order PDF saved successfully at:\n{file_path}!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PO PDF: {e}")

    def on_rfq_template_changed(self, choice):
        specs = ""
        if choice == "Medical PPE":
            specs = "- Product Type: Medical Face Mask / Disposable Hairnet\\n- Conformity: CE EN14683 Type IIR / FDA Approved\\n- Material Specs: 3-ply Non-woven, Meltblown Filter layer (BFE >= 98%)\\n- Packaging: 50pcs per inner box, sterile packed\\n- Weight/Thickness: 25g + 25g + 25g GSM\\n- Dimensions: Standard adult size (17.5cm x 9.5cm)"
        elif choice == "Consumer Electronics":
            specs = "- Input Rating: DC 5V/2A, Type-C Charging port\\n- Battery Capacity: 2000mAh rechargeable lithium-ion\\n- Certifications: CE, RoHS, FCC Conformity\\n- Shell Material: Fire-retardant ABS + PC plastic\\n- Dynamic Testing: Drop test from 1.2m height on concrete\\n- Thermal Limits: Operational range -10°C to 45°C"
        elif choice == "Apparel & Fabrics":
            specs = "- Composition: 95% Organic Cotton, 5% Spandex blend\\n- Weight/Density: 180 GSM single jersey\\n- Dyeing Grade: AZO-free eco-friendly reactive dyeing\\n- Shrinkage Limit: Less than 3% after 5 washes\\n- Stitching Detail: Double-needle flatlock hems\\n- Elasticity: High resilience double elastic banding"
        elif choice == "Industrial Packaging":
            specs = "- Material Grade: Double-wall corrugated Kraft paper (K=K)\\n- Bursting Strength: Min 14 kg/cm²\\n- Water Resistance: Cobra test absorption limit < 150g/m²\\n- Color: Natural Kraft Brown / White exterior\\n- Size/Fit: Standard outer shipping cartons (50x40x40 cm)\\n- Load Limits: Rated for safe transit up to 25 kg weight load"
        
        if specs:
            self.rfq_specs_text.delete("1.0", tk.END)
            self.rfq_specs_text.insert("1.0", specs.replace("\\n", "\n"))
        self.update_rfq_readiness_panel()

    def toggle_navigation_sidebar(self):
        if self.sidebar_visible:
            self.sidebar_frame.grid_forget()
            self.grid_columnconfigure(0, minsize=0, weight=0)
            self.btn_toggle_nav.configure(fg_color="#3c3c3c", text="☰ Show Navigation")
            self.btn_toggle_nav.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], text="Show Navigation")
            self.sidebar_visible = False
        else:
            self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
            self.grid_columnconfigure(0, minsize=240, weight=0)
            self.btn_toggle_nav.configure(fg_color="#1f538d", text="☰ Hide Navigation")
            self.btn_toggle_nav.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], text="Hide Navigation")
            self.sidebar_visible = True

    def toggle_sourcing_files(self):
        if not hasattr(self, 'tab_comp_ref') or not self.tab_comp_ref:
            return
        if self.sourcing_files_visible:
            self.sourcing_files_subframe.grid_forget()
            self.tab_comp_ref.grid_columnconfigure(0, minsize=0, weight=0)
            self.btn_toggle_files.configure(fg_color="#3c3c3c", text="📁 Show Source Files")
            self.btn_toggle_files.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], text="Show Source Files")
            self.sourcing_files_visible = False
        else:
            self.sourcing_files_subframe.grid(row=0, column=0, sticky="nsew", padx=(5, 10), pady=10)
            self.tab_comp_ref.grid_columnconfigure(0, minsize=260, weight=0)
            self.btn_toggle_files.configure(fg_color="#1f538d", text="📁 Hide Source Files")
            self.btn_toggle_files.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], text="Hide Source Files")
            self.sourcing_files_visible = True

    def toggle_document_preview(self):
        if self.document_preview_visible:
            self.preview_frame.grid_forget()
            self.grid_columnconfigure(2, minsize=0, weight=0)
            self.btn_toggle_preview.configure(fg_color="#3c3c3c", text="📄 Show Preview")
            self.btn_toggle_preview.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], text="Show Preview")
            self.document_preview_visible = False
        else:
            self.preview_frame.grid(row=0, column=2, sticky="nsew", padx=(0, 14), pady=14)
            self.grid_columnconfigure(2, minsize=360, weight=0)
            self.btn_toggle_preview.configure(fg_color="#1f538d", text="📄 Hide Preview")
            self.btn_toggle_preview.configure(fg_color=self.THEME["surface_alt"], hover_color="#E2E8F0", text_color=self.THEME["text"], text="Hide Preview")
            self.document_preview_visible = True

    def setup_po_register_tab(self):
        tab = self.logistics_tabview.tab("PO Register")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_columnconfigure(1, weight=0, minsize=360)
        tab.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(tab, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 8))
        header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header, text="Purchase Order Register", font=ctk.CTkFont(size=20, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(header, text="Saved purchase orders with supplier, approved quote, value, status, and generated PDF path.", font=ctk.CTkFont(size=12), text_color=self.THEME["muted"]).grid(row=1, column=0, sticky="w", pady=(3, 0))
        self.make_button(header, "Refresh", command=self.load_po_register, variant="secondary", width=90).grid(row=0, column=1, rowspan=2, padx=5)
        self.make_button(header, "Accepted", command=lambda: self.update_selected_po_status("Supplier Accepted"), variant="primary", width=95).grid(row=0, column=2, rowspan=2, padx=5)
        self.make_button(header, "Shipped", command=lambda: self.update_selected_po_status("Shipped"), variant="warning", width=80).grid(row=0, column=3, rowspan=2, padx=5)
        self.make_button(header, "Close", command=lambda: self.update_selected_po_status("Closed"), variant="success", width=80).grid(row=0, column=4, rowspan=2, padx=5)
        self.make_button(header, "Cancel", command=lambda: self.update_selected_po_status("Cancelled"), variant="danger", width=80).grid(row=0, column=5, rowspan=2, padx=5)

        frame = ctk.CTkFrame(tab, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        frame.grid(row=1, column=0, sticky="nsew", padx=(16, 8), pady=(0, 16))
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        cols = ("id", "po", "supplier", "product", "qty", "unit", "total", "status", "pdf")
        self.po_register_tree = ttk.Treeview(frame, columns=cols, show="headings", style="Treeview")
        for col, label, width in [
            ("id", "ID", 50),
            ("po", "PO No.", 140),
            ("supplier", "Supplier", 150),
            ("product", "Product", 130),
            ("qty", "Qty", 80),
            ("unit", "Unit", 75),
            ("total", "Total", 90),
            ("status", "Status", 120),
            ("pdf", "PDF Path", 240),
        ]:
            self.po_register_tree.heading(col, text=label)
            self.po_register_tree.column(col, width=width, anchor="w")
        self.po_register_tree.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        self.po_register_tree.bind("<<TreeviewSelect>>", lambda event: self.show_po_register_detail())

        detail = ctk.CTkFrame(tab, fg_color=self.THEME["surface"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        detail.grid(row=1, column=1, sticky="nsew", padx=(8, 16), pady=(0, 16))
        detail.grid_columnconfigure(0, weight=1)
        detail.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(detail, text="PO Detail", font=ctk.CTkFont(size=15, weight="bold"), text_color=self.THEME["text"]).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 8))
        self.po_detail_box = ctk.CTkTextbox(detail, wrap="word", fg_color=self.THEME["surface_alt"], text_color=self.THEME["text"], border_width=0)
        self.po_detail_box.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 10))
        self.po_detail_box.insert("1.0", "Select a PO row to view supplier, value, document path, and workflow audit trail.")
        self.po_detail_box.configure(state="disabled")

        po_actions = ctk.CTkFrame(detail, fg_color="transparent")
        po_actions.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 14))
        po_actions.grid_columnconfigure(0, weight=1)
        po_actions.grid_columnconfigure(1, weight=1)
        self.make_button(po_actions, "Open PDF", command=lambda: self.open_selected_register_document("PO"), variant="primary").grid(row=0, column=0, sticky="ew", padx=(0, 4), pady=4)
        self.make_button(po_actions, "Open Folder", command=lambda: self.open_selected_register_folder("PO"), variant="secondary").grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=4)
        self.make_button(po_actions, "Copy Path", command=lambda: self.copy_selected_register_path("PO"), variant="secondary").grid(row=1, column=0, sticky="ew", padx=(0, 4), pady=4)
        self.make_button(po_actions, "Load Generator", command=self.load_selected_po_into_generator, variant="success").grid(row=1, column=1, sticky="ew", padx=(4, 0), pady=4)
        self.load_po_register()

    def load_po_register(self):
        if not hasattr(self, "po_register_tree"):
            return
        self.po_register_tree.delete(*self.po_register_tree.get_children())
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("""
            SELECT id, po_number, supplier_name, product_name, quantity, unit_cost, total_value, status, COALESCE(pdf_path, '')
            FROM po_register
            ORDER BY updated_at DESC, id DESC
        """)
        for row in c.fetchall():
            row = list(row)
            row[5] = f"${float(row[5]):.5f}" if row[5] is not None else "N/A"
            row[6] = f"${float(row[6]):,.2f}" if row[6] is not None else "N/A"
            self.po_register_tree.insert("", tk.END, values=row)
        conn.close()
        if hasattr(self, "po_detail_box"):
            self.set_detail_text(self.po_detail_box, "Select a PO row to view supplier, value, document path, and workflow audit trail.")

    def update_selected_po_status(self, status):
        sel = self.po_register_tree.selection()
        if not sel:
            messagebox.showwarning("Select PO", "Please select a PO record first.")
            return
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        for item in sel:
            po_id = int(self.po_register_tree.item(item, "values")[0])
            c.execute("UPDATE po_register SET status=?, updated_at=datetime('now') WHERE id=?", (status, po_id))
            c.execute("""
                INSERT INTO workflow_audit_log (workflow_type, workflow_id, action, status, note)
                VALUES ('PO', ?, 'Status Changed', ?, ?)
            """, (po_id, status, f"PO marked {status}."))
        conn.commit()
        conn.close()
        self.load_po_register()
        if hasattr(self, "dashboard_cards"):
            self.update_dashboard_page()

    def setup_price_history_tab(self):
        tab_history = self.sourcing_tabview.tab("📈 Price History")
        tab_history.grid_columnconfigure(0, weight=1)
        tab_history.grid_columnconfigure(1, weight=2)
        tab_history.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_history, text="Supplier Historical Price Trend Tracker", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Log Price Point Form ---
        left_frame = ctk.CTkFrame(tab_history)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left_frame, text="📈 Log Historical Price", font=ctk.CTkFont(size=15, weight="bold")).pack(pady=10, padx=15, anchor="w")

        ctk.CTkLabel(left_frame, text="Select Supplier:").pack(padx=15, pady=(5, 2), anchor="w")
        self.hist_supplier_cb = ctk.CTkComboBox(left_frame, values=["Select Supplier"], command=self.on_history_supplier_selected, width=230)
        self.hist_supplier_cb.pack(padx=15, pady=2, anchor="w")

        ctk.CTkLabel(left_frame, text="Select Product:").pack(padx=15, pady=(5, 2), anchor="w")
        self.hist_product_cb = ctk.CTkComboBox(left_frame, values=["Select Product"], command=lambda choice: self.draw_price_history_chart(), width=230)
        self.hist_product_cb.pack(padx=15, pady=2, anchor="w")

        ctk.CTkLabel(left_frame, text="Historical Unit Price (USD):").pack(padx=15, pady=(5, 2), anchor="w")
        self.hist_price_entry = ctk.CTkEntry(left_frame, placeholder_text="e.g. 0.0520", width=230)
        self.hist_price_entry.pack(padx=15, pady=2, anchor="w")

        ctk.CTkLabel(left_frame, text="Log Date (YYYY-MM-DD):").pack(padx=15, pady=(5, 2), anchor="w")
        self.hist_date_entry = ctk.CTkEntry(left_frame, placeholder_text="e.g. 2026-03-15", width=230)
        self.hist_date_entry.pack(padx=15, pady=2, anchor="w")
        import datetime
        self.hist_date_entry.insert(0, datetime.date.today().strftime("%Y-%m-%d"))

        btn_add_pt = ctk.CTkButton(left_frame, text="📈 Log Price Point", fg_color="#1f538d", hover_color="#153e6b", command=self.add_historical_price)
        btn_add_pt.pack(padx=15, pady=20, fill="x")

        # --- RIGHT PANEL: Visual Line Chart Canvas ---
        self.hist_chart_frame = ctk.CTkFrame(tab_history, fg_color=self.THEME["surface_soft"], border_color=self.THEME["border"], border_width=1, corner_radius=8)
        self.hist_chart_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        self.hist_chart_frame.grid_columnconfigure(0, weight=1)
        self.hist_chart_frame.grid_rowconfigure(0, weight=1)

        self.load_price_history_dropdowns()

    def load_price_history_dropdowns(self):
        suppliers = set()
        for r in self.extracted_data:
            s = r.get("supplier")
            if s and s != "Unknown":
                suppliers.add(s)
        sorted_sups = sorted(list(suppliers))
        if hasattr(self, 'hist_supplier_cb'):
            self.hist_supplier_cb.configure(values=sorted_sups)
            if sorted_sups:
                self.hist_supplier_cb.set(sorted_sups[0])
                self.on_history_supplier_selected(sorted_sups[0])

    def on_history_supplier_selected(self, choice):
        products = set()
        for r in self.extracted_data:
            if r.get("supplier") == choice:
                p = (r.get("product") or "").strip().title()
                if p:
                    products.add(p)
        sorted_prods = sorted(list(products))
        if hasattr(self, 'hist_product_cb'):
            self.hist_product_cb.configure(values=sorted_prods)
            if sorted_prods:
                self.hist_product_cb.set(sorted_prods[0])
                self.draw_price_history_chart()

    def add_historical_price(self):
        supplier = self.hist_supplier_cb.get()
        product = self.hist_product_cb.get()
        date_val = self.hist_date_entry.get().strip()
        try:
            price_val = float(self.hist_price_entry.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid numeric historical price!")
            return

        if not supplier or supplier == "Select Supplier" or not product or product == "Select Product":
            messagebox.showwarning("Warning", "Please select a supplier and product first!")
            return
        
        import re
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_val):
            messagebox.showerror("Error", "Date must be in YYYY-MM-DD format!")
            return

        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                INSERT INTO price_history (supplier, product, price, log_date)
                VALUES (?, ?, ?, ?)
            """, (supplier, product, price_val, date_val))
            conn.commit()
            conn.close()

            self.hist_price_entry.delete(0, tk.END)
            messagebox.showinfo("Success", f"Logged historical price point of ${price_val:.4f} on {date_val}!")
            self.draw_price_history_chart()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log price point: {e}")

    def draw_price_history_chart(self):
        for w in self.hist_chart_frame.winfo_children():
            w.destroy()

        supplier = self.hist_supplier_cb.get() if hasattr(self, 'hist_supplier_cb') else ""
        product = self.hist_product_cb.get() if hasattr(self, 'hist_product_cb') else ""

        if not supplier or supplier == "Select Supplier" or not product or product == "Select Product":
            self.make_empty_state(self.hist_chart_frame, "Select a supplier and product", "Choose both fields to view the historical price trend.")
            return

        current_price = None
        for r in self.extracted_data:
            if r.get("supplier") == supplier and (r.get("product") or "").strip().title() == product:
                try:
                    current_price = float(r["price"])
                except Exception:
                    pass
                break

        history_points = []
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("""
                SELECT log_date, price FROM price_history 
                WHERE supplier = ? AND product = ? 
                ORDER BY log_date ASC
            """, (supplier, product))
            history_points = c.fetchall()
            conn.close()
        except Exception as e:
            print("Failed to query historical price points:", e)

        dates = []
        prices = []
        for d, p in history_points:
            dates.append(d)
            prices.append(p)

        import datetime
        if current_price is not None:
            today_str = datetime.date.today().strftime("%Y-%m-%d")
            if today_str not in dates:
                dates.append(today_str)
                prices.append(current_price)

        if len(prices) < 2:
            if len(prices) == 1:
                dates.insert(0, "2026-05-01")
                prices.insert(0, prices[0])
            else:
                self.make_empty_state(self.hist_chart_frame, "No price points yet", "Log the first historical price point to begin tracking this supplier trend.")
                return

        fig, ax = plt.subplots(figsize=(6, 4), facecolor=self.THEME["surface"])
        ax.set_facecolor(self.THEME["surface"])

        ax.plot(dates, prices, marker='o', linestyle='-', color='#1f538d', linewidth=2.5, markersize=8, label="Unit Cost (USD)")
        
        pct_change = ((prices[-1] - prices[0]) / prices[0]) * 100
        trend_color = "#2da832" if pct_change <= 0 else "#a83232"
        ax.plot([dates[0], dates[-1]], [prices[0], prices[-1]], linestyle='--', color=trend_color, alpha=0.5)

        ax.set_ylabel('Unit Cost (USD)', color=self.THEME["muted"])
        ax.set_title(f'Quote Price Trend: {supplier} ({product})', color=self.THEME["text"], pad=15)
        ax.tick_params(colors=self.THEME["muted"])
        ax.grid(color=self.THEME["border"], linestyle='--')
        
        plt.xticks(rotation=15, color=self.THEME["muted"])
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.hist_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    def export_scar_pdf(self):
        selected = self.inc_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an incident log from the table first!")
            return

        item = self.inc_tree.item(selected[0])
        val = item["values"]
        
        log_id = val[0]
        supplier = val[1]
        inc_type = val[2]
        desc = val[3]
        severity = val[4]
        date_logged = val[5]

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf", 
            filetypes=[("PDF files", "*.pdf")], 
            initialfile=f"SCAR_{supplier.replace(' ', '_')}_{log_id}.pdf"
        )
        if not file_path:
            return

        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
            story = []
            styles = getSampleStyleSheet()

            primary_color = colors.HexColor("#bf3b3b")
            text_color = colors.HexColor("#333333")

            title_style = ParagraphStyle(
                'SCARTitle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=22,
                textColor=primary_color,
                spaceAfter=15
            )

            meta_style = ParagraphStyle(
                'SCARMeta',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=11,
                textColor=text_color,
                leading=14
            )

            box_title_style = ParagraphStyle(
                'SCARBoxTitle',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=11,
                textColor=colors.white
            )

            story.append(Paragraph("SUPPLIER CORRECTIVE ACTION REQUEST (SCAR)", title_style))
            story.append(Spacer(1, 10))

            meta_data = [
                [
                    Paragraph(f"<b>SCAR Reference ID:</b> SCAR-{log_id}<br/><b>Supplier Name:</b> {supplier}<br/><b>Date Logged:</b> {date_logged}", meta_style),
                    Paragraph(f"<b>Incident Type:</b> {inc_type}<br/><b>Severity Rating:</b> {severity}<br/><b>Status:</b> PENDING ROOT CAUSE", meta_style)
                ]
            ]
            meta_table = Table(meta_data, colWidths=[270, 270])
            meta_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('PADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(meta_table)
            story.append(Spacer(1, 20))

            story.append(Paragraph("<b>1. Description of Non-Conformance / Quality Incident:</b>", meta_style))
            desc_table = Table([[Paragraph(desc, meta_style)]], colWidths=[540])
            desc_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#f9f9f9")),
                ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#dcdcdc")),
                ('PADDING', (0,0), (-1,-1), 12),
            ]))
            story.append(desc_table)
            story.append(Spacer(1, 20))

            sections = [
                ("2. Containment Action & Immediate Correction Plan (Required in 24 Hours)", "Describe actions taken to isolate defective stock, notify freight forwarders, or suspend production line."),
                ("3. Root Cause Investigation & 5-Why Analysis (Required in 7 Business Days)", "Identify the exact systemic or operational failure (e.g. raw material, machinery tolerance, testing failure)."),
                ("4. Preventive & Corrective Actions Plan (Required in 14 Business Days)", "Define permanent steps to prevent recurrence (e.g. updating standard operating procedures, operator training, calibration schedules).")
            ]

            for s_title, s_placeholder in sections:
                header_t = Table([[Paragraph(s_title, box_title_style)]], colWidths=[540])
                header_t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#333333")),
                    ('PADDING', (0,0), (-1,-1), 6),
                ]))
                story.append(header_t)

                body_t = Table([[Paragraph(f"<font color='grey'><i>{s_placeholder}</i></font><br/><br/><br/><br/><br/>", meta_style)]], colWidths=[540])
                body_t.setStyle(TableStyle([
                    ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#333333")),
                    ('PADDING', (0,0), (-1,-1), 10),
                ]))
                story.append(body_t)
                story.append(Spacer(1, 15))

            story.append(Spacer(1, 20))
            sig_data = [
                [
                    Paragraph("<b>ProcureAI Quality Auditor Signature:</b><br/>___________________________", meta_style),
                    Paragraph("<b>Supplier Representative Signature:</b><br/>___________________________", meta_style)
                ]
            ]
            sig_table = Table(sig_data, colWidths=[270, 270])
            sig_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('PADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(sig_table)

            doc.build(story)
            messagebox.showinfo("Success", f"SCAR Corrective Action PDF generated successfully at:\n{file_path}!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate SCAR PDF: {e}")

    def setup_ai_negotiation_tab(self):
        tab_neg = self.rfqs_tabview.tab("💬 AI Negotiation")
        tab_neg.grid_columnconfigure(0, weight=1)
        tab_neg.grid_columnconfigure(1, weight=1)
        tab_neg.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_neg, text="AI Sourcing Target-Price Negotiation Agent", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Parameters configuration ---
        left_frame = ctk.CTkFrame(tab_neg)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="💬 Sourcing Targets", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, pady=10, padx=15, sticky="w")

        # Supplier Choice
        ctk.CTkLabel(left_frame, text="Select Supplier:").grid(row=1, column=0, padx=15, pady=(5, 2), sticky="w")
        self.neg_supplier_cb = ctk.CTkComboBox(left_frame, values=["Select Supplier"], command=self.on_neg_supplier_selected)
        self.neg_supplier_cb.grid(row=1, column=1, padx=15, pady=2, sticky="ew")

        # Product Choice
        ctk.CTkLabel(left_frame, text="Select Product:").grid(row=2, column=0, padx=15, pady=(5, 2), sticky="w")
        self.neg_product_cb = ctk.CTkComboBox(left_frame, values=["Select Product"])
        self.neg_product_cb.grid(row=2, column=1, padx=15, pady=2, sticky="ew")

        # Target Discount
        ctk.CTkLabel(left_frame, text="Target Discount (%):").grid(row=3, column=0, padx=15, pady=(5, 2), sticky="w")
        self.neg_discount_entry = ctk.CTkEntry(left_frame)
        self.neg_discount_entry.grid(row=3, column=1, padx=15, pady=2, sticky="ew")
        self.neg_discount_entry.insert(0, "15")

        # Target Payment terms
        ctk.CTkLabel(left_frame, text="Propose Payment Term:").grid(row=4, column=0, padx=15, pady=(5, 2), sticky="w")
        self.neg_payment_entry = ctk.CTkEntry(left_frame)
        self.neg_payment_entry.grid(row=4, column=1, padx=15, pady=2, sticky="ew")
        self.neg_payment_entry.insert(0, "100% Letter of Credit (L/C) at sight / Net 30")

        # Action Buttons
        self.btn_neg_draft = ctk.CTkButton(left_frame, text="💬 Draft AI Sourcing Email", command=self.run_ai_negotiation_draft, fg_color="#1f538d", hover_color="#153e6b")
        self.btn_neg_draft.grid(row=5, column=0, columnspan=2, padx=15, pady=25, sticky="ew")

        # --- RIGHT PANEL: Live Draft Output ---
        right_frame = ctk.CTkFrame(tab_neg)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📄 Persuasive Email Draft", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, pady=10, padx=15, sticky="w")

        self.neg_preview_box = ctk.CTkTextbox(right_frame, wrap="word")
        self.neg_preview_box.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))

        # Bottom Actions row
        ctrl_fr = ctk.CTkFrame(right_frame, fg_color="transparent")
        ctrl_fr.grid(row=2, column=0, padx=15, pady=(0, 15), sticky="ew")

        btn_copy = ctk.CTkButton(ctrl_fr, text="📋 Copy Draft", command=lambda: self.copy_to_clipboard(self.neg_preview_box.get("1.0", "end-1c").strip()))
        btn_copy.pack(side="left", fill="x", expand=True, padx=(0, 5))

        btn_mail = ctk.CTkButton(ctrl_fr, text="✉ Open in Outlook", fg_color="#1f7d44", hover_color="#15592e", command=self.launch_neg_email_client)
        btn_mail.pack(side="right", fill="x", expand=True, padx=(5, 0))

        self.load_negotiation_dropdowns()

    def load_negotiation_dropdowns(self):
        suppliers = set()
        for r in self.extracted_data:
            s = r.get("supplier")
            if s and s != "Unknown":
                suppliers.add(s)
        sorted_sups = sorted(list(suppliers))
        if hasattr(self, 'neg_supplier_cb'):
            self.neg_supplier_cb.configure(values=sorted_sups)
            if sorted_sups:
                self.neg_supplier_cb.set(sorted_sups[0])
                self.on_neg_supplier_selected(sorted_sups[0])

    def on_neg_supplier_selected(self, choice):
        products = set()
        for r in self.extracted_data:
            if r.get("supplier") == choice:
                p = (r.get("product") or "").strip().title()
                if p:
                    products.add(p)
        sorted_prods = sorted(list(products))
        if hasattr(self, 'neg_product_cb'):
            self.neg_product_cb.configure(values=sorted_prods)
            if sorted_prods:
                self.neg_product_cb.set(sorted_prods[0])

    def run_ai_negotiation_draft(self):
        supplier = self.neg_supplier_cb.get()
        product = self.neg_product_cb.get()
        discount = self.neg_discount_entry.get().strip()
        payment = self.neg_payment_entry.get().strip()

        if not supplier or supplier == "Select Supplier" or not product or product == "Select Product":
            messagebox.showwarning("Warning", "Please select a supplier and product first!")
            return

        active_quote = None
        other_quotes = []
        for r in self.extracted_data:
            r_prod = (r.get("product") or "").strip().lower()
            if r_prod == product.lower():
                if r.get("supplier") == supplier:
                    active_quote = r
                else:
                    other_quotes.append(r)

        current_price = float(active_quote["price"]) if active_quote and active_quote.get("price") else 0.0
        currency = active_quote.get("unit") or "USD"
        moq = active_quote.get("moq") or "N/A"
        lead_time = active_quote.get("lead_time") or "N/A"

        competitor_prices = []
        for o in other_quotes:
            try:
                competitor_prices.append(float(o["price"]))
            except Exception:
                pass
        best_rival_price = min(competitor_prices) if competitor_prices else None

        self.btn_neg_draft.configure(state="disabled", text="AI Sourcing Agent drafting...")
        self.update()

        def draft_thread():
            competitor_clause = ""
            if best_rival_price is not None and best_rival_price < current_price:
                competitor_clause = f"Please note that we have received competing quotations for this item as low as ${best_rival_price:.4f}/{currency}."

            prompt = f"""
            Draft a highly professional, persuasive supplier target-price negotiation email to:
            - Supplier Name: {supplier}
            - Product Item: {product}
            - Current Quoted Price: {current_price:.4f} {currency}
            - Target Discount requested: {discount}% (Target price: {current_price * (1 - float(discount)/100.0 if discount.isdigit() else 0.85):.4f})
            - Target Payment Terms: {payment}
            - Quoted MOQ: {moq}
            - Quoted Lead Time: {lead_time}
            
            Persuasion strategy:
            - Be extremely polite, respectful, and emphasize a long-term business relationship.
            - Leverage: {competitor_clause} If competitors offer better pricing, suggest that matching or narrowing the gap will secure the volume order.
            - Offer to increase volume commitments or guarantee repeat quarterly orders if target price is matched.
            - Request confirmation of lead time and technical sample availability.
            
            Format the output strictly as a ready-to-send corporate email with Subject, Salutation, Body, and Sign-off block. Do not include markdown syntax.
            """

            try:
                result_text = self.generate_with_fallback([], prompt, json_response=False)
                self.after(0, lambda: display_draft(result_text))
            except Exception as e:
                self.after(0, lambda err=e: display_draft(f"Failed to draft negotiation email:\n{err}"))
            finally:
                self.after(0, lambda: self.btn_neg_draft.configure(state="normal", text="💬 Draft AI Sourcing Email"))

        def display_draft(text):
            self.neg_preview_box.configure(state="normal")
            self.neg_preview_box.delete("1.0", tk.END)
            self.neg_preview_box.insert("1.0", text)
            self.neg_preview_box.configure(state="disabled")

        threading.Thread(target=draft_thread, daemon=True).start()

    def launch_neg_email_client(self):
        body = self.neg_preview_box.get("1.0", "end-1c").strip()
        if not body or "persuasive" in body.lower() or "draft" in body.lower():
            messagebox.showwarning("Warning", "Please draft the email first!")
            return
        
        supplier = self.neg_supplier_cb.get()
        contact_email = "sales@supplier.com"
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT contact_info FROM supplier_contacts WHERE supplier = ?", (supplier,))
            row = c.fetchone()
            if row and "@" in row[0]:
                import re
                emails = re.findall(r'[\w\.-]+@[\w\.-]+', row[0])
                if emails:
                    contact_email = emails[0]
            conn.close()
        except Exception:
            pass

        import urllib.parse
        import webbrowser
        subject = f"Supplier Price Negotiation: {self.neg_product_cb.get()}"
        mail_url = f"mailto:{contact_email}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        webbrowser.open(mail_url)

    def setup_global_barriers_tab(self):
        tab_barriers = self.search_tabview.tab("🌍 Global Trade Barriers")
        tab_barriers.grid_columnconfigure(0, weight=1)
        tab_barriers.grid_columnconfigure(1, weight=1)
        tab_barriers.grid_rowconfigure(1, weight=1)

        # Header Title
        title_lbl = ctk.CTkLabel(tab_barriers, text="Global Sourcing Regulatory Barriers & Customs warnings", font=ctk.CTkFont(size=20, weight="bold"))
        title_lbl.grid(row=0, column=0, columnspan=2, padx=20, pady=(15, 10), sticky="w")

        # --- LEFT PANEL: Configuration Inputs ---
        left_frame = ctk.CTkFrame(tab_barriers)
        left_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(left_frame, text="🌍 Import Compliance parameters", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, columnspan=2, padx=15, pady=10, sticky="w")

        # Product Dropdown Category
        ctk.CTkLabel(left_frame, text="Product Category:").grid(row=1, column=0, padx=15, pady=5, sticky="w")
        self.barrier_category_cb = ctk.CTkComboBox(left_frame, values=["Custom"])
        self.barrier_category_cb.grid(row=1, column=1, padx=15, pady=5, sticky="ew")

        # Custom Product Input Name
        ctk.CTkLabel(left_frame, text="Or Custom Product:").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        self.barrier_custom_name = ctk.CTkEntry(left_frame, placeholder_text="e.g. PPE Masks / Laser Cutter")
        self.barrier_custom_name.grid(row=2, column=1, padx=15, pady=5, sticky="ew")

        # Destination country
        ctk.CTkLabel(left_frame, text="Destination Country:").grid(row=3, column=0, padx=15, pady=5, sticky="w")
        self.barrier_dest_cb = ctk.CTkComboBox(left_frame, values=["United Arab Emirates", "Saudi Arabia", "Qatar", "Oman", "Kuwait", "European Union"])
        self.barrier_dest_cb.grid(row=3, column=1, padx=15, pady=5, sticky="ew")
        self.barrier_dest_cb.set("Saudi Arabia")

        # HS Code
        ctk.CTkLabel(left_frame, text="HS Code (Optional):").grid(row=4, column=0, padx=15, pady=5, sticky="w")
        self.barrier_hs_entry = ctk.CTkEntry(left_frame, placeholder_text="6-digit code, e.g. 630790")
        self.barrier_hs_entry.grid(row=4, column=1, padx=15, pady=5, sticky="ew")

        self.btn_run_barriers = ctk.CTkButton(left_frame, text="🔍 Check Import Restrictions", fg_color="#1f538d", hover_color="#153e6b", command=self.run_global_trade_restrictions_scan)
        self.btn_run_barriers.grid(row=5, column=0, columnspan=2, padx=15, pady=25, sticky="ew")

        # --- RIGHT PANEL: Compliance report output ---
        right_frame = ctk.CTkFrame(tab_barriers)
        right_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(right_frame, text="📋 AI Trade Restrictions Compliance Report", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, pady=10, padx=15, sticky="w")

        self.barrier_output_box = ctk.CTkTextbox(right_frame, wrap="word")
        self.barrier_output_box.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))

        self.load_barrier_categories()

    def load_barrier_categories(self):
        products = set()
        for r in self.extracted_data:
            prod = (r.get("product") or "").strip().title()
            if prod:
                products.add(prod)
        sorted_prods = ["Custom"] + sorted(list(products))
        if hasattr(self, 'barrier_category_cb'):
            self.barrier_category_cb.configure(values=sorted_prods)
            self.barrier_category_cb.set(sorted_prods[0])

    def run_global_trade_restrictions_scan(self):
        choice = self.barrier_category_cb.get()
        custom = self.barrier_custom_name.get().strip()
        dest = self.barrier_dest_cb.get()
        hs_code = self.barrier_hs_entry.get().strip()

        prod_name = custom if choice == "Custom" else choice
        if not prod_name:
            messagebox.showwarning("Warning", "Please select a category or enter a custom product name!")
            return

        self.btn_run_barriers.configure(state="disabled", text="AI compliance scan in progress...")
        self.update()

        def scan_thread():
            ai_prompt = f"""
            Perform a trade compliance audit and import restrictions scan for:
            - Product: {prod_name}
            - Destination Country: {dest}
            - HS Code: {hs_code if hs_code else "Unknown / Auto-detect"}
            
            Determine:
            1. Prohibited & Restricted status (Is the item banned, requires special import permits, or restricted by ministries?).
            2. Anti-Dumping & Countervailing Duties (Check if there are protective tariff penalties on origin countries, specifically from China to this destination).
            3. Regulatory Conformity Requirements (e.g. SABER / SASO certificate of conformity for Saudi Arabia, G-Mark for GCC toy/low-voltage electricals, CE / REACH for EU, MoH approvals for medical devices).
            4. Sourcing Risk Warnings (e.g. customs seizure risks, delay points in documentation, required certificates of origin or legalization steps).
            
            Write the output in a clean, highly structured compliance report.
            Do not use markdown syntax. Make it look like a formal government trade advisory.
            """

            try:
                result_text = self.generate_with_fallback([], ai_prompt, json_response=False)
                
                header_report = f"""==================================================
🌍 GLOBAL IMPORT REGULATORY COMPLIANCE REPORT
==================================================
TARGET PRODUCT : {prod_name}
DESTINATION    : {dest}
HS CODE        : {hs_code if hs_code else "AUTO-DETECTED"}
--------------------------------------------------
AUDIT & REGULATORY WARNING DETAILS:
{result_text}
"""
                self.after(0, lambda: display_report(header_report))
            except Exception as e:
                self.after(0, lambda err=e: display_report(f"Failed to scan global barriers:\n{err}"))
            finally:
                self.after(0, lambda: self.btn_run_barriers.configure(state="normal", text="🔍 Check Import Restrictions"))

        def display_report(text):
            self.barrier_output_box.configure(state="normal")
            self.barrier_output_box.delete("1.0", tk.END)
            self.barrier_output_box.insert("1.0", text)
            self.barrier_output_box.configure(state="disabled")

        threading.Thread(target=scan_thread, daemon=True).start()

if __name__ == "__main__":
    app = App()
    try:
        app.mainloop()
    finally:
        try:
            app.destroy()
        except Exception:
            pass
